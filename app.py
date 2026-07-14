# v19
import os
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.middleware.proxy_fix import ProxyFix
import cloudinary
import cloudinary.uploader
import cloudinary.api
import cloudinary.utils

# Cloudinary config — credentials set via Railway environment variables
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME', ''),
    api_key=os.environ.get('CLOUDINARY_API_KEY', ''),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET', ''),
    secure=True
)

def upload_to_cloudinary(file, folder='tahfeel-documents', public=False):
    """Upload file to Cloudinary. Returns (url, public_id) or (None, None) on failure.
    KYC/company documents are uploaded as `authenticated` (only reachable via signed
    URLs behind the login-protected /documents/... routes). Pass public=True ONLY for
    assets that external services must fetch directly (e.g. WhatsApp media for Meta)."""
    try:
        if not file or not file.filename: return None, None
        result = cloudinary.uploader.upload(
    file,
    folder=folder,
    resource_type='auto',
    use_filename=True,
    unique_filename=True,
    access_mode='public' if public else 'authenticated'
)
        return result.get('secure_url'), result.get('public_id')
    except Exception as e:
        print(f'Cloudinary upload error: {e}')
        return None, None

def signed_document_url(file_url, public_id):
    """Signed Cloudinary delivery URL for an authenticated document asset.
    Falls back to the stored URL for legacy files without a public_id."""
    if not public_id:
        return file_url
    resource_type = 'raw' if (file_url and '/raw/upload/' in file_url) else \
                    'video' if (file_url and '/video/upload/' in file_url) else 'image'
    fmt = None
    if file_url:
        tail = file_url.rsplit('/', 1)[-1].split('?')[0]
        if '.' in tail:
            ext = tail.rsplit('.', 1)[-1].lower()
            # raw public_ids already include the extension — don't double it
            if not public_id.lower().endswith('.' + ext):
                fmt = ext
    url, _ = cloudinary.utils.cloudinary_url(
        public_id, resource_type=resource_type, type='upload',
        format=fmt, sign_url=True, secure=True)
    return url
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timezone
DUBAI_TZ = timezone(timedelta(hours=4))
def now_dubai():
    return datetime.now(DUBAI_TZ).replace(tzinfo=None)
from functools import wraps
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
# Trust exactly ONE proxy hop (Railway's edge) for X-Forwarded-For / -Proto.
# This makes request.remote_addr the real client IP that connected to Railway —
# attacker-supplied X-Forwarded-For entries to the left are ignored, so the login
# throttle can't be bypassed by spoofing the header. (Cloudflare is DNS-only /
# grey-cloud here, so Railway is the only proxy in front of the app.)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
# Secret key signs all session cookies. MUST be set via the SECRET_KEY env var in Railway.
# Security policy:
#   - Production (a real DATABASE_URL is set, i.e. Postgres on Railway): SECRET_KEY is
#     MANDATORY. We refuse to boot without it rather than fall back to a public,
#     git-history key that would let anyone forge admin sessions.
#   - Local dev (no DATABASE_URL / SQLite): generate a random per-boot key so nothing
#     insecure is ever hard-coded. (Restarting logs local dev sessions out — that's fine.)
_secret = os.environ.get('SECRET_KEY')
_is_production = bool(os.environ.get('DATABASE_URL'))
if not _secret:
    if _is_production:
        raise RuntimeError(
            'SECRET_KEY environment variable is not set. Set it in Railway and redeploy. '
            'Refusing to start with an insecure fallback key.')
    import secrets as _secrets
    _secret = _secrets.token_hex(32)
    print('[!] SECRET_KEY not set — using a random per-boot key for local dev only.')
app.secret_key = _secret
# Session configuration for custom domain support
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True  # HTTPS only
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///' + os.path.join(basedir, 'tahfeel.db')).replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ── CSRF protection ───────────────────────────────────────────────────────────
# Protects every POST/PUT/PATCH/DELETE. The token is auto-injected into every
# HTML form and every fetch/XHR by the script added in _inject_csrf() below, so
# individual templates don't each need editing. Webhooks (verified by HMAC
# signature instead) are exempted after their blueprints are registered.
from flask_wtf.csrf import CSRFProtect, generate_csrf, CSRFError
csrf = CSRFProtect(app)
# Give the token a long life so a page left open overnight doesn't fail on submit.
app.config['WTF_CSRF_TIME_LIMIT'] = None

@app.errorhandler(CSRFError)
def _handle_csrf_error(e):
    # Friendlier than a raw 400 — usually a stale tab; ask them to retry.
    flash('Your session expired or the form was stale. Please try that again.', 'warning')
    return redirect(request.referrer or url_for('dashboard'))

# Client-side shim: adds the CSRF token to every non-GET form submit and every
# fetch/XMLHttpRequest, so all existing templates and AJAX calls are covered
# without per-file edits. Injected right before </body> on HTML responses.
_CSRF_JS_TEMPLATE = """<script>(function(){var t=%s;
function inject(f){if(f&&f.tagName==='FORM'&&!f.querySelector('input[name=\"csrf_token\"]')){var i=document.createElement('input');i.type='hidden';i.name='csrf_token';i.value=t;f.appendChild(i);}}
document.addEventListener('submit',function(e){var f=e.target;if(f&&f.tagName==='FORM'){var m=(f.getAttribute('method')||'get').toLowerCase();if(m!=='get')inject(f);}},true);
document.addEventListener('DOMContentLoaded',function(){document.querySelectorAll('form').forEach(function(f){var m=(f.getAttribute('method')||'get').toLowerCase();if(m!=='get')inject(f);});});
// State-changing links are now POST-only routes. Intercept clicks on them and
// submit as a POST form carrying the CSRF token. Runs in the bubble phase so any
// inline onclick="return confirm(...)" has already decided: if the user cancelled
// (defaultPrevented) we do nothing.
var MUT=/\\/(delete|toggle|toggle-leave)$/;
var MUT_EXACT={'/admin/secure-documents':1,'/admin/alerts/disable-all':1};
document.addEventListener('click',function(e){
  if(e.defaultPrevented)return;
  var a=e.target.closest?e.target.closest('a[href]'):null;if(!a)return;
  var href=a.getAttribute('href')||'';if(!href||href.charAt(0)!=='/')return;
  var path=href.split('?')[0];
  if(MUT.test(path)||MUT_EXACT[path]){
    e.preventDefault();
    var f=document.createElement('form');f.method='POST';f.action=href;f.style.display='none';
    var i=document.createElement('input');i.type='hidden';i.name='csrf_token';i.value=t;f.appendChild(i);
    document.body.appendChild(f);f.submit();
  }
},false);
var of=window.fetch;if(of){window.fetch=function(u,o){o=o||{};var m=(o.method||'get').toUpperCase();if(m!=='GET'&&m!=='HEAD'){if(o.headers instanceof Headers){if(!o.headers.has('X-CSRFToken'))o.headers.set('X-CSRFToken',t);}else{o.headers=o.headers||{};if(!o.headers['X-CSRFToken'])o.headers['X-CSRFToken']=t;}}return of(u,o);};}
var oo=XMLHttpRequest.prototype.open,os=XMLHttpRequest.prototype.send;
XMLHttpRequest.prototype.open=function(m){this._csrfM=m;return oo.apply(this,arguments);};
XMLHttpRequest.prototype.send=function(){try{if(this._csrfM&&['POST','PUT','PATCH','DELETE'].indexOf(String(this._csrfM).toUpperCase())>=0)this.setRequestHeader('X-CSRFToken',t);}catch(e){}return os.apply(this,arguments);};
})();</script>"""

@app.after_request
def _inject_csrf(resp):
    try:
        if resp.direct_passthrough:
            return resp
        if 'text/html' not in resp.headers.get('Content-Type', ''):
            return resp
        body = resp.get_data(as_text=True)
        if '</body>' in body:
            import json as _json
            script = _CSRF_JS_TEMPLATE % _json.dumps(generate_csrf())
            resp.set_data(body.replace('</body>', script + '</body>', 1))
    except Exception as e:
        print(f'[csrf inject] {e}')
    return resp

@app.after_request
def set_security_headers(resp):
    # Defensive headers — safe defaults that don't affect existing pages.
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['X-Frame-Options'] = 'SAMEORIGIN'           # blocks clickjacking
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    resp.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    # HSTS: site is already HTTPS-only (SESSION_COOKIE_SECURE=True)
    resp.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return resp

@app.before_request
def restrict_marketing_role():
    # External Marketing-Ext users are sandboxed to the read-only lead report only.
    if session.get('role') != 'marketing':
        return
    allowed = {'marketing_report', 'marketing_export', 'logout', 'static'}
    if (request.endpoint or '') in allowed:
        return
    return redirect(url_for('marketing_report'))

@app.before_request
def restrict_finance_whatsapp():
    # Finance has no need for WhatsApp — block all /whatsapp pages + the unread API.
    if session.get('role') != 'finance':
        return
    path = request.path or ''
    if path.startswith('/whatsapp') or path == '/api/whatsapp-unread-count':
        if request.path == '/api/whatsapp-unread-count':
            return jsonify({'unread': 0})
        flash('WhatsApp is not available for the Finance role.', 'warning')
        return redirect(url_for('dashboard'))

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    print(f"500 ERROR: {error}")
    import traceback
    traceback.print_exc()
    return "<h2>Something went wrong. Please <a href='/'>try again</a> or <a href='/logout'>logout and login</a>.</h2>", 500

@app.errorhandler(404)
def not_found(error):
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='staff')
    is_super = db.Column(db.Boolean, default=False)  # Super Admin: can edit other admins + self
    active = db.Column(db.Boolean, default=True)
    phone = db.Column(db.String(20), nullable=True)
    on_leave = db.Column(db.Boolean, default=False)  # Excludes from Meta lead rotation
    report_from = db.Column(db.Date, nullable=True)  # Marketing-Ext: earliest lead date this user may see

class Lead(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    company = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    address = db.Column(db.String(200))
    source = db.Column(db.String(50))
    service = db.Column(db.String(100))
    representative = db.Column(db.String(100))
    lead_type = db.Column(db.String(20), default='New')
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime, default=lambda: now_dubai() + timedelta(days=1))
    remarks = db.Column(db.Text)
    status = db.Column(db.String(50), default='New')
    created_at = db.Column(db.DateTime, default=datetime.now)
    customer_story = db.Column(db.Text)
    phone2 = db.Column(db.String(20))
    campaign = db.Column(db.String(100))
    meta_lead_id = db.Column(db.String(50), nullable=True, unique=True)  # Prevents duplicate Meta leads
    # ── Lead redesign: quality flag (manual) + channel + timing ──
    genuine = db.Column(db.String(20))            # Genuine / Junk / Unreachable / None (unreviewed)
    junk_reason = db.Column(db.String(100))       # reason when marked Junk/Unreachable
    sub_source = db.Column(db.String(50))         # channel within a source, e.g. Facebook / Instagram
    first_contacted_at = db.Column(db.DateTime)   # first time the lead was actually reached
    attempts = db.Column(db.Integer, default=0)   # contact-attempt counter (info only, never auto-acts)
    assignee = db.relationship('User', foreign_keys=[assigned_to])
    updates = db.relationship('LeadUpdate', backref='lead', lazy=True, order_by='LeadUpdate.created_at.desc()')

class LeadUpdate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=False)
    stage = db.Column(db.String(50))
    activity_type = db.Column(db.String(50))  # Call—connected / Call—no answer / WhatsApp / Email / Quote / Meeting / Note
    remark = db.Column(db.Text, nullable=False)
    staff_name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.now)
    followup_date = db.Column(db.DateTime)
    lost_reason = db.Column(db.String(100))
    future_potential = db.Column(db.String(20))

class Enquiry(db.Model):
    """A quick question/callback capture (lighter than a Lead). Staff log it fast,
    set a remind date, then Resolve it or Convert it to a Lead. Kept separate so raw
    enquiries don't inflate lead metrics; resolution time is tracked for analytics."""
    id           = db.Column(db.Integer, primary_key=True)
    name         = db.Column(db.String(120))
    phone        = db.Column(db.String(30))
    enquiry      = db.Column(db.Text, nullable=False)   # what they asked
    service      = db.Column(db.String(100))            # optional service of interest
    assigned_to  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    remind_date  = db.Column(db.Date, nullable=True)    # follow-up / remind me
    status       = db.Column(db.String(20), default='Open')   # Open / Resolved
    remarks      = db.Column(db.Text)
    created_by   = db.Column(db.String(100))
    created_at   = db.Column(db.DateTime, default=now_dubai)
    resolved_at  = db.Column(db.DateTime, nullable=True)
    converted_lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=True)
    assignee     = db.relationship('User', foreign_keys=[assigned_to])

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(20), default='Medium')  # Low, Medium, High
    status = db.Column(db.String(20), default='Pending')   # Pending, In Progress, Done
    created_at = db.Column(db.DateTime, default=datetime.now)
    lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=True)
    assignee = db.relationship('User', foreign_keys=[assigned_to])
    creator = db.relationship('User', foreign_keys=[created_by])
    lead = db.relationship('Lead', foreign_keys=[lead_id])

class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class Source(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class Campaign(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class ServiceType(db.Model):
    __tablename__ = 'job_type'  # keep same DB table name
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    default_days = db.Column(db.Integer, default=1)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    company = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    phone2 = db.Column(db.String(20))
    email = db.Column(db.String(100))
    address = db.Column(db.String(200))
    source = db.Column(db.String(50))
    nationality = db.Column(db.String(50))
    date_of_birth = db.Column(db.Date, nullable=True)
    customer_type = db.Column(db.String(20), default='Individual')
    contact_person = db.Column(db.String(100))  # used for Company-type customers
    alerts_enabled = db.Column(db.Boolean, default=False)  # document-expiry alerts (email/WhatsApp)
    alert_email = db.Column(db.String(120))
    alert_whatsapp = db.Column(db.String(30))
    # Company profile (UAE) — used for Company-type customers
    ac_code = db.Column(db.String(50))
    trade_name = db.Column(db.String(150))
    legal_form = db.Column(db.String(60))           # LLC, Sole Est, Branch, Free Zone Co, Offshore...
    jurisdiction = db.Column(db.String(30))         # Mainland / Free Zone / Offshore
    licensing_authority = db.Column(db.String(120)) # DED, DMCC, JAFZA, IFZA...
    freezone_name = db.Column(db.String(120))
    emirate = db.Column(db.String(40))
    country_incorp = db.Column(db.String(60))
    business_activity = db.Column(db.String(200))
    ac_status = db.Column(db.String(30))            # Active / Under Formation / Inactive / Closed
    po_box = db.Column(db.String(30))
    mobile = db.Column(db.String(30))
    whatsapp = db.Column(db.String(30))
    website = db.Column(db.String(120))
    ac_opening_date = db.Column(db.Date)
    uae_pass_number = db.Column(db.String(50))   # UAE Pass access / account number
    uae_pass_name = db.Column(db.String(100))    # name on the UAE Pass account
    # Tax filing tracking (shown as Compliance cards; auto-roll on Filed)
    vat_status = db.Column(db.String(20))        # 'Filed' / 'Not filed'
    vat_due_date = db.Column(db.Date)            # next VAT filing due
    corp_tax_status = db.Column(db.String(20))   # 'Filed' / 'Not filed'
    corp_tax_due_date = db.Column(db.Date)       # next corporate-tax filing due
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=True)
    rep = db.relationship('User', foreign_keys=[assigned_to])
    lead = db.relationship('Lead', foreign_keys=[lead_id])
    jobs = db.relationship('Job', backref='customer', lazy=True, order_by='Job.created_at.desc()')

class Job(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    job_type = db.Column(db.String(100), nullable=False)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(20), default='Medium')
    status = db.Column(db.String(50), default='Pending Finance Approval')
    internal_notes = db.Column(db.Text)
    service_note = db.Column(db.String(200))
    amount_invoiced = db.Column(db.Float, default=0)
    amount_received = db.Column(db.Float, default=0)
    num_persons = db.Column(db.Integer, default=1)
    finance_approved_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    finance_approved_at = db.Column(db.DateTime, nullable=True)
    finance_notes = db.Column(db.Text)
    # Completion fields
    final_remarks = db.Column(db.Text, nullable=True)
    future_work_notes = db.Column(db.Text, nullable=True)
    completed_at = db.Column(db.DateTime, nullable=True)
    # Partner commission fields
    partner_commission_expected = db.Column(db.Boolean, default=False)
    partner_name = db.Column(db.String(100))
    partner_amount = db.Column(db.Float)
    partner_due_date = db.Column(db.Date)
    partner_status = db.Column(db.String(20), default='Pending')  # Pending/Received/Written Off
    partner_received_date = db.Column(db.Date)
    revenue = db.Column(db.Float, default=0)  # Revenue counted when no partner OR when partner pays
    revenue_date = db.Column(db.Date)  # Date when revenue is counted (for cash-basis accounting)
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    assignee = db.relationship('User', foreign_keys=[assigned_to])
    creator = db.relationship('User', foreign_keys=[created_by])
    finance_approver = db.relationship('User', foreign_keys=[finance_approved_by])
    updates = db.relationship('JobUpdate', backref='job', lazy=True, order_by='JobUpdate.created_at.desc()')
    subtasks = db.relationship('SubTask', backref='job', lazy=True, order_by='SubTask.created_at')
    partial_revenues = db.relationship('PartialRevenue', backref='job', lazy=True, order_by='PartialRevenue.revenue_date.desc()')

class PartialRevenue(db.Model):
    __tablename__ = 'partial_revenue'
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    revenue_date = db.Column(db.Date, nullable=False)
    notes = db.Column(db.String(500))
    recorded_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.now)
    recorder = db.relationship('User', foreign_keys=[recorded_by])

class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)  # owner (one customer → many companies)
    contact_person = db.Column(db.String(100))
    phone = db.Column(db.String(30))
    email = db.Column(db.String(120))
    trade_license_no = db.Column(db.String(100))
    authority = db.Column(db.String(150))
    address = db.Column(db.String(255))
    notes = db.Column(db.Text)
    # Document-expiry alert settings (per company)
    alerts_enabled = db.Column(db.Boolean, default=False)
    alert_email = db.Column(db.String(120))
    alert_whatsapp = db.Column(db.String(30))
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    owner = db.relationship('Customer', foreign_keys=[customer_id])
    documents = db.relationship('Document', backref='company', lazy=True)

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)  # the company they belong to
    name = db.Column(db.String(120), nullable=False)
    designation = db.Column(db.String(100))
    nationality = db.Column(db.String(60))
    date_of_birth = db.Column(db.Date)
    join_date = db.Column(db.Date)
    mobile = db.Column(db.String(30))
    email = db.Column(db.String(120))
    status = db.Column(db.String(30), default='Active')
    created_at = db.Column(db.DateTime, default=datetime.now)
    company = db.relationship('Customer', foreign_keys=[customer_id])
    documents = db.relationship('Document', backref='employee', lazy=True)

class Owner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)  # the company
    name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(60))            # Shareholder / Director / Manager / Authorized Signatory
    share_pct = db.Column(db.Float)
    nationality = db.Column(db.String(60))
    passport_no = db.Column(db.String(60))
    passport_expiry = db.Column(db.Date)
    eid_no = db.Column(db.String(60))
    eid_expiry = db.Column(db.Date)
    date_of_birth = db.Column(db.Date)         # drives the auto birthday WhatsApp
    mobile = db.Column(db.String(30))          # owner's own WhatsApp — birthday wish target
    created_at = db.Column(db.DateTime, default=datetime.now)
    company = db.relationship('Customer', foreign_keys=[customer_id])

class Partner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

class SubTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    service_type = db.Column(db.String(100))
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(20), default='Medium')
    status = db.Column(db.String(20), default='Pending')
    amount = db.Column(db.Float, default=0)   # optional add-on revenue (e.g. Bank Account Opening); informational
    remarks = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    completed_at = db.Column(db.DateTime, nullable=True)
    assignee = db.relationship('User', foreign_keys=[assigned_to])

class SubTaskTemplate(db.Model):
    """One step inside a named, reusable sub-task GROUP (e.g. "Residence Visa" ->
    E-ID typing, Medical, Stamping…). A group can be added to ANY task from a
    dropdown. (The `job_type` column holds the group name.)"""
    __tablename__ = 'subtask_template'
    id         = db.Column(db.Integer, primary_key=True)
    job_type   = db.Column(db.String(100), index=True)   # the GROUP name
    title      = db.Column(db.String(200), nullable=False)
    sort_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)

class DocType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    doc_type = db.Column(db.String(100), nullable=False)
    belongs_to = db.Column(db.String(20), nullable=False)  # Company / Individual / Staff
    owner_name = db.Column(db.String(100), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    company_id = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=True)  # attach doc to a company
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)  # attach doc to an employee
    expiry_date = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text)
    file_name = db.Column(db.String(255), nullable=True)
    file_url = db.Column(db.Text, nullable=True)
    cloudinary_public_id = db.Column(db.String(255), nullable=True)
    added_by = db.Column(db.String(100), nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.now)
    uploader = db.relationship('User', foreign_keys=[uploaded_by])
    customer = db.relationship('Customer', foreign_keys=[customer_id])


class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    log_date = db.Column(db.Date, nullable=False)
    # Activity counts
    calls_existing = db.Column(db.Integer, default=0)
    calls_cold = db.Column(db.Integer, default=0)
    dm_instagram = db.Column(db.Integer, default=0)
    dm_facebook = db.Column(db.Integer, default=0)
    dm_linkedin = db.Column(db.Integer, default=0)
    posts_social = db.Column(db.Integer, default=0)
    videos_instagram = db.Column(db.Integer, default=0)
    linkedin_writing = db.Column(db.Integer, default=0)
    whatsapp_prospecting = db.Column(db.Integer, default=0)
    community_active = db.Column(db.Integer, default=0)
    google_reviews = db.Column(db.Integer, default=0)
    real_estate_relations = db.Column(db.Integer, default=0)
    content_marketing = db.Column(db.Integer, default=0)
    referral_building = db.Column(db.Integer, default=0)
    networking_activities = db.Column(db.Integer, default=0)
    networking_events = db.Column(db.Integer, default=0)
    off_day = db.Column(db.String(20), nullable=True)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    user = db.relationship('User', foreign_keys=[user_id])


class ActivityType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    field_key = db.Column(db.String(50), nullable=False, unique=True)
    label = db.Column(db.String(150), nullable=False)
    weekly_target = db.Column(db.Float, default=5)
    sort_order = db.Column(db.Integer, default=0)
    active = db.Column(db.Boolean, default=True)

class WhatsAppMessage(db.Model):
    """One WhatsApp message (in or out). Threaded by wa_id (the contact's number)."""
    __tablename__ = 'whats_app_message'
    id           = db.Column(db.Integer, primary_key=True)
    wa_id        = db.Column(db.String(30), index=True)   # contact phone (digits only)
    contact_name = db.Column(db.String(120))              # WhatsApp profile name
    direction    = db.Column(db.String(4))                # 'in' / 'out'
    body         = db.Column(db.Text)
    msg_type     = db.Column(db.String(20), default='text')  # text / template / image…
    wam_id       = db.Column(db.String(80), index=True)   # WhatsApp message id (dedupe)
    status       = db.Column(db.String(20))               # sent/delivered/read/failed
    error        = db.Column(db.String(300))              # Meta's reason when a send fails
    handled_by   = db.Column(db.String(40), default='bot')  # 'bot' or staff name
    is_read      = db.Column(db.Boolean, default=False)   # inbound: has a staff seen it?
    lead_id      = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=True)
    customer_id  = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    media_url    = db.Column(db.String(500), nullable=True)  # Cloudinary permanent URL
    mime_type    = db.Column(db.String(50), nullable=True)
    created_at   = db.Column(db.DateTime, default=now_dubai)
    lead         = db.relationship('Lead', foreign_keys=[lead_id])
    customer     = db.relationship('Customer', foreign_keys=[customer_id])

class WhatsAppThread(db.Model):
    """Thread-level state for a WhatsApp conversation (one row per contact)."""
    __tablename__ = 'whats_app_thread'
    wa_id         = db.Column(db.String(30), primary_key=True)
    assigned_to   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    assigned_at   = db.Column(db.DateTime, nullable=True)
    bot_paused    = db.Column(db.Boolean, default=False)
    bot_paused_by = db.Column(db.String(100), nullable=True)
    resolved      = db.Column(db.Boolean, default=False)   # 'Done' — drops out of the active list
    resolved_at   = db.Column(db.DateTime, nullable=True)
    resolved_by   = db.Column(db.String(100), nullable=True)
    assignee      = db.relationship('User', foreign_keys=[assigned_to])

class QuickReply(db.Model):
    """Canned reply staff can insert with one click in a WhatsApp thread."""
    __tablename__ = 'quick_reply'
    id         = db.Column(db.Integer, primary_key=True)
    label      = db.Column(db.String(100), nullable=False)
    body       = db.Column(db.Text, nullable=False)
    is_global  = db.Column(db.Boolean, default=False)
    staff_id   = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=now_dubai)
    staff      = db.relationship('User', foreign_keys=[staff_id])

class MessageTemplate(db.Model):
    """A WhatsApp template approved in Meta Business Manager, registered in the CRM
    so staff can send it outside the 24h window. meta_name + lang must match Meta
    exactly; body_preview should be kept identical to the approved body (it is what
    the CRM shows in previews and logs — Meta sends its own approved text)."""
    __tablename__ = 'message_template'
    id           = db.Column(db.Integer, primary_key=True)
    label        = db.Column(db.String(100), nullable=False)   # friendly name staff see
    meta_name    = db.Column(db.String(120), nullable=False)   # exact template name in Meta
    category     = db.Column(db.String(20), default='Utility') # Utility / Marketing
    lang         = db.Column(db.String(10), default='en')
    body_preview = db.Column(db.Text, nullable=False)          # body with {{1}}, {{2}}…
    var_fields   = db.Column(db.String(300))                   # comma-separated keys, one per variable
    active       = db.Column(db.Boolean, default=True)
    created_at   = db.Column(db.DateTime, default=now_dubai)

class AppSetting(db.Model):
    """Simple key/value store for CRM settings that used to live in Railway env vars,
    so admins can change them in the CRM instead of touching Railway."""
    __tablename__ = 'app_setting'
    key   = db.Column(db.String(60), primary_key=True)
    value = db.Column(db.String(300))

class LoginAttempt(db.Model):
    """One failed-login timestamp per row, keyed by client IP. DB-backed so the
    brute-force throttle works across ALL gunicorn workers and survives restarts
    (the old in-memory dict was per-process and reset on every redeploy)."""
    __tablename__ = 'login_attempt'
    id         = db.Column(db.Integer, primary_key=True)
    ip         = db.Column(db.String(64), index=True)
    created_at = db.Column(db.DateTime, default=now_dubai, index=True)

def get_setting(key, default=None):
    try:
        row = AppSetting.query.get(key)
        return row.value if (row and row.value is not None) else default
    except Exception:
        return default

def set_setting(key, value):
    row = AppSetting.query.get(key)
    if not row:
        row = AppSetting(key=key)
        db.session.add(row)
    row.value = value
    db.session.commit()

# Default state for each automation toggle. Existing email jobs default ON so this
# change never silently turns off something that is already running; the new
# WhatsApp automations default OFF so nothing new fires until an admin enables it.
AUTOMATION_DEFAULTS = {
    'wa_auto_welcome':     'off',  # WhatsApp welcome to new leads (existing)
    'auto_birthday':       'off',  # WhatsApp birthday wish (new)
    'auto_expiry_wa':      'off',  # WhatsApp expiry reminder at 7 & 3 days (new)
    'auto_expiry_email':   'on',   # weekly document-expiry email (existing behaviour)
    'auto_monthly_report': 'on',   # monthly compliance report email (existing behaviour)
}

def automation_on(key):
    return get_setting(key, AUTOMATION_DEFAULTS.get(key, 'off')) == 'on'

def _mark_run(name, summary):
    """Record when an automation last ran + a one-line summary, for the admin panel."""
    try:
        set_setting(f'run_{name}', f'{now_dubai().strftime("%d %b %Y %H:%M")} — {summary}')
    except Exception:
        pass

class AutoMessageLog(db.Model):
    """Dedupe + audit trail for automatic WhatsApp sends (birthday, expiry).
    `dedupe_key` is unique so a re-run of a cron can never double-send."""
    __tablename__ = 'auto_message_log'
    id         = db.Column(db.Integer, primary_key=True)
    kind       = db.Column(db.String(30), index=True)          # birthday / expiry_wa
    dedupe_key = db.Column(db.String(160), unique=True, index=True)
    detail     = db.Column(db.String(200))
    sent_at    = db.Column(db.DateTime, default=now_dubai, index=True)

# ── Meta Conversions API (CAPI) ───────────────────────────────────────────────
# Sends a "good lead" signal back to Meta when a lead is marked Genuine, so Lead
# Ads optimise for quality, not just form-fills. Fully DORMANT until an admin sets
# the token + dataset id and flips it on. Config lives in AppSetting:
#   capi_enabled ('on'/'off', default off) · capi_token · capi_dataset_id
#   capi_event_name (default 'Qualified') · capi_test_code (optional, for testing)
def capi_send_lead_quality(lead):
    """Fire one CAPI event for a Genuine lead. Only for Meta-sourced leads (need the
    Meta lead id to attribute to the ad). Deduped per lead+event. Never raises."""
    try:
        if get_setting('capi_enabled', 'off') != 'on':
            return
        token = get_setting('capi_token', '')
        dataset = get_setting('capi_dataset_id', '')
        if not token or not dataset:
            return
        if not lead or not lead.meta_lead_id:
            return  # only Meta-ad leads can be matched back to a campaign
        event_name = get_setting('capi_event_name', 'Qualified') or 'Qualified'
        key = f'capi:{lead.id}:{event_name}'
        if AutoMessageLog.query.filter_by(dedupe_key=key).first():
            return  # already sent this signal for this lead
        import hashlib, time, requests
        def _h(v):
            return hashlib.sha256(str(v).strip().lower().encode()).hexdigest() if v else None
        # Meta lead id attributes to the ad; hashed phone/email help matching.
        try:
            lead_id_val = int(lead.meta_lead_id)
        except (TypeError, ValueError):
            lead_id_val = lead.meta_lead_id
        user_data = {'lead_id': lead_id_val}
        ph = ''.join(ch for ch in (lead.phone or '') if ch.isdigit())
        if ph:
            user_data['ph'] = [hashlib.sha256(ph.encode()).hexdigest()]
        if lead.email:
            user_data['em'] = [_h(lead.email)]
        payload = {
            'data': [{
                'event_name': event_name,
                'event_time': int(time.time()),
                'action_source': 'system_generated',
                'user_data': user_data,
                'custom_data': {'lead_event_source': 'crm', 'crm': 'Tahfeel CRM'},
            }],
            'access_token': token,
        }
        test_code = get_setting('capi_test_code', '')
        if test_code:
            payload['test_event_code'] = test_code
        r = requests.post(f'https://graph.facebook.com/v19.0/{dataset}/events',
                          json=payload, timeout=10)
        db.session.add(AutoMessageLog(kind='capi', dedupe_key=key,
                                      detail=f'{event_name} · {(lead.name or lead.id)} · HTTP {r.status_code}'))
        set_setting('run_capi', f'{now_dubai().strftime("%d %b %Y %H:%M")} — {event_name} -> HTTP {r.status_code}')
        db.session.commit()
        print(f'[CAPI] {event_name} sent for lead {lead.id}: HTTP {r.status_code} {r.text[:200]}')
    except Exception as e:
        print(f'[CAPI] send failed: {e}')

def wa_template_active(meta_name):
    """Return an active MessageTemplate by its Meta name, or None."""
    try:
        return MessageTemplate.query.filter_by(meta_name=meta_name, active=True).first()
    except Exception:
        return None

# Auto-fill keys available for template variables (order in var_fields = {{1}}, {{2}}…)
WA_VAR_LABELS = {
    'first_name': 'First name',
    'full_name':  'Full name',
    'company':    'Company name',
    'job_type':   'Service / task type',
    'service':    'Service enquired (from the lead)',
    'custom':     'Custom text (typed at send time)',
}

def _wa_resolve_var(key, customer=None, job=None):
    """Best-effort auto-fill of one template variable from CRM data."""
    if key == 'first_name' and customer:
        base = (customer.contact_person or customer.name or '').strip()
        return base.split()[0] if base else ''
    if key == 'full_name' and customer:
        return (customer.contact_person or customer.name or '').strip()
    if key == 'company' and customer:
        return (customer.trade_name or customer.company
                or (customer.name if customer.customer_type == 'Company' else '') or '').strip()
    if key == 'job_type' and job:
        return job.job_type or ''
    return ''

def wa_send_context(customer=None, job=None):
    """Active templates + pre-filled variable values for the send modal."""
    if job and customer is None:
        customer = job.customer
    out = []
    for t in MessageTemplate.query.filter_by(active=True).order_by(MessageTemplate.label).all():
        keys = [k.strip() for k in (t.var_fields or '').split(',') if k.strip()]
        params = [{'n': i, 'key': k,
                   'label': WA_VAR_LABELS.get(k, k),
                   'value': _wa_resolve_var(k, customer, job)}
                  for i, k in enumerate(keys, start=1)]
        out.append({'id': t.id, 'label': t.label, 'category': t.category,
                    'lang': t.lang, 'body': t.body_preview, 'params': params})
    return out

class Broadcast(db.Model):
    """One bulk WhatsApp send to a filtered group. Individual messages are logged
    to WhatsAppMessage as usual; this row tracks the campaign + live progress."""
    __tablename__ = 'broadcast'
    id             = db.Column(db.Integer, primary_key=True)
    template_id    = db.Column(db.Integer, db.ForeignKey('message_template.id'))
    template_label = db.Column(db.String(100))
    filter_summary = db.Column(db.String(400))
    total          = db.Column(db.Integer, default=0)
    sent           = db.Column(db.Integer, default=0)
    failed         = db.Column(db.Integer, default=0)
    status         = db.Column(db.String(20), default='sending')  # sending / done
    created_by     = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at     = db.Column(db.DateTime, default=now_dubai)
    creator        = db.relationship('User', foreign_keys=[created_by])

def broadcast_filter_customers(args):
    """Build the customer list for a broadcast from filter args (works for the
    preview page, Excel export, and the send loop — one source of truth)."""
    q = Customer.query
    activity = (args.get('business_activity') or '').strip()
    if activity:
        q = q.filter(Customer.business_activity.ilike(f'%{activity}%'))
    for field in ('jurisdiction', 'emirate', 'licensing_authority', 'customer_type', 'nationality'):
        val = (args.get(field) or '').strip()
        if val and val != 'All':
            q = q.filter(getattr(Customer, field) == val)
    custs = q.order_by(Customer.name).all()
    # Document-expiry window is across the customer's documents — filter in Python
    expiry_days = args.get('expiry_days', type=int)
    if expiry_days:
        today = now_dubai().date()
        keep = []
        for c in custs:
            docs = Document.query.filter_by(customer_id=c.id)\
                    .filter(Document.expiry_date.isnot(None)).all()
            if any(0 <= (d.expiry_date.date() - today).days <= expiry_days for d in docs):
                keep.append(c)
        custs = keep
    return custs

def _cust_wa_number(c):
    from whatsapp_webhook import normalize_phone
    return normalize_phone(c.whatsapp or c.mobile or c.phone or c.phone2 or '')

class JobUpdate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    status = db.Column(db.String(50))
    status_note = db.Column(db.String(100))
    remark = db.Column(db.Text, nullable=False)
    staff_name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.now)

class CompanyDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    doc_type = db.Column(db.String(100), nullable=False)  # E-migration Card, E-channel
    issue_date = db.Column(db.Date)
    expiry_date = db.Column(db.Date, nullable=False)
    authority = db.Column(db.String(255))  # Issuer/Authority
    owner = db.Column(db.String(255), nullable=False)  # Staff name or "Company"
    document_url = db.Column(db.String(500))  # Cloudinary URL
    cloudinary_public_id = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.String(100))
    category = db.Column(db.String(20))  # Tahfeel / Staff / Management
    staff_id = db.Column(db.Integer, db.ForeignKey('tahfeel_staff.id'), nullable=True)  # for Staff/Management docs
    staff = db.relationship('TahfeelStaff')

    def days_until_expiry(self):
        if self.expiry_date:
            delta = self.expiry_date - now_dubai().date()
            return delta.days
        return None
    
    def expiry_status(self):
        days = self.days_until_expiry()
        if days is None:
            return 'unknown'
        elif days < 0:
            return 'expired'
        elif days < 30:
            return 'critical'  # Red
        elif days < 60:
            return 'warning'  # Yellow
        else:
            return 'ok'  # Green

class TahfeelStaff(db.Model):
    """A person (Staff or Management) that Tahfeel documents belong to."""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    category = db.Column(db.String(20), nullable=False)  # Staff / Management
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

class MonthlyTarget(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    lead_target = db.Column(db.Integer, default=0)
    conversion_target = db.Column(db.Integer, default=0)
    amount_target = db.Column(db.Float, default=0)
    user = db.relationship('User', foreign_keys=[user_id])

class DeskNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    text = db.Column(db.Text, nullable=False)
    reminder_date = db.Column(db.Date, nullable=True)
    mention_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    is_done = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User', foreign_keys=[user_id])
    mention_user = db.relationship('User', foreign_keys=[mention_user_id])

@app.context_processor
def inject_globals():
    result = {'birthdays_today': [], 'show_backup_reminder': False}
    try:
        if 'user_id' in session and session.get('role') == 'admin':
            last_backup = session.get('last_backup_date')
            if not last_backup:
                result['show_backup_reminder'] = True
            else:
                from datetime import datetime as _dt
                last = _dt.strptime(last_backup, '%Y-%m-%d')
                if (now_dubai() - last).days >= 3:
                    result['show_backup_reminder'] = True
    except:
        pass
    return result

def wa_unread_count():
    """Unread WhatsApp badge on the nav. Reps see only chats assigned to THEM
    (their targeted alert); admins see all unread (oversight). Done/resolved
    chats never count."""
    try:
        if 'user_id' not in session:
            return 0
        resolved_ids = {t.wa_id for t in WhatsAppThread.query.filter_by(resolved=True).all()}
        q = WhatsAppMessage.query.filter_by(direction='in', is_read=False)
        if session.get('role') == 'admin':
            msgs = q.all()
        else:
            my_ids = {t.wa_id for t in WhatsAppThread.query.filter_by(assigned_to=session['user_id']).all()}
            if not my_ids:
                return 0
            msgs = q.filter(WhatsAppMessage.wa_id.in_(list(my_ids))).all()
        return sum(1 for m in msgs if m.wa_id not in resolved_ids)
    except Exception:
        return 0
app.jinja_env.globals['wa_unread_count'] = wa_unread_count

def job_status_label(s):
    """Display label for a job status. The stored value 'Done' is shown as
    'Done - Work Completed' to staff; the internal value is unchanged so all
    finance/revenue logic keeps working."""
    return 'Done - Work Completed' if s == 'Done' else (s or '')
app.jinja_env.globals['job_status_label'] = job_status_label

@app.context_processor
def inject_birthdays():
    try:
        if 'user_id' in session:
            today = now_dubai()
            bdays = []
            try:
                result = db.session.execute(db.text(
                    "SELECT id, name, phone FROM customer WHERE date_of_birth IS NOT NULL AND EXTRACT(MONTH FROM date_of_birth)=:m AND EXTRACT(DAY FROM date_of_birth)=:d"
                ), {'m': today.month, 'd': today.day}).fetchall()
                bdays = [{'id': r[0], 'name': r[1], 'phone': r[2]} for r in result]
            except:
                try:
                    all_c = db.session.execute(db.text("SELECT id, name, phone, date_of_birth FROM customer WHERE date_of_birth IS NOT NULL")).fetchall()
                    bdays = [{'id':r[0],'name':r[1],'phone':r[2]} for r in all_c if r[3] and r[3].month==today.month and r[3].day==today.day]
                except:
                    pass
            return {'birthdays_today': bdays}
    except Exception as e:
        print(f'Birthday error: {e}')
    return {'birthdays_today': []}

def _safe_redirect(target, fallback_endpoint='dashboard', **fallback_kwargs):
    """Redirect only to a same-site relative path. Blocks open-redirect payloads
    (e.g. //evil.com or https://evil.com) coming from user-supplied `next`/return
    URLs by falling back to a known-safe internal endpoint."""
    from urllib.parse import urlparse
    if target:
        p = urlparse(target)
        # Accept only relative paths (no scheme, no netloc) that start with a single '/'
        if not p.scheme and not p.netloc and target.startswith('/') and not target.startswith('//'):
            return redirect(target)
    return redirect(url_for(fallback_endpoint, **fallback_kwargs))

def _can_view_document(doc):
    """Authorization for viewing a customer/employee/company document file.
    Admin & Finance & Operations see all. Sales/Staff only see documents for a
    customer they are the assigned rep of, or documents they uploaded themselves."""
    role = session.get('role')
    if role in ('admin', 'finance', 'operations'):
        return True
    uid = session.get('user_id')
    if getattr(doc, 'uploaded_by', None) and doc.uploaded_by == uid:
        return True
    cust = getattr(doc, 'customer', None)
    if cust and cust.assigned_to == uid:
        return True
    return False

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def can_manage_user(target):
    """A regular admin may manage only non-admin staff. A super admin may manage
    anyone (including other admins and themselves). Returns True if allowed."""
    if session.get('is_super'):
        return True
    # Not a super admin: block managing admins or super admins
    return not (target.role == 'admin' or target.is_super)

def finance_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ['admin', 'finance']:
            flash('Finance access required')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def apply_lead_filters(leads, args, now):
    date_filter = args.get('date')
    status_filter = args.get('status')
    staff_filter = args.get('staff')
    due_filter = args.get('due')  # NEW: Due Date filter
    source_filter = args.get('source')
    
    if date_filter == 'today':
        leads = [l for l in leads if l.created_at and l.created_at.date() == now.date()]
    elif date_filter == 'week':
        week_start = now.date() - timedelta(days=now.weekday())
        week_end = week_start + timedelta(days=6)
        leads = [l for l in leads if l.created_at and week_start <= l.created_at.date() <= week_end]
    elif date_filter == 'month':
        leads = [l for l in leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
    elif date_filter == 'custom':
        from_date = args.get('from')
        to_date = args.get('to')
        if from_date:
            try:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                leads = [l for l in leads if l.created_at and l.created_at.date() >= from_dt]
            except: pass
        if to_date:
            try:
                to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                leads = [l for l in leads if l.created_at and l.created_at.date() <= to_dt]
            except: pass
    
    if status_filter:
        if status_filter == 'Overdue':
            leads = [l for l in leads if l.due_date < now and l.status not in ['Converted', 'Lost']]
        elif status_filter == 'Initiated':
            # Initiated = any action taken (not New, Converted, Lost, or Future)
            leads = [l for l in leads if l.status not in ['New', 'Converted', 'Lost', 'Future']]
        else:
            # Includes status_filter == 'Future' (shows all parked + due future leads)
            leads = [l for l in leads if l.status == status_filter]
    
    if staff_filter:
        try:
            sf = int(staff_filter)
            leads = [l for l in leads if l.assigned_to == sf]
        except: pass
    
    # NEW: Due Date filter
    if due_filter:
        if due_filter == 'overdue':
            leads = [l for l in leads if l.due_date and l.due_date < now and l.status not in ['Converted', 'Lost']]
        elif due_filter == 'today':
            leads = [l for l in leads if l.due_date and l.due_date.date() == now.date()]
        elif due_filter == 'tomorrow':
            tomorrow = now.date() + timedelta(days=1)
            leads = [l for l in leads if l.due_date and l.due_date.date() == tomorrow]
        elif due_filter == 'this_week':
            week_end = now.date() + timedelta(days=7)
            leads = [l for l in leads if l.due_date and now.date() <= l.due_date.date() <= week_end]
        elif due_filter == 'next_week':
            next_week_start = now.date() + timedelta(days=7)
            next_week_end = next_week_start + timedelta(days=7)
            leads = [l for l in leads if l.due_date and next_week_start <= l.due_date.date() <= next_week_end]
    
    if source_filter:
        leads = [l for l in leads if l.source == source_filter]

    quality_filter = args.get('quality')
    if quality_filter == 'unreviewed':
        leads = [l for l in leads if not l.genuine]
    elif quality_filter:
        leads = [l for l in leads if l.genuine == quality_filter]

    return leads

@app.route('/')
def index():
    return redirect(url_for('login'))

_LOGIN_MAX_ATTEMPTS = 10               # per window, per IP
_LOGIN_WINDOW = 300                    # seconds (5 minutes)

def _client_ip():
    # ProxyFix has already set remote_addr to the real client IP (rightmost
    # trusted hop), so this is not spoofable via a forged X-Forwarded-For.
    return request.remote_addr or 'unknown'

def _login_blocked(ip):
    """True if this IP has hit the failed-login cap within the window. DB-backed,
    so it holds across workers. Fails OPEN on any DB error (never lock everyone out)."""
    try:
        window_start = now_dubai() - timedelta(seconds=_LOGIN_WINDOW)
        return LoginAttempt.query.filter(
            LoginAttempt.ip == ip,
            LoginAttempt.created_at >= window_start
        ).count() >= _LOGIN_MAX_ATTEMPTS
    except Exception as e:
        print(f'[login-throttle] check skipped: {e}')
        return False

def _record_login_failure(ip):
    try:
        # Opportunistic prune of anything older than the window (keeps table tiny).
        window_start = now_dubai() - timedelta(seconds=_LOGIN_WINDOW)
        LoginAttempt.query.filter(LoginAttempt.created_at < window_start).delete()
        db.session.add(LoginAttempt(ip=ip))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'[login-throttle] record skipped: {e}')

def _clear_login_failures(ip):
    try:
        LoginAttempt.query.filter(LoginAttempt.ip == ip).delete()
        db.session.commit()
    except Exception:
        db.session.rollback()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        # Brute-force throttle (per IP, DB-backed so it holds across workers).
        ip = _client_ip()
        if _login_blocked(ip):
            flash('Too many failed attempts. Please wait a few minutes and try again.')
            return render_template('login.html', email=email), 429
        user = User.query.filter_by(email=email, active=True).first()
        if user and check_password_hash(user.password, password):
            session.permanent = True  # Enable persistent session
            session['user_id'] = user.id
            session['user_name'] = user.name
            session['user_email'] = user.email
            session['role'] = user.role
            session['is_super'] = bool(user.is_super)
            try:
                session['unread_mentions'] = DeskNote.query.filter_by(mention_user_id=user.id, is_done=False).count()
            except:
                session['unread_mentions'] = 0
            _clear_login_failures(ip)  # reset on success
            if user.role == 'marketing':
                return redirect(url_for('marketing_report'))
            return redirect(url_for('dashboard'))
        _record_login_failure(ip)  # record failed attempt
        flash('Invalid email or password')
        return render_template('login.html', email=email)
    return render_template('login.html', email='')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/healthz')
def healthz():
    """Lightweight health probe for an uptime monitor (e.g. free UptimeRobot).
    Returns 200 when the app + database are reachable, 503 otherwise. No login and
    no data is exposed — just a status word — so it's safe to leave public."""
    try:
        db.session.execute(db.text('SELECT 1'))
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        print(f'[healthz] DB check failed: {e}')
        return jsonify({'status': 'error'}), 503

@app.route('/api/lead-alerts')
@login_required
def api_lead_alerts():
    """Lightweight JSON feed for the new-lead bell/toast in the top bar.
    Admin sees all 'New' leads; everyone else sees leads assigned to them."""
    from flask import jsonify
    since = request.args.get('since', 0, type=int)
    base = Lead.query.filter(Lead.status == 'New')
    if session.get('role') != 'admin':
        base = base.filter(Lead.assigned_to == session['user_id'])
    count = base.count()
    # "unseen" = New leads newer than the last one the user acknowledged (bell click)
    unseen = base.filter(Lead.id > since).count()
    new_leads = base.order_by(Lead.id.desc()).limit(8).all()
    recent = [{
        'id': l.id,
        'name': l.name or 'Lead',
        'source': l.source or '',
        'created_at': l.created_at.strftime('%d %b, %I:%M %p') if l.created_at else '',
    } for l in new_leads]
    latest_id = new_leads[0].id if new_leads else 0
    return jsonify({'count': count, 'unseen': unseen, 'latest_id': latest_id, 'recent': recent})

@app.route('/dashboard')
@login_required
def dashboard():
    now = now_dubai()
    role = session['role']

    # ── Finance dashboard ────────────────────────────────────────────────────
    if role == 'finance':
        date_filter = request.args.get('date', 'month')  # Default to current month
        
        try:
            # Eager-load partial_revenues — the revenue loops below iterate them
            # per job, which would otherwise fire one query per job (N+1).
            all_jobs = Job.query.options(db.subqueryload(Job.partial_revenues)).order_by(Job.created_at.desc()).all()

            # Filter jobs by date
            if date_filter == 'today':
                jobs = [j for j in all_jobs if j.created_at and j.created_at.date() == now.date()]
            elif date_filter == 'week':
                week_start = now.date() - timedelta(days=now.weekday())  # Monday
                week_end = week_start + timedelta(days=6)  # Sunday
                jobs = [j for j in all_jobs if j.created_at and week_start <= j.created_at.date() <= week_end]
            elif date_filter == 'month':
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
            elif date_filter == 'all':
                jobs = all_jobs
            else:
                # Default to month
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
            
            active_jobs = [j for j in jobs if j.status != 'Done']
            pending_approval = [j for j in jobs if j.status in ['Pending Finance Approval', 'Done']]
            pending_close = [j for j in jobs if j.status == 'Pending Finance Close']
            total_invoiced = sum((j.amount_invoiced or 0) for j in active_jobs)
            total_received = sum((j.amount_received or 0) for j in active_jobs)
            total_pending = total_invoiced - total_received
            completed_value = sum((j.amount_received or 0) for j in jobs if j.status == 'Done')
            
            # Revenue calculations (closed tasks + partial revenues from in-progress tasks)
            # Revenue filtered by revenue_date (when closed), not created_at
            if date_filter == 'today':
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and j.revenue_date == now.date()]
                partial_revenue_jobs = all_jobs
            elif date_filter == 'week':
                week_start = now.date() - timedelta(days=now.weekday())
                week_end = week_start + timedelta(days=6)
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and week_start <= j.revenue_date <= week_end]
                partial_revenue_jobs = all_jobs
            elif date_filter == 'month':
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
                partial_revenue_jobs = all_jobs
            elif date_filter == 'all':
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission']]
                partial_revenue_jobs = all_jobs
            else:
                # Default to month
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
                partial_revenue_jobs = all_jobs
            
            total_revenue = sum((j.revenue or 0) for j in closed_jobs)
            
            # Add partial revenues from period (filter partial revenues by revenue_date too)
            partial_revenue_total = 0
            if date_filter == 'today':
                for j in partial_revenue_jobs:
                    for pr in j.partial_revenues:
                        if pr.revenue_date == now.date():
                            partial_revenue_total += pr.amount
            elif date_filter == 'week':
                for j in partial_revenue_jobs:
                    for pr in j.partial_revenues:
                        if week_start <= pr.revenue_date <= week_end:
                            partial_revenue_total += pr.amount
            elif date_filter == 'month':
                for j in partial_revenue_jobs:
                    for pr in j.partial_revenues:
                        if pr.revenue_date.year == now.year and pr.revenue_date.month == now.month:
                            partial_revenue_total += pr.amount
            else:  # all
                for j in partial_revenue_jobs:
                    partial_revenue_total += sum(pr.amount for pr in j.partial_revenues)
            
            total_revenue += partial_revenue_total
            
            # Customer Advances calculation (same as Admin dashboard)
            all_received = sum((j.amount_received or 0) for j in jobs)
            closed_received = sum((j.amount_received or 0) for j in closed_jobs)
            customer_advances = all_received - closed_received
            
            # Partner commission calculations
            partner_jobs = [j for j in all_jobs if j.partner_commission_expected and j.partner_status == 'Pending']
            total_partner_pending = sum((j.partner_amount or 0) for j in partner_jobs)
            
            # Monthly target
            targets = MonthlyTarget.query.filter_by(month=now.month, year=now.year).all()
            total_monthly_target = sum((t.amount_target or 0) for t in targets)
        except:
            active_jobs = pending_approval = pending_close = []
            total_invoiced = total_received = total_pending = completed_value = customer_advances = 0
            total_revenue = total_partner_pending = total_monthly_target = 0
        try:
            all_docs = Document.query.all()
            docs_30 = len([d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30])
            docs_60 = len([d for d in all_docs if d.expiry_date and 30 < (d.expiry_date - now).days <= 60])
            docs_90 = len([d for d in all_docs if d.expiry_date and 60 < (d.expiry_date - now).days <= 90])
            total_docs = len(all_docs)
        except:
            docs_30 = docs_60 = docs_90 = total_docs = 0
        all_active_jobs = Job.query.filter(Job.status.notin_(['Closed'])).all()
        tasks_active = len([j for j in all_active_jobs if j.status not in ['Pending Finance Approval','Pending Finance Close','Closed']])
        tasks_overdue = len([j for j in all_active_jobs if j.due_date and j.due_date < now and j.status not in ['Closed','Done']])
        tasks_processing = len([j for j in all_active_jobs if j.status == 'Processing'])
        tasks_pending_approval = len(pending_approval)
        return render_template('dashboard_finance.html',
                               now=now,
                               date_filter=date_filter,
                               docs_30=docs_30, docs_60=docs_60, docs_90=docs_90, total_docs=total_docs,
                               all_jobs=active_jobs,
                               pending_approval=pending_approval,
                               pending_close=pending_close,
                               tasks_active=tasks_active,
                               tasks_overdue=tasks_overdue,
                               tasks_processing=tasks_processing,
                               tasks_pending_approval=tasks_pending_approval,
                               total_invoiced=total_invoiced,
                               total_received=total_received,
                               total_pending=total_pending,
                               completed_value=completed_value,
                               customer_advances=customer_advances,
                               birthdays_today=[],
                               total_revenue=total_revenue,
                               total_partner_pending=total_partner_pending,
                               total_monthly_target=total_monthly_target)

    # ── Admin dashboard ──────────────────────────────────────────────────────
    if role == 'admin':
        all_leads = Lead.query.order_by(Lead.due_date).all()
        date_filter = request.args.get('date', 'month')  # DEFAULT TO CURRENT MONTH
        from_date = request.args.get('from', '')
        to_date = request.args.get('to', '')

        leads = all_leads
        try:
            # Eager-load partial_revenues (iterated per job in the revenue loops below).
            all_jobs = Job.query.options(db.subqueryload(Job.partial_revenues)).order_by(Job.due_date).all()
            if date_filter == 'today':
                leads = [l for l in all_leads if l.created_at and l.created_at.date() == now.date()]
                jobs = [j for j in all_jobs if j.created_at and j.created_at.date() == now.date()]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and j.revenue_date == now.date()]
            elif date_filter == 'week':
                week_start = now.date() - timedelta(days=now.weekday())  # Monday
                week_end = week_start + timedelta(days=6)  # Sunday
                leads = [l for l in all_leads if l.created_at and week_start <= l.created_at.date() <= week_end]
                jobs = [j for j in all_jobs if j.created_at and week_start <= j.created_at.date() <= week_end]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and week_start <= j.revenue_date <= week_end]
            elif date_filter == 'month':
                leads = [l for l in all_leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
                # For revenue (cash-basis): use revenue_date instead of created_at
                revenue_jobs = [j for j in all_jobs if j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
            elif date_filter == 'custom' and from_date and to_date:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                leads = [l for l in all_leads if l.created_at and from_dt <= l.created_at.date() <= to_dt]
                jobs = [j for j in all_jobs if j.created_at and from_dt <= j.created_at.date() <= to_dt]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and from_dt <= j.revenue_date <= to_dt]
            elif date_filter == 'all':
                # Show all time only if explicitly selected
                leads = all_leads  # FIX: Include all leads
                jobs = all_jobs
                revenue_jobs = all_jobs
            else:
                # Default to current month
                leads = [l for l in all_leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
            active_jobs = [j for j in jobs if j.status not in ['Done', 'Closed', 'Closed - Pending Partner Commission']]
            done_jobs = [j for j in jobs if j.status == 'Done']
            closed_jobs = [j for j in jobs if j.status in ['Closed', 'Closed - Pending Partner Commission']]
            # Revenue calculations use revenue_jobs (cash-basis)
            closed_revenue_jobs = [j for j in revenue_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission']]
            
            # Finance totals: Count ALL jobs (active + done + closed) for consistency
            all_invoiced = sum((j.amount_invoiced or 0) for j in jobs)
            all_received = sum((j.amount_received or 0) for j in jobs)
            closed_received = sum((j.amount_received or 0) for j in closed_jobs)
            
            # Display values
            total_invoiced = all_invoiced
            total_received = all_received
            total_pending = all_invoiced - all_received
            completed_value = closed_received  # Money from CLOSED tasks (not Done)
            
            # Customer Advances = Money received but work not closed yet
            customer_advances = all_received - closed_received
            
            try:
                total_revenue = sum((j.revenue or 0) for j in closed_revenue_jobs)
                
                # Add partial revenues filtered by revenue_date (same period as revenue_jobs)
                partial_revenue_total = 0
                if date_filter == 'today':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if pr.revenue_date == now.date():
                                partial_revenue_total += pr.amount
                elif date_filter == 'week':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if week_start <= pr.revenue_date <= week_end:
                                partial_revenue_total += pr.amount
                elif date_filter == 'month':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if pr.revenue_date.year == now.year and pr.revenue_date.month == now.month:
                                partial_revenue_total += pr.amount
                elif date_filter == 'custom':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if from_dt <= pr.revenue_date <= to_dt:
                                partial_revenue_total += pr.amount
                else:  # all
                    for j in all_jobs:
                        partial_revenue_total += sum(pr.amount for pr in j.partial_revenues)
                
                total_revenue += partial_revenue_total
                
                # Partner commission pending
                partner_jobs = [j for j in all_jobs if j.partner_commission_expected and j.partner_status == 'Pending']
                total_partner_pending = sum((j.partner_amount or 0) for j in partner_jobs)
            except:
                total_revenue = 0
                total_partner_pending = 0
            overdue_jobs = [j for j in jobs if j.due_date and j.due_date < now and j.status not in ['Done', 'Pending Finance Approval']]
            pending_approval = [j for j in jobs if j.status == 'Pending Finance Approval']
            pending_close = [j for j in jobs if j.status == 'Pending Finance Close']
            recent_jobs = [j for j in all_jobs if j.status not in ['Done', 'Closed', 'Pending Finance Approval']][:10]

            # Global task counts — match All Tasks page exactly
            closed_statuses_g = ['Closed', 'Closed - Pending Partner Commission']
            dash_stat_total = len(all_jobs)
            dash_stat_done_closed = len([j for j in all_jobs if j.status in ['Done'] + closed_statuses_g])
            dash_stat_overdue = len([j for j in all_jobs if j.due_date and j.due_date < now and j.status not in ['Done'] + closed_statuses_g])
            dash_stat_active = len([j for j in all_jobs if j.status not in ['Done'] + closed_statuses_g])
            dash_stat_pending_finance = len([j for j in all_jobs if j.status in ['Pending Finance Approval', 'Pending Finance Close']])
        except:
            jobs = all_jobs = active_jobs = done_jobs = closed_jobs = overdue_jobs = pending_approval = pending_close = recent_jobs = []
            total_invoiced = total_received = total_pending = completed_value = customer_advances = total_revenue = 0

        # Lead stats
        total = len(leads)
        overdue_leads = [l for l in leads if l.due_date and l.due_date < now and l.status not in ['Converted', 'Lost', 'Future']]
        converted = [l for l in leads if l.status == 'Converted']
        lost = [l for l in leads if l.status == 'Lost']
        new_leads = [l for l in leads if l.status == 'New']
        future = [l for l in leads if l.status == 'Future']
        future_due = [l for l in future if l.due_date and l.due_date.date() <= now.date()]
        # Initiated = any action taken (not New, Converted, Lost, or Future)
        initiated = [l for l in leads if l.status not in ['New', 'Converted', 'Lost', 'Future']]

        users = User.query.filter_by(active=True).filter(User.role != 'admin').all()
        # Workload — always this month
        wl_filter = 'month'
        wl_from = wl_to = ''
        all_leads_db = Lead.query.all()
        # The per-staff loop below reads j.customer.assigned_to and j.partial_revenues
        # for every job, once per user — eager-load both to avoid a large N+1.
        all_jobs_db = Job.query.options(
            db.joinedload(Job.customer), db.subqueryload(Job.partial_revenues)
        ).all()
        def in_period(dt, f):
            if not dt: return False
            d = dt.date() if hasattr(dt,'date') else dt
            return d.month == now.month and d.year == now.year
        # Targets
        try:
            staff_targets = {t.user_id: t for t in MonthlyTarget.query.filter_by(month=now.month, year=now.year).all()}
            total_monthly_target = sum((t.amount_target or 0) for t in staff_targets.values())
        except:
            staff_targets = {}
            total_monthly_target = 0
        staff_stats = []
        for u in users:
            u_leads = [l for l in all_leads_db if l.assigned_to == u.id and in_period(l.created_at, wl_filter)]
            u_jobs_all = [j for j in all_jobs_db if j.assigned_to == u.id]
            u_jobs = [j for j in u_jobs_all if in_period(j.created_at, wl_filter)]
            u_closed = [j for j in u_jobs_all if j.status == 'Closed']
            # Sales value: credited to the customer's representative (assigned_to on customer)
            u_sales_jobs = [j for j in all_jobs_db if j.customer and j.customer.assigned_to == u.id and in_period(j.created_at, wl_filter)]
            u_sales_closed = [j for j in all_jobs_db if j.customer and j.customer.assigned_to == u.id and j.status == 'Closed']
            u_invoiced = sum((j.amount_invoiced or 0) for j in u_sales_jobs if j.status not in ['Pending Finance Approval'])
            u_closed_val = sum((j.amount_received or 0) for j in u_sales_closed)
            
            # Initiated = leads where staff took action (status NOT "New")
            u_initiated = len([l for l in u_leads if l.status != 'New'])
            
            # New leads = leads with status "New" (not yet contacted)
            u_new_leads = len([l for l in u_leads if l.status == 'New'])
            
            try:
                u_revenue = sum((j.revenue or 0) for j in u_sales_closed)
                # Add partial revenues from non-closed jobs for this staff
                for j in u_sales_jobs:
                    if j.status not in ['Closed', 'Closed - Pending Partner Commission']:
                        u_revenue += sum(pr.amount for pr in j.partial_revenues)
            except:
                u_revenue = 0
            t = staff_targets.get(u.id)
            amount_target = (t.amount_target or 0) if t else 0
            staff_stats.append({
                'name': u.name,
                'role': u.role,
                'leads': len(u_leads),
                'initiated': u_initiated,
                'new_leads': u_new_leads,
                'overdue_leads': len([l for l in u_leads if l.due_date and l.due_date < now and l.status not in ['Converted','Lost']]),
                'conversions': len([l for l in u_leads if l.status == 'Converted']),
                'lost': len([l for l in u_leads if l.status == 'Lost']),
                'future': len([l for l in u_leads if l.status == 'Future']),
                'active_jobs': len([j for j in u_jobs_all if j.status not in ['Done','Closed','Pending Finance Approval']]),
                'overdue_jobs': len([j for j in u_jobs_all if j.due_date and j.due_date < now and j.status not in ['Done','Closed','Pending Finance Approval']]),
                'invoiced': u_invoiced,
                'closed_val': u_closed_val,
                'revenue': u_revenue,
                'amount_target': amount_target,
                'leads_this_month': len(u_leads),
            })
        today_leads = [l for l in all_leads_db if l.created_at and l.created_at.date() == now.date()][:10]

        try:
            all_docs = Document.query.all()
            docs_30 = len([d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30])
            docs_60 = len([d for d in all_docs if d.expiry_date and 30 < (d.expiry_date - now).days <= 60])
            docs_90 = len([d for d in all_docs if d.expiry_date and 60 < (d.expiry_date - now).days <= 90])
            total_docs = len(all_docs)
        except:
            docs_30 = docs_60 = docs_90 = total_docs = 0
        # Birthdays today
        try:
            today = now.date()
            all_customers = Customer.query.filter(Customer.date_of_birth != None).all()
            birthdays_today = [c for c in all_customers if c.date_of_birth and c.date_of_birth.month == today.month and c.date_of_birth.day == today.day]
        except:
            birthdays_today = []
        return render_template('dashboard_admin.html',
                               leads=leads, today_leads=today_leads,
                               birthdays_today=birthdays_today,
                               wl_filter=wl_filter, wl_from=wl_from, wl_to=wl_to,
                               total=total, overdue_leads=overdue_leads,
                               converted=converted, lost=lost, new_leads=new_leads, initiated=initiated,
                               future=future, future_due=future_due,
                               jobs=jobs, active_jobs=active_jobs,
                               overdue_jobs=overdue_jobs, done_jobs=done_jobs,
                               pending_approval=pending_approval,
                               pending_close=pending_close,
                               recent_jobs=recent_jobs,
                               total_invoiced=total_invoiced,
                               total_received=total_received,
                               total_pending=total_pending,
                               completed_value=completed_value,
                               customer_advances=customer_advances,
                               total_revenue=total_revenue,
                               total_partner_pending=total_partner_pending,
                               total_monthly_target=total_monthly_target,
                               staff_stats=staff_stats,
                               docs_30=docs_30, docs_60=docs_60, docs_90=docs_90, total_docs=total_docs,
                               now=now, date_filter=date_filter,
                               from_date=from_date, to_date=to_date,
                               dash_stat_total=dash_stat_total, dash_stat_done_closed=dash_stat_done_closed,
                               dash_stat_overdue=dash_stat_overdue, dash_stat_active=dash_stat_active,
                               dash_stat_pending_finance=dash_stat_pending_finance)

    # ── Staff dashboard ──────────────────────────────────────────────────────
    period = request.args.get('period', 'month')
    all_leads = Lead.query.filter_by(assigned_to=session['user_id']).order_by(Lead.due_date).all()
    if period == 'today':
        leads = [l for l in all_leads if l.created_at and l.created_at.date() == now.date()]
    elif period == 'week':
        week_start = now.date() - timedelta(days=now.weekday())  # Monday
        week_end = week_start + timedelta(days=6)  # Sunday
        leads = [l for l in all_leads if l.created_at and week_start <= l.created_at.date() <= week_end]
    elif period == 'month':
        leads = [l for l in all_leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
    else:
        leads = all_leads
    overdue = [l for l in leads if l.due_date and l.due_date < now and l.status not in ['Converted', 'Lost', 'Future']]
    converted = [l for l in leads if l.status == 'Converted']
    lost = [l for l in leads if l.status == 'Lost']
    new_leads = [l for l in leads if l.status == 'New']
    future = [l for l in leads if l.status == 'Future']
    future_due = [l for l in future if l.due_date and l.due_date.date() <= now.date()]
    initiated = [l for l in leads if l.status not in ['New', 'Converted', 'Lost', 'Future']]
    try:
        my_jobs = Job.query.filter_by(assigned_to=session['user_id']).filter(Job.status.notin_(['Done','Closed'])).order_by(Job.due_date).all()
        pending_approval_jobs = [j for j in my_jobs if j.status == 'Pending Finance Approval']
        overdue_jobs = [j for j in my_jobs if j.due_date and j.due_date < now and j.status != 'Pending Finance Approval']
        active_jobs = [j for j in my_jobs if j.status != 'Pending Finance Approval']
        total_invoiced = sum((j.amount_invoiced or 0) for j in active_jobs)
        total_received = sum((j.amount_received or 0) for j in active_jobs)
        total_pending = total_invoiced - total_received
        done_jobs = Job.query.filter_by(assigned_to=session['user_id'], status='Done').all()
        closed_jobs = Job.query.filter_by(assigned_to=session['user_id']).filter(Job.status.in_(['Closed', 'Closed - Pending Partner Commission'])).all()
        completed_value = sum((j.amount_received or 0) for j in done_jobs)
        try:
            total_revenue = sum((j.revenue or 0) for j in closed_jobs)
        except:
            total_revenue = 0
    except:
        my_jobs = []
        overdue_jobs = []
        total_invoiced = total_received = total_pending = completed_value = total_revenue = 0
    followups = LeadUpdate.query.filter(
        LeadUpdate.staff_name == session['user_name'],
        LeadUpdate.followup_date <= now + timedelta(days=1),
        LeadUpdate.followup_date >= now
    ).all()
    try:
        all_docs = Document.query.all()
        docs_30 = len([d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30])
        docs_60 = len([d for d in all_docs if d.expiry_date and 30 < (d.expiry_date - now).days <= 60])
        docs_90 = len([d for d in all_docs if d.expiry_date and 60 < (d.expiry_date - now).days <= 90])
        total_docs = len(all_docs)
    except:
        docs_30 = docs_60 = docs_90 = total_docs = 0
    try:
        today_date = now.date()
        all_customers_bday = Customer.query.filter(Customer.date_of_birth != None).all()
        birthdays_today = [c for c in all_customers_bday if c.date_of_birth and c.date_of_birth.month == today_date.month and c.date_of_birth.day == today_date.day]
    except:
        birthdays_today = []
    return render_template('dashboard_staff.html', leads=leads, overdue=overdue,
                           birthdays_today=birthdays_today,
                           converted=converted, lost=lost, new_leads=new_leads, initiated=initiated,
                           future=future, future_due=future_due,
                           my_jobs=my_jobs, overdue_jobs=overdue_jobs,
                           total_invoiced=total_invoiced,
                           total_received=total_received,
                           total_pending=total_pending,
                           completed_value=completed_value,
                           total_revenue=total_revenue,
                           docs_30=docs_30, docs_60=docs_60, docs_90=docs_90, total_docs=total_docs,
                           pending_approval_jobs=pending_approval_jobs,
                           followups=followups, now=now, period=period)

@app.route('/leads')
@login_required
def all_leads():
    if session['role'] == 'finance':
        flash('Access denied')
        return redirect(url_for('dashboard'))
    now = now_dubai()
    role = session.get('role')
    user_id = session.get('user_id')

    FILTER_KEYS = ['date', 'status', 'staff', 'search', 'from', 'to', 'due', 'source', 'quality']

    # If reset requested — clear saved filters and redirect clean
    if request.args.get('reset') == '1':
        session.pop('leads_filters', None)
        return redirect(url_for('all_leads'))

    # If any filter param is present in URL — save to session
    if any(request.args.get(k) for k in FILTER_KEYS):
        session['leads_filters'] = {k: request.args.get(k, '') for k in FILTER_KEYS}
        args = request.args
    # If no filter in URL but session has saved filters — restore them
    elif 'leads_filters' in session:
        args = session['leads_filters']
    else:
        args = request.args

    search = (args.get('search') or '').strip().lower()
    is_default = not any(args.get(k) for k in FILTER_KEYS)

    leads = Lead.query.order_by(Lead.due_date).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()

    # For sales: default to their own leads unless staff filter explicitly set
    if role == 'sales' and not args.get('staff'):
        leads = [l for l in leads if l.assigned_to == user_id]

    if search:
        sdigits = ''.join(ch for ch in search if ch.isdigit())
        def _lmatch(l):
            if search in (l.name or '').lower() or search in (l.company or '').lower():
                return True
            # digits-only phone match, so a partial number works despite spaces/+ formatting
            if sdigits:
                for p in (l.phone, getattr(l, 'phone2', None)):
                    if p and sdigits in ''.join(ch for ch in p if ch.isdigit()):
                        return True
            return False
        leads = [l for l in leads if _lmatch(l)]

    if is_default:
        # Default: show this week's leads
        week_start = (now - timedelta(days=now.weekday())).date()
        today = now.date()
        # This week's leads, but exclude Future leads still parked (revisit date not yet due)
        base_leads = [l for l in leads
                      if l.created_at and l.created_at.date() >= week_start
                      and not (l.status == 'Future' and l.due_date and l.due_date.date() > today)]
        # Always resurface Future leads whose revisit date is now due — even older ones
        due_future = [l for l in leads
                      if l.status == 'Future' and l.due_date and l.due_date.date() <= today]
        ids = {l.id for l in base_leads}
        leads = base_leads + [l for l in due_future if l.id not in ids]
        leads.sort(key=lambda l: (l.due_date or now))
    else:
        leads = apply_lead_filters(leads, args, now)

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = 50
    total = len(leads)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated = leads[(page - 1) * per_page: page * per_page]

    sources = Source.query.order_by(Source.name).all()
    bot_lead_ids = _bot_replied_lead_ids(paginated)
    return render_template('all_leads.html', leads=paginated, now=now, users=users,
                           search=search, is_default=is_default,
                           page=page, total_pages=total_pages, total=total,
                           sources=sources, saved_filters=session.get('leads_filters', {}),
                           bot_lead_ids=bot_lead_ids)

@app.route('/leads/export')
@login_required
def export_leads():
    if session['role'] not in ['admin', 'sales']:
        flash('Access denied')
        return redirect(url_for('all_leads'))
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import make_response
    import io
    now = now_dubai()
    leads = Lead.query.order_by(Lead.due_date).all()
    leads = apply_lead_filters(leads, request.args, now)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ['Name', 'Company', 'Phone', 'Phone 2', 'Email', 'Address', 'Source',
               'Service', 'Lead Type', 'Assigned To', 'Due Date', 'Status', 'Lead Quality',
               'Remarks', 'Created']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="133E87", end_color="133E87", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for lead in leads:
        ws.append([
            lead.name, lead.company or '', lead.phone or '', lead.phone2 or '',
            lead.email or '', lead.address or '', lead.source or '', lead.service or '',
            lead.lead_type or '', lead.assignee.name if lead.assignee else '',
            lead.due_date.strftime('%d %b %Y') if lead.due_date else '',
            lead.status or '', lead.genuine or 'Not reviewed', lead.remarks or '',
            lead.created_at.strftime('%d %b %Y') if lead.created_at else '',
        ])
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4
    # Excel column filters (dropdown arrows on the header row)
    ws.auto_filter.ref = ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    quality = (request.args.get('quality') or '').lower()
    suffix = f'_{quality}' if quality in ('genuine', 'junk', 'unreachable', 'unreviewed') else ''
    response.headers['Content-Disposition'] = f'attachment; filename=tahfeel_leads{suffix}_{now.strftime("%Y%m%d")}.xlsx'
    return response

@app.route('/leads/add', methods=['GET', 'POST'])
@login_required
def add_lead():
    now = now_dubai()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        due = request.form.get('due_date')
        lead_date = request.form.get('lead_date')
        # Capture the actual time the lead is entered (combine chosen date with current time)
        if lead_date:
            created_dt = datetime.combine(datetime.strptime(lead_date, '%Y-%m-%d').date(), now_dubai().time())
        else:
            created_dt = now_dubai()
        due_dt = datetime.strptime(due, '%Y-%m-%d') if due else created_dt + timedelta(days=1)
        lead = Lead(
            name=request.form['name'],
            company=request.form.get('company'),
            phone=request.form.get('phone'),
            phone2=request.form.get('phone2'),
            email=request.form.get('email'),
            address=request.form.get('address'),
            source=request.form.get('source'),
            service=request.form.get('service'),
            representative=session['user_name'],
            lead_type=request.form.get('lead_type', 'New'),
            assigned_to=int(request.form['assigned_to']) if request.form.get('assigned_to') else None,
            due_date=due_dt,
            remarks=request.form.get('remarks'),
            campaign=request.form.get('campaign') or None,
            created_at=created_dt
        )
        db.session.add(lead)
        db.session.commit()
        # Auto-greet on WhatsApp (approved template). Only for this single manual add —
        # never bulk import. No-ops unless WA_AUTO_WELCOME is on. Never breaks lead add.
        try:
            from whatsapp_webhook import notify_new_lead
            notify_new_lead(lead)
        except Exception as e:
            print(f'[WA] manual-lead greet skipped: {e}')
        flash('Lead added successfully')
        return redirect(url_for('all_leads'))
    campaigns = Campaign.query.order_by(Campaign.name).all()
    tomorrow = (now_dubai() + timedelta(days=1)).strftime('%Y-%m-%d')
    return render_template('add_lead.html', users=users, services=services, sources=sources, campaigns=campaigns, now=now, tomorrow=tomorrow)

@app.route('/leads/<int:lead_id>', methods=['GET', 'POST'])
@login_required
def lead_detail(lead_id):
    now = now_dubai()
    lead = Lead.query.get_or_404(lead_id)
    if request.method == 'POST':
        stage = request.form['stage']
        activity_type = request.form.get('activity_type') or None
        remark = request.form['remark']
        followup = request.form.get('followup_date')
        followup_dt = datetime.strptime(followup, '%Y-%m-%d') if followup else None
        # A "Future" lead must have a revisit date — that's what brings it back later.
        if stage == 'Future' and not followup_dt:
            flash('Please pick a future revisit date for a Future lead.')
            return redirect(url_for('lead_detail', lead_id=lead_id))
        # ── Lead quality is mandatory once a real contact is made ──
        # A "positive action" = any logged activity that isn't a no-answer or a plain
        # note, OR moving the lead forward. You can't judge quality on a no-answer, so
        # those (and blank activity) can still be saved without a quality tag.
        new_quality = request.form.get('genuine') or None
        junk_reason_q = request.form.get('junk_reason') or None
        _act = (activity_type or '').lower()
        is_positive_action = bool(activity_type) and 'no answer' not in _act and _act != 'note'
        if (lead.genuine is None and new_quality is None
                and (is_positive_action or stage in ('Qualified', 'Proposal', 'Converted', 'Lost'))):
            flash('Please mark the lead quality (Genuine / Junk / Unreachable) — required once you have reached the customer.')
            return redirect(url_for('lead_detail', lead_id=lead_id))
        update = LeadUpdate(
            lead_id=lead.id, stage=stage, activity_type=activity_type, remark=remark,
            staff_name=session['user_name'], followup_date=followup_dt,
            lost_reason=request.form.get('lost_reason'),
            future_potential=request.form.get('future_potential')
        )
        lead.status = stage
        # Apply a quality tag if one was chosen on the update form
        if new_quality in ('Genuine', 'Junk', 'Unreachable'):
            lead.genuine = new_quality
            lead.junk_reason = junk_reason_q if new_quality in ('Junk', 'Unreachable') else None
        # Attempt counter (info only) + first-contact timestamp, driven by the activity
        act = (activity_type or '').lower()
        if 'no answer' in act or 'not reach' in act:
            lead.attempts = (lead.attempts or 0) + 1
        if ('connected' in act or 'meeting' in act) and not lead.first_contacted_at:
            lead.first_contacted_at = now_dubai()
        # Update lead's due_date with new followup date
        if followup_dt:
            lead.due_date = followup_dt
        if request.form.get('customer_story'):
            lead.customer_story = request.form.get('customer_story')
        db.session.add(update)
        db.session.commit()
        # Send the "good lead" signal to Meta when quality is set to Genuine (dormant
        # unless CAPI is configured + enabled; safe no-op otherwise).
        if new_quality == 'Genuine':
            capi_send_lead_quality(lead)
        flash('Update saved')
        return redirect(url_for('lead_detail', lead_id=lead_id))
    # WhatsApp chat link: show a "View WhatsApp chat" button only when this lead's
    # number actually has a logged conversation.
    wa_id = None
    wa_has_chat = False
    bot_replied = False
    try:
        from whatsapp_webhook import normalize_phone
        wa_id = normalize_phone(lead.phone) if lead.phone else None
        if wa_id:
            wa_has_chat = WhatsAppMessage.query.filter_by(wa_id=wa_id).first() is not None
            bot_replied = WhatsAppMessage.query.filter_by(wa_id=wa_id, direction='in').first() is not None
    except Exception:
        pass
    return render_template('lead_detail.html', lead=lead, now=now,
                           wa_id=wa_id, wa_has_chat=wa_has_chat, bot_replied=bot_replied)

@app.route('/leads/<int:lead_id>/quality', methods=['POST'])
@login_required
def set_lead_quality(lead_id):
    """Manually mark lead quality (Genuine / Junk / Unreachable). Never automatic."""
    lead = Lead.query.get_or_404(lead_id)
    value = request.form.get('genuine')
    if value not in ('Genuine', 'Junk', 'Unreachable', ''):
        flash('Invalid quality value')
        return redirect(url_for('lead_detail', lead_id=lead_id))
    lead.genuine = value or None
    lead.junk_reason = request.form.get('junk_reason') or None
    # Record the judgment in the lead history for the audit trail
    note = LeadUpdate(lead_id=lead.id, stage=lead.status, activity_type='Quality',
                      remark=f'Marked {value or "unreviewed"}' + (f' — {lead.junk_reason}' if lead.junk_reason else ''),
                      staff_name=session['user_name'])
    db.session.add(note)
    db.session.commit()
    if value == 'Genuine':
        capi_send_lead_quality(lead)
    flash(f'Lead marked {value or "unreviewed"}')
    return redirect(url_for('lead_detail', lead_id=lead_id))

# ── Enquiries — quick question/callback capture (lighter than a lead) ──────────
@app.route('/enquiries')
@login_required
def enquiries():
    if session.get('role') == 'finance':
        flash('Access denied')
        return redirect(url_for('dashboard'))
    now = now_dubai()
    flt = (request.args.get('status') or 'Open').strip()
    q = Enquiry.query
    if flt in ('Open', 'Resolved'):
        q = q.filter(Enquiry.status == flt)
    if session.get('role') == 'sales' and not request.args.get('all'):
        q = q.filter(Enquiry.assigned_to == session.get('user_id'))
    items = q.order_by(Enquiry.created_at.desc()).all()
    all_e = Enquiry.query.all()
    resolved = [e for e in all_e if e.status == 'Resolved' and e.resolved_at]
    avg_res = _fmt_duration(sum((e.resolved_at - e.created_at).total_seconds() for e in resolved) / len(resolved)) if resolved else '—'
    stats = {
        'total': len(all_e),
        'open': sum(1 for e in all_e if e.status == 'Open'),
        'resolved': len(resolved),
        'avg_res': avg_res,
        'due': sum(1 for e in all_e if e.status == 'Open' and e.remind_date and e.remind_date <= now.date()),
    }
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales', 'operations', 'admin'])).all()
    tomorrow = (now + timedelta(days=1)).strftime('%Y-%m-%d')
    from collections import Counter
    by_service = Counter((e.service or 'Unspecified') for e in all_e).most_common()
    by_staff = Counter((e.assignee.name if e.assignee else 'Unassigned') for e in all_e).most_common()
    return render_template('enquiries.html', items=items, stats=stats, flt=flt, users=users, now=now,
                           tomorrow=tomorrow, by_service=by_service, by_staff=by_staff)

@app.route('/enquiries/add', methods=['POST'])
@login_required
def enquiry_add():
    text = (request.form.get('enquiry') or '').strip()
    if not text:
        flash('Please type the enquiry.')
        return redirect(url_for('enquiries'))
    remind = request.form.get('remind_date')
    try:
        rd = datetime.strptime(remind, '%Y-%m-%d').date() if remind else None
    except ValueError:
        rd = None
    assigned = request.form.get('assigned_to')
    e = Enquiry(name=(request.form.get('name') or '').strip() or None,
                phone=(request.form.get('phone') or '').strip() or None,
                enquiry=text, service=(request.form.get('service') or '').strip() or None,
                assigned_to=int(assigned) if assigned else session.get('user_id'),
                remind_date=rd, created_by=session.get('user_name'))
    db.session.add(e)
    db.session.commit()
    flash('Enquiry logged.')
    return redirect(url_for('enquiries'))

@app.route('/enquiries/<int:eid>/resolve', methods=['POST'])
@login_required
def enquiry_resolve(eid):
    e = Enquiry.query.get_or_404(eid)
    if e.status != 'Resolved':
        e.status, e.resolved_at = 'Resolved', now_dubai()
    else:
        e.status, e.resolved_at = 'Open', None
    db.session.commit()
    return redirect(request.referrer or url_for('enquiries'))

@app.route('/enquiries/<int:eid>/convert-lead', methods=['POST'])
@login_required
def enquiry_convert_lead(eid):
    e = Enquiry.query.get_or_404(eid)
    if e.converted_lead_id:
        flash('This enquiry was already converted to a lead.')
        return redirect(url_for('lead_detail', lead_id=e.converted_lead_id))
    lead = Lead(name=e.name or 'Enquiry', phone=e.phone, source='Enquiry',
                service=e.service, status='New',
                assigned_to=e.assigned_to or session.get('user_id'),
                remarks=e.enquiry)
    db.session.add(lead)
    db.session.commit()
    e.converted_lead_id, e.status, e.resolved_at = lead.id, 'Resolved', now_dubai()
    db.session.commit()
    flash('Converted to a lead.')
    return redirect(url_for('lead_detail', lead_id=lead.id))

@app.route('/enquiries/export')
@login_required
def enquiries_export():
    if session.get('role') == 'finance':
        flash('Access denied')
        return redirect(url_for('dashboard'))
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook(); ws = wb.active; ws.title = 'Enquiries'
    headers = ['#', 'Date', 'Name', 'Phone', 'Enquiry', 'Service', 'Assigned To',
               'Remind Date', 'Status', 'Resolved On', 'Time to Resolve', 'Converted to Lead #']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    rows = Enquiry.query.order_by(Enquiry.created_at.desc()).all()
    for i, e in enumerate(rows, 1):
        rt = _fmt_duration((e.resolved_at - e.created_at).total_seconds()) if e.resolved_at else ''
        ws.append([
            i, e.created_at.strftime('%d/%m/%Y %H:%M') if e.created_at else '',
            e.name or '', e.phone or '', e.enquiry or '', e.service or '',
            e.assignee.name if e.assignee else '', e.remind_date.strftime('%d/%m/%Y') if e.remind_date else '',
            e.status or '', e.resolved_at.strftime('%d/%m/%Y') if e.resolved_at else '', rt,
            e.converted_lead_id or '',
        ])
    for i, w in enumerate([4, 16, 20, 16, 50, 18, 18, 14, 12, 14, 16, 16], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f'tahfeel_enquiries_{now_dubai().strftime("%Y%m%d")}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/enquiries/<int:eid>/delete', methods=['POST'])
@login_required
def enquiry_delete(eid):
    e = Enquiry.query.get_or_404(eid)
    db.session.delete(e)
    db.session.commit()
    flash('Enquiry deleted.')
    return redirect(url_for('enquiries'))

@app.route('/leads/import', methods=['GET', 'POST'])
@login_required
def import_leads():
    if session.get('role') not in ['admin', 'sales']:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx file')
            return redirect(url_for('import_leads'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            errors = []
            all_staff = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
            staff_map = {u.name.strip().lower(): u.id for u in all_staff}
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0] if len(row) > 0 else None
                company = row[1] if len(row) > 1 else None
                phone = str(row[2]) if len(row) > 2 and row[2] else None
                email = row[3] if len(row) > 3 else None
                address = row[4] if len(row) > 4 else None
                source = row[5] if len(row) > 5 else None
                service = row[6] if len(row) > 6 else None
                lead_type = row[7] if len(row) > 7 else 'New'
                remarks = row[8] if len(row) > 8 else None
                assigned_name = str(row[9]).strip() if len(row) > 9 and row[9] else None
                lead_date_str = str(row[10]).strip() if len(row) > 10 and row[10] else None
                campaign = str(row[11]).strip() if len(row) > 11 and row[11] else None
                assigned_id = None
                if assigned_name:
                    assigned_id = staff_map.get(assigned_name.lower())
                    if not assigned_id:
                        errors.append(f'Row {i}: Staff "{assigned_name}" not found — lead imported unassigned')
                if not name:
                    errors.append(f'Row {i}: Name is required — skipped')
                    continue
                if not phone:
                    errors.append(f'Row {i}: Phone is required — skipped')
                    continue
                try:
                    if isinstance(lead_date_str, datetime):
                        created_dt = lead_date_str
                    elif lead_date_str:
                        created_dt = datetime.strptime(str(lead_date_str).split(' ')[0].split('T')[0], '%Y-%m-%d')
                    else:
                        created_dt = now_dubai()
                except:
                    created_dt = now_dubai()
                lead = Lead(
                    name=str(name), company=str(company) if company else None,
                    phone=str(phone), email=str(email) if email else None,
                    address=str(address) if address else None,
                    source=str(source) if source else None,
                    service=str(service) if service else None,
                    lead_type=str(lead_type) if lead_type else 'New',
                    remarks=str(remarks) if remarks else None,
                    representative=session['user_name'],
                    assigned_to=assigned_id,
                    campaign=campaign,
                    created_at=created_dt,
                    due_date=created_dt + timedelta(days=1)
                )
                db.session.add(lead)
                count += 1
            db.session.commit()
            if errors:
                flash(f'Imported {count} leads. Notes: ' + ' | '.join(errors))
            else:
                flash(f'Successfully imported {count} leads!')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}')
            return redirect(url_for('import_leads'))
    return render_template('import_leads.html')

@app.route('/leads/template')
@login_required
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from flask import make_response
    import io
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    staff = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    campaigns = Campaign.query.order_by(Campaign.name).all()
    headers = ['Name*', 'Company', 'Phone*', 'Email', 'Address',
               'Source', 'Service', 'Lead Type', 'Remarks', 'Assigned To', 'Lead Date', 'Campaign']
    ws.append(headers)
    ws.append(['John Smith', 'ABC Trading LLC', '+971501234567',
               'john@abc.ae', 'Dubai',
               sources[0].name if sources else 'WhatsApp',
               services[0].name if services else 'Trade License',
               'New', 'Interested in mainland license',
               staff[0].name if staff else '', '2026-04-16',
               campaigns[0].name if campaigns else ''])
    ws.append(['Sara Ahmed', '', '+971509876543', '', 'Sharjah',
               sources[1].name if len(sources) > 1 else '',
               services[1].name if len(services) > 1 else '',
               'New', '',
               staff[1].name if len(staff) > 1 else '', '2026-04-16', ''])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="133E87", end_color="133E87", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4
    ref = wb.create_sheet(title="Reference")
    ref['A1'] = 'Services'
    for i, s in enumerate(services, start=2):
        ref.cell(row=i, column=1, value=s.name)
    ref['B1'] = 'Sources'
    for i, s in enumerate(sources, start=2):
        ref.cell(row=i, column=2, value=s.name)
    ref['C1'] = 'Staff'
    for i, u in enumerate(staff, start=2):
        ref.cell(row=i, column=3, value=u.name)
    ref['D1'] = 'Lead Type'
    ref['D2'] = 'New'
    ref['D3'] = 'Old Follow-up'
    ref['E1'] = 'Campaigns'
    for i, c in enumerate(campaigns, start=2):
        ref.cell(row=i, column=5, value=c.name)
    ref.sheet_state = 'hidden'
    service_count = len(services) + 1
    source_count = len(sources) + 1
    staff_count = len(staff) + 1
    campaign_count = len(campaigns) + 1
    dv_source = DataValidation(type="list", formula1=f"Reference!$B$2:$B${source_count}", allow_blank=True, showDropDown=False)
    dv_source.sqref = "F2:F1000"
    ws.add_data_validation(dv_source)
    dv_service = DataValidation(type="list", formula1=f"Reference!$A$2:$A${service_count}", allow_blank=True, showDropDown=False)
    dv_service.sqref = "G2:G1000"
    ws.add_data_validation(dv_service)
    dv_type = DataValidation(type="list", formula1="Reference!$D$2:$D$3", allow_blank=True, showDropDown=False)
    dv_type.sqref = "H2:H1000"
    ws.add_data_validation(dv_type)
    dv_staff = DataValidation(type="list", formula1=f"Reference!$C$2:$C${staff_count}", allow_blank=True, showDropDown=False)
    dv_staff.sqref = "J2:J1000"
    ws.add_data_validation(dv_staff)
    if campaigns:
        dv_campaign = DataValidation(type="list", formula1=f"Reference!$E$2:$E${campaign_count}", allow_blank=True, showDropDown=False)
        dv_campaign.sqref = "L2:L1000"
        ws.add_data_validation(dv_campaign)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=tahfeel_leads_template.xlsx'
    return response

@app.route('/leads/<int:lead_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_lead(lead_id):
    now = now_dubai()
    lead = Lead.query.get_or_404(lead_id)
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        lead.name = request.form['name']
        lead.company = request.form.get('company')
        lead.phone = request.form.get('phone')
        lead.phone2 = request.form.get('phone2')
        lead.email = request.form.get('email')
        lead.address = request.form.get('address')
        lead.source = request.form.get('source')
        lead.service = request.form.get('service')
        lead.campaign = request.form.get('campaign') or None
        lead.lead_type = request.form.get('lead_type', 'New')
        lead.remarks = request.form.get('remarks')
        assigned = request.form.get('assigned_to')
        lead.assigned_to = int(assigned) if assigned else None
        due = request.form.get('due_date')
        if due:
            lead.due_date = datetime.strptime(due, '%Y-%m-%d')
        db.session.commit()
        flash('Lead updated successfully')
        return redirect(url_for('lead_detail', lead_id=lead_id))
    campaigns = Campaign.query.order_by(Campaign.name).all()
    return render_template('edit_lead.html', lead=lead, users=users, services=services, sources=sources, campaigns=campaigns, now=now)

@app.route('/leads/<int:lead_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_lead(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    
    # Check if lead is converted (has associated customer)
    if lead.status == 'Converted':
        customer = Customer.query.filter_by(lead_id=lead_id).first()
        if customer:
            flash('Cannot delete converted leads. This lead has been converted to a customer and has associated records.', 'error')
            return redirect(url_for('lead_detail', lead_id=lead_id))
    
    LeadUpdate.query.filter_by(lead_id=lead_id).delete()
    # Unlink (don't delete) records that reference this lead but can outlive it —
    # otherwise the DB foreign key blocks deletion with an error.
    Customer.query.filter_by(lead_id=lead_id).update({'lead_id': None})
    WhatsAppMessage.query.filter_by(lead_id=lead_id).update({'lead_id': None})
    Task.query.filter_by(lead_id=lead_id).update({'lead_id': None})
    db.session.delete(lead)
    db.session.commit()
    flash('Lead deleted successfully')
    ref = request.referrer
    if ref and '/leads' in ref:
        return redirect(ref)
    return redirect(url_for('all_leads'))

@app.route('/leads/bulk-delete', methods=['POST'])
@login_required
@admin_required
def bulk_delete_leads():
    ids = request.form.getlist('lead_ids')
    if not ids:
        flash('No leads selected', 'error')
        return redirect(url_for('all_leads'))
    count = 0
    skipped = 0
    for lead_id in ids:
        lead = Lead.query.get(int(lead_id))
        if lead:
            # Keep converted leads that still own a real customer record
            if lead.status == 'Converted' and Customer.query.filter_by(lead_id=lead.id).first():
                skipped += 1
                continue
            # Unlink everything that references this lead so the FK doesn't block delete
            Customer.query.filter_by(lead_id=lead.id).update({'lead_id': None})
            LeadUpdate.query.filter_by(lead_id=lead.id).delete()
            WhatsAppMessage.query.filter_by(lead_id=lead.id).update({'lead_id': None})
            Task.query.filter_by(lead_id=lead.id).update({'lead_id': None})
            db.session.delete(lead)
            count += 1
    db.session.commit()
    if count > 0:
        flash(f'{count} lead(s) deleted successfully')
    if skipped > 0:
        flash(f'{skipped} converted lead(s) skipped (cannot delete converted leads)', 'warning')
    return redirect(url_for('all_leads'))

@app.route('/admin', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_panel():
    users = User.query.order_by(User.name).all()
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    campaigns = Campaign.query.order_by(Campaign.name).all()
    job_types = ServiceType.query.order_by(ServiceType.name).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    partners = Partner.query.order_by(Partner.name).all()
    wa_auto_welcome = automation_on('wa_auto_welcome')
    autos = {k: automation_on(k) for k in AUTOMATION_DEFAULTS}
    runs = {k: get_setting(f'run_{k}') for k in ('birthday', 'expiry_wa', 'expiry_email', 'monthly_report')}
    capi = {
        'enabled': get_setting('capi_enabled', 'off') == 'on',
        'token_set': bool(get_setting('capi_token', '')),
        'dataset_id': get_setting('capi_dataset_id', '') or '',
        'event_name': get_setting('capi_event_name', 'Qualified') or 'Qualified',
        'test_code': get_setting('capi_test_code', '') or '',
        'last_run': get_setting('run_capi', ''),
    }
    # Flat master list of common sub-tasks (pick from when creating a task)
    subtask_list = SubTaskTemplate.query.order_by(SubTaskTemplate.sort_order, SubTaskTemplate.id).all()
    return render_template('admin_panel.html', users=users, services=services,
                           sources=sources, campaigns=campaigns, job_types=job_types, doc_types=doc_types, partners=partners,
                           wa_auto_welcome=wa_auto_welcome, autos=autos, runs=runs, capi=capi,
                           subtask_list=subtask_list)

@app.route('/admin/whatsapp-settings', methods=['POST'])
@login_required
@admin_required
def admin_whatsapp_settings():
    """Save WhatsApp settings (one-time config) — currently the auto-welcome toggle."""
    set_setting('wa_auto_welcome', 'on' if request.form.get('wa_auto_welcome') == 'on' else 'off')
    flash('WhatsApp settings saved.')
    return redirect(url_for('admin_panel'))

@app.route('/admin/automations', methods=['POST'])
@login_required
@admin_required
def admin_automations():
    """Save the on/off state of every automation from the Automations panel.
    A checkbox that is unticked simply isn't in the form, so it becomes 'off'."""
    for key in AUTOMATION_DEFAULTS:
        set_setting(key, 'on' if request.form.get(key) == 'on' else 'off')
    flash('Automation settings saved.')
    return redirect(url_for('admin_panel') + '#automations')

@app.route('/admin/subtask-list/add', methods=['POST'])
@login_required
@admin_required
def admin_subtask_list_add():
    """Add one or more sub-task names to the master list (textarea, one per line)."""
    names = [ln.strip() for ln in (request.form.get('steps') or '').splitlines() if ln.strip()]
    existing = {t.title for t in SubTaskTemplate.query.all()}
    base = SubTaskTemplate.query.count()
    added = 0
    for name in names:
        if name not in existing:
            db.session.add(SubTaskTemplate(job_type='list', title=name[:200], sort_order=base + added))
            existing.add(name)
            added += 1
    db.session.commit()
    flash(f'Added {added} sub-task(s) to the list.')
    return redirect(url_for('admin_panel') + '#step-templates')

@app.route('/admin/subtask-list/<int:item_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_subtask_list_delete(item_id):
    t = SubTaskTemplate.query.get_or_404(item_id)
    db.session.delete(t)
    db.session.commit()
    return redirect(url_for('admin_panel') + '#step-templates')

@app.route('/admin/capi-settings', methods=['POST'])
@login_required
@admin_required
def admin_capi_settings():
    """Save Meta Conversions API config. The token is only overwritten when a new
    one is typed, so saving other fields never wipes an existing token."""
    set_setting('capi_enabled', 'on' if request.form.get('capi_enabled') == 'on' else 'off')
    set_setting('capi_dataset_id', (request.form.get('capi_dataset_id') or '').strip())
    set_setting('capi_event_name', (request.form.get('capi_event_name') or 'Qualified').strip() or 'Qualified')
    set_setting('capi_test_code', (request.form.get('capi_test_code') or '').strip())
    new_token = (request.form.get('capi_token') or '').strip()
    if new_token:
        set_setting('capi_token', new_token)
    flash('Meta CAPI settings saved.')
    return redirect(url_for('admin_panel') + '#capi')

@app.route('/admin/capi-test', methods=['POST'])
@login_required
@admin_required
def admin_capi_test():
    """Fire one test event to Meta using the saved token + dataset, so an admin can
    confirm the connection in Events Manager -> Test Events without a real lead."""
    token = get_setting('capi_token', '')
    dataset = get_setting('capi_dataset_id', '')
    if not token or not dataset:
        flash('Add the access token and Dataset/Pixel ID (and Save) first.')
        return redirect(url_for('admin_panel') + '#capi')
    event_name = get_setting('capi_event_name', 'Qualified') or 'Qualified'
    test_code = get_setting('capi_test_code', '')
    import time, hashlib, requests
    payload = {
        'data': [{
            'event_name': event_name,
            'event_time': int(time.time()),
            'action_source': 'system_generated',
            'user_data': {'ph': [hashlib.sha256('971500000000'.encode()).hexdigest()]},
            'custom_data': {'lead_event_source': 'crm', 'crm': 'Tahfeel CRM', 'test': True},
        }],
        'access_token': token,
    }
    if test_code:
        payload['test_event_code'] = test_code
    try:
        r = requests.post(f'https://graph.facebook.com/v19.0/{dataset}/events', json=payload, timeout=10)
        if r.status_code == 200:
            flash(f'✅ Test event sent — HTTP 200. Check Events Manager → Test Events. {r.text[:200]}')
        else:
            flash(f'⚠️ Meta returned HTTP {r.status_code}: {r.text[:250]}')
    except Exception as e:
        flash(f'Test failed: {e}')
    return redirect(url_for('admin_panel') + '#capi')


# ── Marketing-Ext: read-only lead report for the external marketing agency ──
def _mask_phone(p):
    if not p:
        return '—'
    d = ''.join(c for c in p if c.isdigit())
    if len(d) < 2:
        return '••••'
    prefix = ('+971 ' + d[3:5]) if (d.startswith('971') and len(d) >= 5) else d[:2]
    return prefix + ' ••• ••' + d[-2:]

def _fmt_duration(seconds):
    if seconds is None:
        return '—'
    s = max(int(seconds), 0)
    d, rem = divmod(s, 86400); h, rem = divmod(rem, 3600); m, _ = divmod(rem, 60)
    if d:
        return f'{d}d {h}h'
    if h:
        return f'{h}h {m}m'
    return f'{m}m'

def _staff_remarks(lead):
    """Only staff-typed update remarks (exclude the auto Meta note + system 'Marked …'), newest first."""
    out = []
    for u in lead.updates:  # ordered newest-first by the model relationship
        rm = (u.remark or '').strip()
        if not rm or rm.lower().startswith(('auto-received', 'marked ')):
            continue
        out.append({'when': u.created_at, 'text': rm})
    return out

def _attempt_count(lead):
    """Total real contact attempts = every logged outreach activity (call connected,
    call no-answer, WhatsApp sent, email sent, quote sent, meeting). Excludes plain
    'Note' and blank rows. Counts from the full update history, so it is correct
    retroactively for older leads too."""
    return sum(1 for u in lead.updates
               if (u.activity_type or '').strip()
               and (u.activity_type or '').strip().lower() not in ('note', 'quality'))

def _inbound_wa_ids():
    """Normalized phone numbers that sent us at least one INBOUND WhatsApp message
    (i.e. the customer actually replied on the bot)."""
    try:
        rows = db.session.query(WhatsAppMessage.wa_id).filter_by(direction='in').distinct().all()
        return {r[0] for r in rows if r[0]}
    except Exception:
        return set()

def _bot_replied_lead_ids(leads):
    """Subset of the given leads' ids whose customer replied on WhatsApp (inbound msg)."""
    try:
        from whatsapp_webhook import normalize_phone
    except Exception:
        return set()
    inbound = _inbound_wa_ids()
    return {l.id for l in leads if l.phone and normalize_phone(l.phone) in inbound}

def _marketing_leads():
    """Leads for the Marketing-Ext report: Meta-only + per-user date floor + UI filters.
    Returns (leads, floor, filters_dict)."""
    from sqlalchemy import or_
    user = User.query.get(session['user_id'])
    floor = user.report_from if (user and user.role == 'marketing') else None
    f = {k: (request.args.get(k) or '').strip() for k in ('from', 'to', 'q', 'stage', 'quality', 'src')}
    # Meta ads only: leads that came from Meta (have a meta_lead_id or a Meta* source)
    q = Lead.query.filter(or_(Lead.meta_lead_id.isnot(None), Lead.source.like('Meta%')))
    if floor:
        q = q.filter(Lead.created_at >= datetime.combine(floor, datetime.min.time()))
    try:
        if f['from']:
            q = q.filter(Lead.created_at >= datetime.strptime(f['from'], '%Y-%m-%d'))
        if f['to']:
            q = q.filter(Lead.created_at < datetime.strptime(f['to'], '%Y-%m-%d') + timedelta(days=1))
    except ValueError:
        pass
    if f['q']:
        q = q.filter(Lead.name.ilike('%' + f['q'] + '%'))
    if f['stage']:
        q = q.filter(Lead.status == f['stage'])
    if f['quality']:
        q = q.filter(Lead.genuine == f['quality'])
    if f['src']:
        q = q.filter(Lead.sub_source == f['src'])
    return q.order_by(Lead.created_at.desc()).all(), floor, f

@app.route('/marketing-report')
@login_required
def marketing_report():
    if session.get('role') not in ('marketing', 'admin'):
        flash('Access denied.')
        return redirect(url_for('dashboard'))
    from sqlalchemy import or_
    from collections import Counter
    from whatsapp_webhook import normalize_phone
    leads, floor, f = _marketing_leads()
    bot_ids = _inbound_wa_ids()
    rows = []
    for l in leads:
        resp = (l.first_contacted_at - l.created_at).total_seconds() if (l.first_contacted_at and l.created_at) else None
        rows.append({
            'date': l.created_at, 'name': l.name or '—', 'phone': _mask_phone(l.phone),
            'source': l.sub_source or l.source or '—', 'stage': l.status or 'New',
            'genuine': l.genuine or '', 'junk_reason': l.junk_reason or '',
            'attempts': _attempt_count(l), 'remarks': _staff_remarks(l),
            'campaign': l.campaign or '—', 'response': _fmt_duration(resp),
            'bot_replied': bool(l.phone and normalize_phone(l.phone) in bot_ids),
        })
    n = len(leads)
    pct = lambda x: round(100 * x / n) if n else 0
    # ── Lead breakdown by stage — mutually EXCLUSIVE buckets that TALLY to the total ──
    sc = Counter((l.status or 'New') for l in leads)
    b_new        = sc.get('New', 0)                                                    # not yet initiated
    b_processing = sum(v for k, v in sc.items() if k in ('Contacted', 'Qualified', 'Proposal'))  # being worked
    b_converted  = sc.get('Converted', 0)
    b_lost       = sc.get('Lost', 0)
    b_future     = sc.get('Future', 0)
    b_other      = n - (b_new + b_processing + b_converted + b_lost + b_future)         # any legacy/unknown status
    breakdown = {
        'total': n, 'new': b_new, 'processing': b_processing, 'converted': b_converted,
        'lost': b_lost, 'future': b_future, 'other': b_other,
        'p_new': pct(b_new), 'p_processing': pct(b_processing), 'p_converted': pct(b_converted),
        'p_lost': pct(b_lost), 'p_future': pct(b_future), 'p_other': pct(b_other),
    }
    funnel = {'rate': pct(b_converted)}
    # ── Lead quality — also tallies to the total (includes "Not reviewed") ──
    genuine      = sum(1 for l in leads if l.genuine == 'Genuine')
    unreachable  = sum(1 for l in leads if l.genuine == 'Unreachable')
    junk         = sum(1 for l in leads if l.genuine == 'Junk')
    not_reviewed = n - (genuine + unreachable + junk)
    junk_reasons = sorted(Counter((l.junk_reason or 'Unspecified') for l in leads if l.genuine == 'Junk').items(),
                          key=lambda x: -x[1])
    quality = {'genuine': genuine, 'unreachable': unreachable, 'junk': junk, 'not_reviewed': not_reviewed,
               'junk_reasons': junk_reasons, 'total': n,
               'p_genuine': pct(genuine), 'p_junk': pct(junk),
               'p_unreachable': pct(unreachable), 'p_not_reviewed': pct(not_reviewed)}
    resp_secs = [(l.first_contacted_at - l.created_at).total_seconds()
                 for l in leads if l.first_contacted_at and l.created_at]
    avg_response = _fmt_duration(sum(resp_secs) / len(resp_secs)) if resp_secs else '—'
    # ── Response-time buckets — how fast leads are actually reached (tallies to total) ──
    resp_buckets = {'lt1h': 0, 'h1_4': 0, 'h4_24': 0, 'd1_3': 0, 'gt3d': 0, 'not_reached': 0}
    for l in leads:
        if l.first_contacted_at and l.created_at:
            hrs = (l.first_contacted_at - l.created_at).total_seconds() / 3600
            if hrs < 1:     resp_buckets['lt1h'] += 1
            elif hrs < 4:   resp_buckets['h1_4'] += 1
            elif hrs < 24:  resp_buckets['h4_24'] += 1
            elif hrs < 72:  resp_buckets['d1_3'] += 1
            else:           resp_buckets['gt3d'] += 1
        else:
            resp_buckets['not_reached'] += 1
    counts = {'total': n, 'processing': b_processing, 'converted': b_converted, 'genuine': genuine}
    # ── Daily leads received — a bar per day over the filtered range (capped to 31 bars) ──
    day_counts = Counter(l.created_at.date() for l in leads if l.created_at)
    daily = []
    if day_counts:
        dmax, dmin = max(day_counts), min(day_counts)
        if (dmax - dmin).days > 30:
            dmin = dmax - timedelta(days=30)
        day = dmin
        while day <= dmax:
            daily.append({'date': day, 'count': day_counts.get(day, 0)})
            day += timedelta(days=1)
    daily_max = max((d['count'] for d in daily), default=0)
    daily_avg = round(sum(d['count'] for d in daily) / len(daily), 1) if daily else 0
    # ── By source (fb / ig / …): volume + quality + conversions ──
    src_stats_map = {}
    for l in leads:
        s = l.sub_source or l.source or '—'
        d = src_stats_map.setdefault(s, {'total': 0, 'genuine': 0, 'converted': 0})
        d['total'] += 1
        if l.genuine == 'Genuine':
            d['genuine'] += 1
        if l.status == 'Converted':
            d['converted'] += 1
    src_stats = sorted(src_stats_map.items(), key=lambda kv: -kv[1]['total'])
    src_options = sorted({r[0] for r in db.session.query(Lead.sub_source).filter(
        or_(Lead.meta_lead_id.isnot(None), Lead.source.like('Meta%'))).all() if r[0]})
    # ── Paginate the table only (the analytics above use ALL filtered leads) ──
    per_page = 50
    total_rows = len(rows)
    total_pages = max(1, (total_rows + per_page - 1) // per_page)
    try:
        page = int(request.args.get('page', 1))
    except (TypeError, ValueError):
        page = 1
    page = max(1, min(page, total_pages))
    page_rows = rows[(page - 1) * per_page: page * per_page]
    return render_template('marketing_report.html', rows=page_rows, counts=counts, funnel=funnel, breakdown=breakdown,
                           quality=quality, avg_response=avg_response, floor=floor, f=f, src_options=src_options,
                           daily=daily, daily_max=daily_max, daily_avg=daily_avg, src_stats=src_stats,
                           resp=resp_buckets,
                           page=page, total_pages=total_pages, total_rows=total_rows, per_page=per_page)

@app.route('/marketing-report/export')
@login_required
def marketing_export():
    if session.get('role') not in ('marketing', 'admin'):
        flash('Access denied.')
        return redirect(url_for('dashboard'))
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from whatsapp_webhook import normalize_phone
    leads, floor, f = _marketing_leads()
    bot_ids = _inbound_wa_ids()
    wb = Workbook(); ws = wb.active; ws.title = 'Leads'
    headers = ['Date', 'Name', 'Phone', 'Source', 'Stage', 'Lead Quality', 'Attempts', 'Bot Reply', 'Remarks (staff updates, with time)']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    for l in leads:
        quality = l.genuine or ''
        if l.genuine == 'Junk' and l.junk_reason:
            quality = f'Junk — {l.junk_reason}'
        # Full staff-update history, each line timestamped (newest first)
        history = '\n'.join(
            f"{r['when'].strftime('%d %b %Y %H:%M')} — {r['text']}" for r in _staff_remarks(l)
        )
        bot_reply = 'Yes' if (l.phone and normalize_phone(l.phone) in bot_ids) else ''
        ws.append([
            l.created_at.strftime('%d/%m/%Y') if l.created_at else '',
            l.name or '', _mask_phone(l.phone), l.sub_source or l.source or '',
            l.status or 'New', quality, _attempt_count(l), bot_reply, history,
        ])
    widths = [11, 22, 16, 10, 12, 18, 9, 10, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    # wrap the remarks column so multi-update history expands the row
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        row[0].alignment = Alignment(wrap_text=True, vertical='top')
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_leads.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# TEMPORARILY DISABLED - WILL RE-ADD AFTER FIXING
# @app.route('/admin/import-data')
# @login_required
# @admin_required
# def import_data_page():
#     """Temporary page for importing historical Jan-March data"""
#     # Import tool for historical data
#     return render_template('import_data.html')

# @app.route('/admin/import-customers', methods=['POST'])
# @login_required
# @admin_required
# def import_customers():
#     """Import customers from Excel"""
#     from openpyxl import load_workbook
#     import io
#     
#     file = request.files.get('customers_file')
#     if not file:
#         flash('No file uploaded', 'error')
#         return redirect(url_for('import_data_page'))
#     
#     try:
#         wb = load_workbook(io.BytesIO(file.read()))
#         ws = wb.active
#         
#         # Get staff mapping
#         users = User.query.all()
#         staff_map = {u.name.lower(): u.id for u in users}
#         
#         imported = 0
#         skipped = 0
#         errors = []
#         
#         # Skip header row, start from row 2
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             if not row[0]:  # Skip empty rows
#                 continue
#             
#             name = str(row[0]).strip() if row[0] else None
#             company = str(row[1]).strip() if row[1] else None
#             phone = str(row[2]).strip() if row[2] else None
#             email = str(row[3]).strip() if row[3] else None
#             assigned_to_name = str(row[4]).strip().lower() if row[4] else None
#             customer_type = str(row[5]).strip() if row[5] else 'Individual'
#             
#             if not name:
#                 skipped += 1
#                 continue
#             
#             # Check if customer already exists
#             existing = Customer.query.filter_by(phone=phone).first() if phone else None
#             if existing:
#                 skipped += 1
#                 continue
#             
#             # Map assigned_to
#             assigned_to_id = staff_map.get(assigned_to_name) if assigned_to_name else None
#             
#             customer = Customer(
#                 name=name,
#                 company=company,
#                 phone=phone,
#                 email=email,
#                 assigned_to=assigned_to_id,
#                 customer_type=customer_type,
#                 created_at=now_dubai()  # Will be backdated in tasks
#             )
#             db.session.add(customer)
#             imported += 1
#         
#         db.session.commit()
#         flash(f'✅ Imported {imported} customers. Skipped {skipped} (already exist or invalid).', 'success')
#     
#     except Exception as e:
#         db.session.rollback()
#         flash(f'Error: {str(e)}', 'error')
#     
#     return redirect(url_for('import_data_page'))

# @app.route('/admin/import-tasks', methods=['POST'])
# @login_required
# @admin_required  
# def import_tasks():
#     """Import tasks from Excel with historical dates"""
#     from openpyxl import load_workbook
#     import io
#     
#     file = request.files.get('tasks_file')
#     if not file:
#         flash('No file uploaded', 'error')
#         return redirect(url_for('import_data_page'))
#     
#     try:
#         wb = load_workbook(io.BytesIO(file.read()))
#         ws = wb.active
#         
#         # Get mappings
#         users = User.query.all()
#         staff_map = {u.name.lower(): u.id for u in users}
#         
#         customers = Customer.query.all()
#         customer_map = {c.name.lower(): c.id for c in customers}
#         
#         imported = 0
#         skipped = 0
#         
#         # Skip header row
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             if not row[0]:
#                 continue
#             
#             customer_name = str(row[0]).strip().lower() if row[0] else None
#             job_type = str(row[1]).strip() if row[1] else None
#             status = str(row[2]).strip() if row[2] else 'Closed'
#             created_date = row[3] if row[3] else None
#             due_date = row[4] if row[4] else None
#             completed_date = row[5] if row[5] else None
#             assigned_to_name = str(row[6]).strip().lower() if row[6] else None
#             invoiced = float(row[7]) if row[7] else 0
#             received = float(row[8]) if row[8] else 0
#             priority = str(row[9]).strip() if row[9] else 'Medium'
#             remarks = str(row[10]).strip() if row[10] else None
#             
#             if not customer_name or not job_type:
#                 skipped += 1
#                 continue
#             
#             customer_id = customer_map.get(customer_name)
#             if not customer_id:
#                 skipped += 1
#                 continue
#             
#             assigned_to_id = staff_map.get(assigned_to_name) if assigned_to_name else None
#             
#             # Parse dates
#             try:
#                 if isinstance(created_date, str):
#                     created_dt = datetime.strptime(created_date, '%d/%m/%Y')
#                 else:
#                     created_dt = created_date
#             except:
#                 created_dt = now_dubai()
#             
#             try:
#                 if isinstance(due_date, str):
#                     due_dt = datetime.strptime(due_date, '%d/%m/%Y')
#                 else:
#                     due_dt = due_date
#             except:
#                 due_dt = None
#             
#             try:
#                 if isinstance(completed_date, str):
#                     completed_dt = datetime.strptime(completed_date, '%d/%m/%Y')
#                 else:
#                     completed_dt = completed_date
#             except:
#                 completed_dt = None
#             
#             # Calculate revenue (cash-basis)
#             revenue = received if status == 'Closed' else 0
#             
#             job = Job(
#                 customer_id=customer_id,
#                 job_type=job_type,
#                 status=status,
#                 assigned_to=assigned_to_id,
#                 created_by=assigned_to_id,
#                 priority=priority,
#                 created_at=created_dt,
#                 due_date=due_dt,
#                 completed_at=completed_dt,
#                 amount_invoiced=invoiced,
#                 amount_received=received,
#                 revenue=revenue,
#                 remarks=remarks
#             )
#             db.session.add(job)
#             imported += 1
#         
#         db.session.commit()
#         flash(f'✅ Imported {imported} tasks. Skipped {skipped} (missing customer or invalid data).', 'success')
#     
#     except Exception as e:
#         db.session.rollback()
#         flash(f'Error: {str(e)}', 'error')
#     
#     return redirect(url_for('import_data_page'))



@app.route('/documents/<int:doc_id>/file')
@login_required
def document_file(doc_id):
    """Login-protected access to a customer/employee document: redirects to a
    signed Cloudinary URL. Raw Cloudinary links no longer work (authenticated)."""
    d = Document.query.get_or_404(doc_id)
    if not _can_view_document(d):
        flash('You are not authorised to view this document.', 'error')
        return redirect(url_for('dashboard'))
    if not d.file_url:
        return 'No file attached', 404
    return redirect(signed_document_url(d.file_url, d.cloudinary_public_id))

@app.route('/tahfeel-docs/<int:doc_id>/file')
@login_required
def company_document_file(doc_id):
    """Internal Tahfeel document (CompanyDocument) — restricted to the same people
    who manage them: Admin or Saada (matches the /tahfeel-doc page access)."""
    if session.get('role') != 'admin' and session.get('user_email') != 'saadatahfeel@gmail.com':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    d = CompanyDocument.query.get_or_404(doc_id)
    if not d.document_url:
        return 'No file attached', 404
    return redirect(signed_document_url(d.document_url, d.cloudinary_public_id))

@app.route('/admin/secure-documents', methods=['POST'])
@login_required
@admin_required
def secure_documents():
    """One-time migration: switch every existing Cloudinary document from public
    to authenticated access, so raw links stop working and files are only
    reachable via the login-protected /documents routes."""
    secured, failed = 0, 0
    for model, url_attr in ((Document, 'file_url'), (CompanyDocument, 'document_url')):
        for d in model.query.filter(model.cloudinary_public_id != None).all():
            url = getattr(d, url_attr) or ''
            rt = 'raw' if '/raw/upload/' in url else 'video' if '/video/upload/' in url else 'image'
            done = False
            # resource_type may be mislabeled in legacy URLs — try the likely one, then the others
            for try_rt in [rt] + [x for x in ('image', 'raw', 'video') if x != rt]:
                try:
                    cloudinary.api.update(d.cloudinary_public_id, resource_type=try_rt,
                                          access_mode='authenticated')
                    done = True
                    break
                except Exception:
                    continue
            if done:
                secured += 1
            else:
                failed += 1
                print(f'[secure-docs] could not secure {model.__name__} id={d.id} pid={d.cloudinary_public_id}')
    flash(f'🔐 Secured {secured} document file(s) on Cloudinary.' +
          (f' {failed} could not be updated — check logs.' if failed else ' Raw links are now disabled.'))
    return redirect(url_for('admin_panel'))


@app.route('/admin/staff/add', methods=['POST'])
@login_required
@admin_required
def admin_add_staff():
    email = request.form['email']
    existing = User.query.filter_by(email=email).first()
    if existing:
        flash('This email already exists')
        return redirect(url_for('admin_panel'))
    try:
        rf = request.form.get('report_from', '').strip()
        user = User(
            name=request.form['name'],
            email=email,
            password=generate_password_hash(request.form['password']),
            role=request.form.get('role', 'staff'),
            phone=request.form.get('phone', '').strip() or None,
            report_from=datetime.strptime(rf, '%Y-%m-%d').date() if rf else None
        )
        db.session.add(user)
        db.session.commit()
        flash('Staff member added successfully')
    except Exception as e:
        db.session.rollback()
        flash('Error — ' + str(e))
    return redirect(url_for('admin_panel'))

@app.route('/admin/staff/<int:user_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_staff(user_id):
    user = User.query.get_or_404(user_id)
    if not can_manage_user(user):
        flash('Only a Super Admin can edit an admin account')
        return redirect(url_for('admin_panel'))
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    new_password = request.form.get('password', '').strip()
    new_role = request.form.get('role', '').strip()
    if name:
        user.name = name
    if email:
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != user_id:
            flash('That email is already in use')
            return redirect(url_for('admin_panel'))
        user.email = email
    if new_role in ['staff', 'sales', 'operations', 'admin', 'finance', 'marketing']:
        user.role = new_role
    if new_password:
        user.password = generate_password_hash(new_password)
    user.phone = request.form.get('phone', '').strip() or None
    rf = request.form.get('report_from', '').strip()
    user.report_from = datetime.strptime(rf, '%Y-%m-%d').date() if rf else None
    db.session.commit()
    flash('Staff member updated successfully')
    return redirect(url_for('admin_panel'))

@app.route('/admin/service/add', methods=['POST'])
@login_required
@admin_required
def admin_add_service():
    name = request.form.get('name', '').strip()
    if name:
        existing = Service.query.filter_by(name=name).first()
        if existing:
            flash('Service already exists')
        else:
            db.session.add(Service(name=name))
            db.session.commit()
            flash(f'Service "{name}" added')
    return redirect(url_for('admin_panel'))

@app.route('/admin/service/<int:service_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_service(service_id):
    service = Service.query.get_or_404(service_id)
    db.session.delete(service)
    db.session.commit()
    flash(f'Service "{service.name}" removed')
    return redirect(url_for('admin_panel'))

@app.route('/admin/source/add', methods=['POST'])
@login_required
@admin_required
def admin_add_source():
    name = request.form.get('name', '').strip()
    if name:
        existing = Source.query.filter_by(name=name).first()
        if existing:
            flash('Source already exists')
        else:
            db.session.add(Source(name=name))
            db.session.commit()
            flash(f'Source "{name}" added')
    return redirect(url_for('admin_panel'))

@app.route('/admin/source/<int:source_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_source(source_id):
    source = Source.query.get_or_404(source_id)
    db.session.delete(source)
    db.session.commit()
    flash(f'Source "{source.name}" removed')
    return redirect(url_for('admin_panel'))

@app.route('/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.active = not user.active
    db.session.commit()
    flash(f'{"Activated" if user.active else "Deactivated"} {user.name}')
    return redirect(url_for('admin_panel'))

@app.route('/admin/staff/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_staff(user_id):
    user = User.query.get_or_404(user_id)
    if not can_manage_user(user):
        flash('Only a Super Admin can deactivate an admin account')
        return redirect(url_for('admin_panel'))
    if user.id == session.get('user_id') and user.active:
        flash('You cannot deactivate your own account')
        return redirect(url_for('admin_panel'))
    user.active = not user.active
    db.session.commit()
    flash(f'{"Activated" if user.active else "Deactivated"} {user.name}')
    return redirect(url_for('admin_panel'))

@app.route('/admin/staff/<int:user_id>/toggle-leave', methods=['POST'])
@login_required
@admin_required
def toggle_staff_leave(user_id):
    user = User.query.get_or_404(user_id)
    user.on_leave = not user.on_leave
    db.session.commit()
    status = 'On Leave' if user.on_leave else 'Available'
    flash(f'{user.name} marked as {status}. Meta leads will {"skip" if user.on_leave else "include"} them.')
    return redirect(url_for('admin_panel'))

# ── Customers ────────────────────────────────────────────────────────────────

@app.route('/customers')
@login_required
def customers():
    now = now_dubai()
    search = request.args.get('search', '').strip().lower()
    birthday_filter = request.args.get('birthday', '')
    rep_filter = request.args.get('representative', '')
    type_filter = request.args.get('type', '')
    status_filter = request.args.get('status', '')
    birthdays_today = []
    try:
        # Ensure columns exist first
        with db.engine.connect() as conn:
            for col, typ in [('phone2','VARCHAR(20)'),('assigned_to','INTEGER'),('date_of_birth','DATE')]:
                try:
                    conn.execute(db.text(f'ALTER TABLE customer ADD COLUMN IF NOT EXISTS {col} {typ}'))
                    conn.commit()
                except: pass
        customer_list = Customer.query.order_by(Customer.created_at.desc()).all()
        if search:
            csdigits = ''.join(ch for ch in search if ch.isdigit())
            def _cmatch(c):
                if search in (c.name or '').lower() or search in (c.company or '').lower():
                    return True
                if csdigits:
                    for p in (c.phone, c.phone2, c.mobile, c.whatsapp):
                        if p and csdigits in ''.join(ch for ch in p if ch.isdigit()):
                            return True
                return False
            customer_list = [c for c in customer_list if _cmatch(c)]
        if birthday_filter == 'today':
            customer_list = [c for c in customer_list if c.date_of_birth and
                             c.date_of_birth.month == now.month and
                             c.date_of_birth.day == now.day]
        if rep_filter:
            customer_list = [c for c in customer_list if c.assigned_to == int(rep_filter)]
        if type_filter:
            customer_list = [c for c in customer_list if (c.customer_type or 'Individual') == type_filter]
        if status_filter:
            customer_list = [c for c in customer_list if c.ac_status == status_filter]
        try:
            bday_list = Customer.query.filter(Customer.date_of_birth != None).all()
            birthdays_today = [c for c in bday_list if c.date_of_birth and
                               c.date_of_birth.month == now.month and c.date_of_birth.day == now.day]
        except:
            birthdays_today = []
    except Exception as e:
        customer_list = []
        flash(f'Error loading customers: {e}')
    page = int(request.args.get('page', 1))
    per_page = 25
    total = len(customer_list)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated = customer_list[(page-1)*per_page : page*per_page]
    users = User.query.filter_by(active=True).order_by(User.name).all()
    return render_template('customers.html', customers=paginated, page=page, total_pages=total_pages,
                           total=total, search=request.args.get('search',''),
                           birthdays_today=birthdays_today, now=now, birthday_filter=birthday_filter,
                           users=users, rep_filter=rep_filter, type_filter=type_filter, status_filter=status_filter)

@app.route('/api/customer-phone-exists')
@login_required
def api_customer_phone_exists():
    from flask import jsonify
    phone = (request.args.get('phone') or '').strip()
    if not phone:
        return jsonify({'exists': False, 'name': ''})
    c = Customer.query.filter_by(phone=phone).first()
    return jsonify({'exists': bool(c), 'name': c.name if c else ''})


def _add_months(d, months):
    import calendar
    m = d.month - 1 + months
    y = d.year + m // 12
    m = m % 12 + 1
    return d.replace(year=y, month=m, day=min(d.day, calendar.monthrange(y, m)[1]))

def _save_tax_fields(customer):
    """Save VAT + Corporate-tax filing status/due from the form.
    Auto-roll: marking 'Filed' advances the due date one period (VAT quarterly,
    corp tax yearly) and resets status to 'Not filed' (next period now due)."""
    for prefix, months in (('vat', 3), ('corp_tax', 12)):
        if (prefix + '_status') not in request.form and (prefix + '_due_date') not in request.form:
            continue  # form didn't include these — preserve existing values
        status = (request.form.get(prefix + '_status') or '').strip()
        due_str = (request.form.get(prefix + '_due_date') or '').strip()
        due = None
        if due_str:
            try:
                due = datetime.strptime(due_str, '%Y-%m-%d').date()
            except ValueError:
                due = None
        if status == 'Filed' and due:
            due = _add_months(due, months)
            status = 'Not filed'
        setattr(customer, prefix + '_status', status or None)
        setattr(customer, prefix + '_due_date', due)

def _customer_type_template(ctype):
    return 'add_customer_company.html' if ctype == 'Company' else 'add_customer_individual.html'

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    converted_leads = Lead.query.filter_by(status='Converted').order_by(Lead.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        ctype = request.form.get('customer_type', 'Individual')
        doc_types = DocType.query.order_by(DocType.name).all()
        if ctype == 'Company' and not (request.form.get('contact_person') or '').strip():
            flash('Contact Person is required for a Company client', 'error')
            return redirect(url_for('add_customer', type=ctype))
        # Validate required fields
        if not request.form.get('lead_id'):
            if not request.form.get('source'):
                flash('Source is required', 'error')
                users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
                return render_template(_customer_type_template(ctype), users=users, sources=sources, converted_leads=converted_leads, doc_types=doc_types)
            if not request.form.get('assigned_to'):
                flash('Primary Representative is required', 'error')
                users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
                return render_template(_customer_type_template(ctype), users=users, sources=sources, converted_leads=converted_leads, doc_types=doc_types)

        # Check for duplicate phone number
        phone_to_check = None
        lead_id = request.form.get('lead_id') or None
        if lead_id:
            lead = Lead.query.get(int(lead_id))
            phone_to_check = lead.phone
        else:
            phone_to_check = request.form.get('phone', '').strip()

        if phone_to_check and not request.form.get('allow_duplicate'):
            existing_customer = Customer.query.filter_by(phone=phone_to_check).first()
            if existing_customer:
                flash(f'⚠️ Phone {phone_to_check} already exists for "{existing_customer.name}". Submit again to add anyway.', 'error')
                users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
                return render_template(_customer_type_template(ctype), users=users, sources=sources, converted_leads=converted_leads, doc_types=doc_types)
        
        if lead_id:
            customer = Customer(
                name=lead.name, company=lead.company, phone=lead.phone,
                phone2=getattr(lead, 'phone2', None),
                email=lead.email, address=lead.address, source=lead.source,
                notes=request.form.get('notes'), lead_id=int(lead_id)
            )
        else:
            customer = Customer(
                name=request.form.get('name', '').strip(),
                company=request.form.get('company', '').strip() or None,
                phone=request.form.get('phone', '').strip(),
                phone2=request.form.get('phone2', '').strip() or None,
                email=request.form.get('email', '').strip() or None,
                address=request.form.get('address', '').strip() or None,
                source=request.form.get('source', '').strip() or None,
                nationality=request.form.get('nationality', '').strip() or None,
                customer_type=request.form.get('customer_type', 'Individual'),
                contact_person=request.form.get('contact_person', '').strip() or None,
                assigned_to=int(request.form.get('assigned_to')) if request.form.get('assigned_to') else None,
                notes=request.form.get('notes', '').strip() or None,
                date_of_birth=datetime.strptime(request.form.get('date_of_birth'), '%Y-%m-%d').date() if request.form.get('date_of_birth') else None
            )
        db.session.add(customer)
        db.session.flush()  # get customer.id before commit

        # Company profile fields (UAE) — applied for any customer; blank for individuals
        for _f in ['ac_code','trade_name','legal_form','jurisdiction','licensing_authority','freezone_name','emirate','country_incorp','business_activity','ac_status','po_box','mobile','whatsapp','website','uae_pass_number','uae_pass_name']:
            setattr(customer, _f, request.form.get(_f, '').strip() or None)
        _aod = request.form.get('ac_opening_date', '').strip()
        customer.ac_opening_date = datetime.strptime(_aod, '%Y-%m-%d').date() if _aod else None
        _save_tax_fields(customer)

        # Save inline documents
        doc_types_inline = request.form.getlist('doc_type[]')
        doc_owners = request.form.getlist('doc_owner[]')
        doc_expiries = request.form.getlist('doc_expiry[]')
        doc_notes_list = request.form.getlist('doc_notes[]')

        for i, dt in enumerate(doc_types_inline):
            if not dt: continue
            expiry = None
            try:
                if i < len(doc_expiries) and doc_expiries[i]:
                    expiry = datetime.strptime(doc_expiries[i], '%Y-%m-%d')
            except: pass
            file_url, file_name, public_id = None, None, None
            doc_file = request.files.get(f'doc_file_{i+1}')
            if doc_file and doc_file.filename:
                file_url, public_id = upload_to_cloudinary(doc_file)
                file_name = doc_file.filename
            doc = Document(
                customer_id=customer.id,
                doc_type=dt,
                owner_name=doc_owners[i] if i < len(doc_owners) else customer.name,
                belongs_to='Individual',
                expiry_date=expiry,
                notes=doc_notes_list[i] if i < len(doc_notes_list) else None,
                added_by=session['user_name'],
                file_url=file_url,
                file_name=file_name,
                cloudinary_public_id=public_id,
                uploaded_by=session.get('user_id')
            )
            db.session.add(doc)

        db.session.commit()
        flash('Customer added successfully')
        return redirect(url_for('customer_detail', customer_id=customer.id))
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales','operations','admin'])).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    ctype = request.args.get('type', '')
    if ctype not in ('Individual', 'Company'):
        return render_template('add_customer_choose.html', converted_leads=converted_leads)
    return render_template(_customer_type_template(ctype), converted_leads=converted_leads, sources=sources, users=users, doc_types=doc_types)

@app.route('/customers/<int:customer_id>')
@login_required
def customer_detail(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    now = now_dubai()
    jobs = Job.query.filter_by(customer_id=customer_id).order_by(Job.created_at.desc()).all()
    docs = Document.query.filter_by(customer_id=customer_id, employee_id=None).order_by(Document.expiry_date).all()
    employees = Employee.query.filter_by(customer_id=customer_id).order_by(Employee.name).all()
    owners = Owner.query.filter_by(customer_id=customer_id).order_by(Owner.id).all()
    total_invoiced = sum(j.amount_invoiced or 0 for j in jobs)
    total_received = sum(j.amount_received or 0 for j in jobs)
    return render_template('customer_detail.html', customer=customer, jobs=jobs,
                           documents=docs, employees=employees, owners=owners, now=now, today=now.date(),
                           total_invoiced=total_invoiced, total_received=total_received,
                           wa_templates=wa_send_context(customer=customer))


@app.route('/customers/<int:customer_id>/health')
@login_required
def customer_health(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    now = now_dubai()
    today = now.date()
    all_docs = Document.query.filter_by(customer_id=customer_id).all()
    docs = [d for d in all_docs if d.expiry_date]
    docs.sort(key=lambda d: d.expiry_date)
    def dleft(d):
        return (d.expiry_date.date() - today).days
    expired = [d for d in docs if dleft(d) < 0]
    expiring = [d for d in docs if 0 <= dleft(d) <= 90]   # "expiring soon" = within 90 days
    valid = [d for d in docs if dleft(d) > 90]
    total = len(all_docs)
    scored = len(docs)
    score = round(100 * (len(valid) + 0.5 * len(expiring)) / scored) if scored else None
    if score is None:
        band = 'No data'
    elif score >= 90:
        band = 'Excellent'
    elif score >= 70:
        band = 'Good'
    elif score >= 50:
        band = 'Average'
    else:
        band = 'Poor'
    company_docs = sorted([d for d in all_docs if not d.employee_id], key=lambda d: (d.expiry_date or datetime.max))
    employees = Employee.query.filter_by(customer_id=customer_id).order_by(Employee.name).all()
    owners_count = Owner.query.filter_by(customer_id=customer_id).count()  # partners / UBO
    # Employee document summary (across all employees of this company)
    emp_docs = [d for d in all_docs if d.employee_id and d.expiry_date]
    emp_valid = len([d for d in emp_docs if dleft(d) > 90])
    emp_expiring = len([d for d in emp_docs if 0 <= dleft(d) <= 90])
    emp_expired = len([d for d in emp_docs if dleft(d) < 0])
    # Upcoming renewals = anything due within the next 30 days (docs + tax filings)
    renewals_30 = len([d for d in docs if 0 <= dleft(d) <= 30])
    for _due in (customer.vat_due_date, customer.corp_tax_due_date):
        if _due and 0 <= (_due - today).days <= 30:
            renewals_30 += 1
    wa_number = (customer.whatsapp or customer.mobile or customer.phone or '').replace(' ', '').replace('+', '')
    from_email = os.environ.get('SMTP_FROM') or os.environ.get('SMTP_USER') or 'info@tahfeel.ae'
    return render_template('customer_health.html', customer=customer, now=now, today=today,
                           total=total, valid=valid, expiring=expiring, expired=expired,
                           score=score, band=band, company_docs=company_docs, employees=employees,
                           emp_docs_total=len(emp_docs), emp_valid=emp_valid, emp_expiring=emp_expiring,
                           emp_expired=emp_expired, owners_count=owners_count, renewals_30=renewals_30,
                           wa_number=wa_number, from_email=from_email)

@app.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    sources = Source.query.order_by(Source.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales','operations','admin'])).all()
    if request.method == 'POST':
        if request.form.get('customer_type') == 'Company' and not (request.form.get('contact_person') or '').strip():
            flash('Contact Person is required for a Company client', 'error')
            return redirect(url_for('edit_customer', customer_id=customer_id))
        customer.name = request.form.get('name', '').strip() or customer.name
        if 'company' in request.form:
            customer.company = request.form.get('company', '').strip()
        customer.phone = request.form.get('phone', '').strip()
        customer.phone2 = request.form.get('phone2', '').strip() or None
        customer.email = request.form.get('email', '').strip()
        customer.address = request.form.get('address', '').strip()
        customer.source = request.form.get('source', '').strip()
        customer.nationality = request.form.get('nationality', '').strip() or None
        dob_str = request.form.get('date_of_birth', '').strip()
        customer.date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date() if dob_str else None
        customer.customer_type = request.form.get('customer_type', 'Individual')
        customer.contact_person = request.form.get('contact_person', '').strip() or None
        customer.alert_email = request.form.get('alert_email', '').strip() or None
        customer.alert_whatsapp = request.form.get('alert_whatsapp', '').strip() or None
        customer.alerts_enabled = bool(request.form.get('alerts_enabled'))
        # Company profile fields (UAE) — only update fields actually present in the
        # submitted form, so trimmed/removed fields keep their existing values (no wipe).
        for _f in ['ac_code','trade_name','legal_form','jurisdiction','licensing_authority','freezone_name','emirate','country_incorp','business_activity','ac_status','po_box','mobile','whatsapp','website','uae_pass_number','uae_pass_name']:
            if _f in request.form:
                setattr(customer, _f, request.form.get(_f, '').strip() or None)
        _aod = request.form.get('ac_opening_date', '').strip()
        customer.ac_opening_date = datetime.strptime(_aod, '%Y-%m-%d').date() if _aod else None
        _save_tax_fields(customer)
        try:
            customer.assigned_to = int(request.form.get('assigned_to')) if request.form.get('assigned_to') else None
        except:
            pass
        customer.notes = request.form.get('notes', '').strip()
        # Save any inline documents added
        doc_types_inline = request.form.getlist('doc_type[]')
        doc_owners = request.form.getlist('doc_owner[]')
        doc_expiries = request.form.getlist('doc_expiry[]')
        doc_notes_list = request.form.getlist('doc_notes[]')
        for i, dt in enumerate(doc_types_inline):
            if not dt: continue
            expiry = None
            try:
                if i < len(doc_expiries) and doc_expiries[i]:
                    expiry = datetime.strptime(doc_expiries[i], '%Y-%m-%d')
            except: pass
            # Handle file upload for this doc
            doc_file_key = f'doc_file_{i+1}'
            file_name = None
            file_url = None
            public_id = None
            if doc_file_key in request.files:
                f = request.files[doc_file_key]
                if f and f.filename:
                    file_name = f.filename
                    file_url, public_id = upload_to_cloudinary(f)
            doc = Document(
                customer_id=customer_id,
                doc_type=dt,
                owner_name=doc_owners[i] if i < len(doc_owners) and doc_owners[i] else customer.name,
                belongs_to='Individual',
                expiry_date=expiry,
                notes=doc_notes_list[i] if i < len(doc_notes_list) else None,
                file_name=file_name,
                file_url=file_url,
                cloudinary_public_id=public_id,
                added_by=session['user_name']
            )
            db.session.add(doc)
        db.session.commit()
        flash('Customer updated successfully')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    doc_types = DocType.query.order_by(DocType.name).all()
    existing_docs = Document.query.filter_by(customer_id=customer_id).order_by(Document.expiry_date).all()
    now = now_dubai()
    return render_template('edit_customer.html', customer=customer, sources=sources, users=users, doc_types=doc_types, existing_docs=existing_docs, now=now)

@app.route('/customers/<int:customer_id>/toggle-alerts', methods=['POST'])
@login_required
def toggle_customer_alerts(customer_id):
    c = Customer.query.get_or_404(customer_id)
    c.alerts_enabled = not c.alerts_enabled
    db.session.commit()
    flash(('🔔 Periodic compliance emails ON' if c.alerts_enabled else '🔕 Periodic compliance emails OFF') + f' for {c.name}')
    if request.args.get('next') == 'health':
        return redirect(url_for('customer_health', customer_id=customer_id))
    return redirect(url_for('customer_detail', customer_id=customer_id))

@app.route('/admin/alerts/disable-all', methods=['POST'])
@login_required
@admin_required
def disable_all_alerts():
    """One-click safety switch: turn OFF report alerts for every customer.
    Used when customer details aren't finalised yet, so no automated
    weekly/monthly emails go out until each customer is re-enabled."""
    n = Customer.query.filter_by(alerts_enabled=True).update({'alerts_enabled': False})
    db.session.commit()
    flash(f'🔕 Report alerts turned OFF for {n} customer(s). No automated emails will send until you re-enable them individually.')
    return redirect(url_for('admin_panel'))

@app.route('/customers/<int:customer_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    # Only delete if no jobs linked (tasks carry finance/revenue — remove them first)
    if customer.jobs:
        flash('Cannot delete customer with existing tasks. Remove tasks first.')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    try:
        company_ids = [c.id for c in Company.query.filter_by(customer_id=customer_id).all()]
        employee_ids = [e.id for e in Employee.query.filter_by(customer_id=customer_id).all()]
        # 1) documents attached to the customer, its companies, and its employees
        Document.query.filter_by(customer_id=customer_id).delete(synchronize_session=False)
        if company_ids:
            Document.query.filter(Document.company_id.in_(company_ids)).delete(synchronize_session=False)
        if employee_ids:
            Document.query.filter(Document.employee_id.in_(employee_ids)).delete(synchronize_session=False)
        # 2) the company / employee / owner records that belong to this customer
        Employee.query.filter_by(customer_id=customer_id).delete(synchronize_session=False)
        Owner.query.filter_by(customer_id=customer_id).delete(synchronize_session=False)
        Company.query.filter_by(customer_id=customer_id).delete(synchronize_session=False)
        # 3) unlink WhatsApp history (keep the messages, just detach)
        WhatsAppMessage.query.filter_by(customer_id=customer_id).update({'customer_id': None})
        db.session.delete(customer)
        db.session.commit()
        flash('Customer deleted')
    except Exception as e:
        db.session.rollback()
        print(f'[delete_customer] failed for {customer_id}: {e}')
        flash('Could not delete this customer — it still has linked records. Tell the admin.', 'error')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    return redirect(url_for('customers'))


@app.route('/customers/export')
@login_required
def export_customers():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Customers'
    headers = ['Name','Company','Phone','Phone 2','Email','Address','Nationality','Date of Birth','Customer Type','Source','Primary Representative','Notes']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    customers = Customer.query.order_by(Customer.created_at.desc()).all()
    for r, c in enumerate(customers, 2):
        # Get assigned representative name
        rep_name = ''
        if c.assigned_to:
            rep_user = User.query.get(c.assigned_to)
            if rep_user:
                rep_name = rep_user.name
        ws.append([
            c.name,
            c.company or '',
            c.phone or '',
            c.phone2 or '',
            c.email or '',
            c.address or '',
            c.nationality or '',
            c.date_of_birth.strftime('%d/%m/%Y') if c.date_of_birth else '',
            c.customer_type or 'Individual',
            c.source or '',
            rep_name,
            c.notes or ''
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='tahfeel_customers.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/customers/template')
@login_required
def customer_import_template():
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Customers'
    from openpyxl.worksheet.datavalidation import DataValidation
    headers = ['Name *', 'Phone *', 'Company', 'Phone 2', 'Email', 'Address', 'Nationality', 'Date of Birth', 'Type', 'Source', 'Primary Representative *', 'Notes']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A3B8B')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 18)
    # Sample rows
    samples = [
        ['Ahmed Al Mansoori', '+971501234567', 'Al Mansoori Trading LLC', '+971551234567', 'ahmed@example.com', 'Dubai, UAE', 'Emirati', '15/01/1985', 'Company', 'Referral', 'Aslam', 'VIP client'],
        ['Priya Sharma', '+971507654321', '', '', 'priya@gmail.com', 'Sharjah, UAE', 'Indian', '20/05/1990', 'Individual', 'WhatsApp', 'Anfal', ''],
        ['XYZ Investments', '+971509876543', 'XYZ Investments LLC', '', '', 'Abu Dhabi, UAE', 'British', '', 'Company', 'Website', 'Lukman', 'Golden visa interested'],
    ]
    for row in samples:
        ws.append(row)
    # Reference sheet for dropdowns
    ref = wb.create_sheet('Reference')
    ref.sheet_state = 'hidden'
    sources = Source.query.order_by(Source.name).all()
    src_names = [s.name for s in sources] or ['Referral', 'WhatsApp', 'Website', 'Walk-in', 'Social Media', 'Other']
    for i, s in enumerate(src_names, 1):
        ref.cell(i, 1, s)
    types = ['Individual', 'Company', 'Investor']
    for i, t in enumerate(types, 1):
        ref.cell(i, 2, t)
    # Get staff names for Primary Representative dropdown
    users = User.query.filter_by(active=True).order_by(User.name).all()
    staff_names = [u.name for u in users]
    for i, name in enumerate(staff_names, 1):
        ref.cell(i, 3, name)
    # Source dropdown — column J (10)
    src_range = f'Reference!$A$1:$A${len(src_names)}'
    dv_src = DataValidation(type='list', formula1=src_range, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_src)
    dv_src.add('J2:J1000')
    # Type dropdown — column I (9)
    dv_type = DataValidation(type='list', formula1='Reference!$B$1:$B$3', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_type)
    dv_type.add('I2:I1000')
    # Primary Representative dropdown — column K (11)
    staff_range = f'Reference!$C$1:$C${len(staff_names)}'
    dv_staff = DataValidation(type='list', formula1=staff_range, allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv_staff)
    dv_staff.add('K2:K1000')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='customer_import_template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/customers/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_customers():
    sources = Source.query.order_by(Source.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales','operations','admin'])).all()
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload an .xlsx file')
            return redirect(url_for('import_customers'))
        from openpyxl import load_workbook
        wb = load_workbook(f)
        ws = wb.active
        imported = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            name = str(row[0]).strip()
            if not name: continue
            phone = str(row[1]).strip() if row[1] else ''
            company = str(row[2]).strip() if row[2] else ''
            phone2 = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            email = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            address = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            source = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            nationality = str(row[7]).strip() if len(row) > 7 and row[7] else ''
            ctype = str(row[8]).strip() if len(row) > 8 and row[8] else 'Individual'
            notes = str(row[9]).strip() if len(row) > 9 and row[9] else ''
            c = Customer(name=name, phone=phone, company=company, phone2=phone2 or None,
                        email=email, address=address, source=source,
                        nationality=nationality or None, customer_type=ctype, notes=notes)
            db.session.add(c)
            imported += 1
        db.session.commit()
        flash(f'Import complete: {imported} customers added')
        return redirect(url_for('customers'))
    return render_template('import_customers.html', sources=sources, users=users)

# ── Jobs ──────────────────────────────────────────────────────────────────────

JOB_STATUSES = ['Assigned', 'Job Started', 'Processing', 'Pending Authority', 'On Hold', 'Delayed', 'Partially Completed', 'Done']
# Rolling stages: these progress stages only move FORWARD for non-admin users.
# Pending Authority / On Hold / Delayed are situational states, not ladder steps —
# a task can enter/leave them any time, but on exit can't land below its highest stage.
JOB_STAGE_RANK = {'Assigned': 1, 'Job Started': 2, 'Processing': 3, 'Partially Completed': 4, 'Done': 5}

def job_stage_floor(job):
    """Rank of the task's current ladder stage. If it's parked in a situational
    state (On Hold etc.), the most recent ladder stage before that — so it can't
    exit the parking state to an earlier stage. Uses current position, not the
    all-time high, so an admin rollback lets staff progress forward again."""
    if job.status in JOB_STAGE_RANK:
        return JOB_STAGE_RANK[job.status]
    for u in job.updates:  # newest first
        if u.status in JOB_STAGE_RANK:
            return JOB_STAGE_RANK[u.status]
    return 0
JOB_STATUSES_FINANCE = ['Closed']  # Finance-only status
JOB_STATUSES_ALL = ['Pending Finance Approval'] + JOB_STATUSES + ['Pending Finance Close', 'Closed']

@app.route('/jobs')
@login_required
def jobs():
    now = now_dubai()
    role = session['role']
    # Remember the last-used filters so returning from a task keeps the same view.
    FJK = ['status', 'priority', 'assigned_to', 'staff', 'representative', 'date', 'from_date', 'to_date', 'customer', 'sort', 'order', 'all']
    if request.args.get('reset') == '1':
        session.pop('jobs_filters', None)
        return redirect(url_for('jobs'))
    if any(k in request.args for k in FJK):
        # MERGE the params present in the URL over the saved set (don't replace the
        # whole dict) — otherwise a link carrying only ?sort=... or ?all=1 would wipe
        # the user's saved status/staff filters. The filter form submits every field
        # (including empties), so clearing filters via the form still works.
        saved = dict(session.get('jobs_filters') or {})
        saved.update({k: request.args.get(k, '') for k in FJK if k in request.args})
        session['jobs_filters'] = saved
        args = saved
    elif 'jobs_filters' in session:
        args = session['jobs_filters']
    else:
        args = request.args
    sort = args.get('sort') or 'due'
    order = args.get('order') or 'asc'
    status_filter = args.get('status', '')

    try:
        # Exclude Done and Closed by default unless explicitly filtered
        if not status_filter:
            if role in ['admin', 'finance']:
                job_list = Job.query.options(db.joinedload(Job.customer).joinedload(Customer.rep)).all()
            else:
                job_list = Job.query.options(db.joinedload(Job.customer).joinedload(Customer.rep)).filter(Job.status.notin_(['Done', 'Closed', 'Closed - Pending Partner Commission'])).order_by(Job.due_date.asc()).all()
        elif status_filter == 'Closed':
            job_list = Job.query.options(db.joinedload(Job.customer).joinedload(Customer.rep)).filter(Job.status.in_(['Closed', 'Closed - Pending Partner Commission'])).all()
        elif status_filter == 'Done':
            job_list = Job.query.options(db.joinedload(Job.customer).joinedload(Customer.rep)).filter(Job.status == 'Done').all()
        else:
            job_list = Job.query.options(db.joinedload(Job.customer).joinedload(Customer.rep)).filter(Job.status == status_filter).order_by(Job.due_date.asc()).all()
        
        priority_filter = args.get('priority', '')
        assigned_filter = args.get('assigned_to', '') or args.get('staff', '')
        date_filter = args.get('date', '')
        from_date = args.get('from_date', '')
        to_date = args.get('to_date', '')

        customer_search = (args.get('customer', '') or '').strip().lower()
        if customer_search:
            job_list = [j for j in job_list if customer_search in (j.customer.name or '').lower() or customer_search in (j.customer.company or '').lower()]
        if status_filter and status_filter not in ["Closed", "Done"]:
            job_list = [j for j in job_list if j.status == status_filter]
        if priority_filter:
            job_list = [j for j in job_list if j.priority == priority_filter]
        if assigned_filter:
            try:
                job_list = [j for j in job_list if j.assigned_to == int(assigned_filter)]
            except: pass
        representative_filter = args.get('representative', '')
        if representative_filter:
            try:
                job_list = [j for j in job_list if j.customer and j.customer.assigned_to == int(representative_filter)]
            except: pass
        # Due date filters
        if date_filter == 'today':
            job_list = [j for j in job_list if j.due_date and j.due_date.date() == now.date()]
        elif date_filter == 'week':
            week_end = now + timedelta(days=7)
            job_list = [j for j in job_list if j.due_date and now.date() <= j.due_date.date() <= week_end.date()]
        elif date_filter == 'month':
            job_list = [j for j in job_list if j.due_date and j.due_date.month == now.month and j.due_date.year == now.year]
        elif date_filter == 'custom':
            if from_date:
                try:
                    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
                    job_list = [j for j in job_list if j.due_date and j.due_date.date() >= fd]
                except: pass
            if to_date:
                try:
                    td = datetime.strptime(to_date, '%Y-%m-%d').date()
                    job_list = [j for j in job_list if j.due_date and j.due_date.date() <= td]
                except: pass
        # Operations staff default to THEIR own tasks (their to-do list) unless they
        # explicitly view all (?all=1) or filter by a specific staff member.
        mine_only = (role == 'operations' and args.get('all') != '1'
                     and not assigned_filter and not representative_filter)
        if mine_only:
            job_list = [j for j in job_list if j.assigned_to == session.get('user_id')]

        overdue = [j for j in job_list if j.due_date and j.due_date < now and j.status not in ['Done', 'Pending Finance Approval']]

        # Global stats — same for all roles, pulled from full DB
        all_jobs_global = Job.query.all()
        closed_statuses = ['Closed', 'Closed - Pending Partner Commission']
        stat_total = len(all_jobs_global)
        stat_done_closed = len([j for j in all_jobs_global if j.status in ['Done'] + closed_statuses])
        stat_overdue = len([j for j in all_jobs_global if j.due_date and j.due_date < now and j.status not in ['Done'] + closed_statuses])
        stat_processing = len([j for j in all_jobs_global if j.status not in ['Done'] + closed_statuses])
        stat_pending_finance = len([j for j in all_jobs_global if j.status in ['Pending Finance Approval', 'Pending Finance Close']])
        users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
        jobs_invoiced = sum((j.amount_invoiced or 0) for j in job_list)
        jobs_received = sum((j.amount_received or 0) for j in job_list)
        jobs_pending = jobs_invoiced - jobs_received
        jobs_completed = sum((j.amount_received or 0) for j in job_list if j.status == 'Done')
    except Exception as e:
        # DB migration not complete yet — run it now
        try:
            with db.engine.connect() as conn:
                for col, typ in [
                    ('amount_invoiced', 'FLOAT DEFAULT 0'),
                    ('amount_received', 'FLOAT DEFAULT 0'),
                    ('num_persons', 'INTEGER DEFAULT 1'),
                    ('finance_approved_by', 'INTEGER'),
                    ('finance_approved_at', 'TIMESTAMP'),
                    ('finance_notes', 'TEXT'),
                ]:
                    try:
                        conn.execute(db.text(f'ALTER TABLE job ADD COLUMN IF NOT EXISTS {col} {typ}'))
                    except:
                        pass
                conn.commit()
        except:
            pass
        flash('System update applied. Please refresh.')
        return redirect(url_for('dashboard'))
    return render_template('jobs.html', jobs=job_list, now=now, overdue=overdue, mine_only=mine_only,
                           statuses=JOB_STATUSES + (['Closed'] if session.get('role') in ['admin','finance'] else []), users=users,
                           status_filter=status_filter, priority_filter=priority_filter,
                           assigned_filter=assigned_filter, date_filter=date_filter,
                           sort=sort, order=order,
                           jobs_invoiced=jobs_invoiced, jobs_received=jobs_received,
                           jobs_pending=jobs_pending, jobs_completed=jobs_completed,
                           stat_total=stat_total, stat_done_closed=stat_done_closed,
                           stat_overdue=stat_overdue, stat_processing=stat_processing,
                           stat_pending_finance=stat_pending_finance)


@app.route('/jobs/export')
@login_required
def export_jobs():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tasks'
    headers = ['ID','Customer','Company','Service Type','Assigned To','Status','Priority','Due Date','Invoiced (AED)','Received (AED)','Pending (AED)','Created']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    jobs = Job.query.order_by(Job.due_date.asc()).all()
    for j in jobs:
        ws.append([
            j.id,
            j.customer.name if j.customer else '',
            j.customer.company if j.customer and j.customer.company else '',
            j.job_type or '',
            j.assignee.name if j.assignee else '',
            j.status or '',
            j.priority or '',
            j.due_date.strftime('%d/%m/%Y') if j.due_date else '',
            j.amount_invoiced or 0,
            j.amount_received or 0,
            (j.amount_invoiced or 0) - (j.amount_received or 0),
            j.created_at.strftime('%d/%m/%Y') if j.created_at else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_tasks.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/jobs/add', methods=['GET', 'POST'])
@login_required
def add_job():
    if session.get('role') not in ['admin', 'operations']:
        flash('Access denied — only Operations can add tasks')
        return redirect(url_for('jobs'))
    customers = Customer.query.order_by(Customer.name).all()
    job_types = ServiceType.query.order_by(ServiceType.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    if request.method == 'POST':
        due = request.form.get('due_date')
        due_dt = datetime.strptime(due, '%Y-%m-%d') if due else None
        amount_invoiced = request.form.get('amount_invoiced') or 0
        assigned = request.form.get('assigned_to')
        job = Job(
            customer_id=int(request.form['customer_id']),
            job_type=request.form['job_type'],
            assigned_to=int(assigned) if assigned else None,
            due_date=due_dt,
            priority=request.form.get('priority', 'Medium'),
            internal_notes=request.form.get('internal_notes'),
            service_note=request.form.get('service_note', '').strip() or None,
            amount_invoiced=float(amount_invoiced),
            amount_received=0,
            num_persons=int(request.form.get('num_persons') or 1),
            created_by=session['user_id'],
            status='Pending Finance Approval'
        )
        db.session.add(job)
        db.session.commit()
        update = JobUpdate(job_id=job.id, status='Pending Finance Approval',
                           remark='Task created — awaiting finance approval',
                           staff_name=session['user_name'])
        db.session.add(update)
        # Process sub-tasks submitted inline
        st_titles = request.form.getlist('st_title[]')
        st_service_types = request.form.getlist('st_service_type[]')
        st_assigned_tos = request.form.getlist('st_assigned_to[]')
        st_due_dates = request.form.getlist('st_due_date[]')
        st_priorities = request.form.getlist('st_priority[]')
        st_amounts = request.form.getlist('st_amount[]')
        try:
          for i, title in enumerate(st_titles):
            if not title.strip():
                continue
            st_assigned = st_assigned_tos[i] if i < len(st_assigned_tos) and st_assigned_tos[i] else None
            st_due_str = st_due_dates[i] if i < len(st_due_dates) and st_due_dates[i] else None
            st_due = datetime.strptime(st_due_str, '%Y-%m-%d') if st_due_str else now_dubai() + timedelta(days=1)
            try:
                st_amt = float(st_amounts[i]) if i < len(st_amounts) and st_amounts[i] else 0
            except ValueError:
                st_amt = 0
            subtask = SubTask(
                job_id=job.id,
                title=title.strip(),
                service_type=st_service_types[i] if i < len(st_service_types) else None,
                assigned_to=int(st_assigned) if st_assigned else (job.assigned_to or session['user_id']),
                due_date=st_due,
                priority=st_priorities[i] if i < len(st_priorities) else 'Medium',
                amount=st_amt,
            )
            db.session.add(subtask)
            # Fold add-on revenue into the task's invoice so it flows through finance
            if st_amt:
                job.amount_invoiced = (job.amount_invoiced or 0) + st_amt
          db.session.commit()
        except Exception as e:
          db.session.rollback()
          print(f'SubTask error: {e}')

        # Handle additional tasks for same customer
        extra_types = request.form.getlist('extra_job_type[]')
        extra_assigned = request.form.getlist('extra_assigned_to[]')
        extra_due = request.form.getlist('extra_due_date[]')
        extra_priority = request.form.getlist('extra_priority[]')
        extra_amount = request.form.getlist('extra_amount[]')
        extra_persons = request.form.getlist('extra_persons[]')
        extra_notes = request.form.getlist('extra_notes[]')
        extra_service_notes = request.form.getlist('extra_service_note[]')

        for i, jt in enumerate(extra_types):
            if not jt: continue
            try: ea = int(extra_assigned[i]) if i < len(extra_assigned) and extra_assigned[i] else None
            except: ea = None
            try: ed = datetime.strptime(extra_due[i], '%Y-%m-%d') if i < len(extra_due) and extra_due[i] else now_dubai() + timedelta(days=1)
            except: ed = now_dubai() + timedelta(days=1)
            try: eamt = float(extra_amount[i]) if i < len(extra_amount) and extra_amount[i] else 0
            except: eamt = 0
            try: ep = int(extra_persons[i]) if i < len(extra_persons) and extra_persons[i] else 1
            except: ep = 1
            extra_job = Job(
                customer_id=job.customer_id,
                job_type=jt,
                assigned_to=ea,
                due_date=ed,
                priority=extra_priority[i] if i < len(extra_priority) else 'Medium',
                amount_invoiced=eamt,
                num_persons=ep,
                internal_notes=extra_notes[i] if i < len(extra_notes) else None,
                service_note=extra_service_notes[i].strip() if i < len(extra_service_notes) and extra_service_notes[i].strip() else None,
                status='Pending Finance Approval',
                created_by=session['user_id']
            )
            db.session.add(extra_job)

        db.session.commit()
        count = 1 + len([t for t in extra_types if t])
        flash(f'{count} task(s) created successfully')
        return redirect(url_for('jobs'))
    tomorrow = (now_dubai() + timedelta(days=1)).strftime('%Y-%m-%d')
    # Pass the raw dict and let Jinja's |tojson encode+escape it safely in the
    # template (avoids XSS via admin-entered service-type names).
    service_days = {jt.name: (getattr(jt, 'default_days', None) or 1) for jt in job_types}
    all_jobs = Job.query.order_by(Job.created_at.desc()).all()
    subtask_list_names = [t.title for t in SubTaskTemplate.query.order_by(SubTaskTemplate.sort_order, SubTaskTemplate.id).all()]
    return render_template('add_job.html', customers=customers, job_types=job_types, users=users, tomorrow=tomorrow, service_days_json=service_days, all_jobs=all_jobs, subtask_list_names=subtask_list_names)

@app.route('/jobs/<int:job_id>', methods=['GET', 'POST'])
@login_required
def job_detail(job_id):
    job = Job.query.get_or_404(job_id)
    now = now_dubai()
    role = session['role']
    # Sales and operations can view all tasks (not just assigned ones)
    # Only restrict if somehow a non-authorised role gets here
    if role not in ['admin', 'sales', 'operations', 'finance', 'staff']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    if request.method == 'POST':
        # Sales cannot update tasks at all
        if role == 'sales':
            flash('Sales cannot update task status. Contact Operations.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Closed — no updates except admin/finance
        if job.status == 'Closed' and role not in ['admin', 'finance']:
            flash('This task is closed. No further updates allowed.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Done/Pending Finance Close — no further updates from non-admin/finance
        if job.status in ['Done', 'Pending Finance Close'] and role not in ['admin', 'finance']:
            flash('Task is already marked Done. Contact Finance/Admin for changes.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Block sales/staff from updating if pending finance approval
        if job.status == 'Pending Finance Approval' and role in ['staff', 'sales']:
            flash('This task is pending finance approval. You cannot update it yet.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Block sales/staff from updating if pending finance close
        if job.status == 'Pending Finance Close' and role in ['staff', 'sales']:
            flash('Work is complete. Awaiting finance to close this task.')
            return redirect(url_for('job_detail', job_id=job_id))
        remark = request.form.get('remark', '').strip()
        if not remark:
            flash('Remark is required')
            return redirect(url_for('job_detail', job_id=job_id))
        new_status = request.form.get('status', job.status)
        # Rolling stages: non-admin can never move a task back to an earlier stage
        if (role != 'admin' and new_status != job.status and new_status in JOB_STAGE_RANK
                and JOB_STAGE_RANK[new_status] < job_stage_floor(job)):
            flash('Task stages only move forward — this task already passed that stage. Ask an admin if it was set by mistake.')
            return redirect(url_for('job_detail', job_id=job_id))
        if role == 'staff' and new_status == 'Pending Finance Approval':
            new_status = job.status
        # When ops marks Done → stays as Done, appears in Finance queue
        # Finance will verify payment and close the task
        if new_status == 'Done' and role not in ['admin', 'finance']:
            pass  # Keep as Done — Finance will close it
        job.status = new_status
        # Save completion fields when marking Done or Pending Finance Close
        if new_status in ['Done', 'Pending Finance Close']:
            if not job.completed_at:
                job.completed_at = now_dubai()
            job.final_remarks = request.form.get('final_remarks') or None
            job.future_work_notes = request.form.get('future_work_notes') or None
            # Log completion to timeline
            completion_note = 'Task completed.'
            if job.final_remarks: completion_note += f' Remarks: {job.final_remarks}'
            update_completion = JobUpdate(job_id=job.id, status=new_status, remark=completion_note, staff_name=session['user_name'])
            db.session.add(update_completion)
        update = JobUpdate(job_id=job.id, status=new_status,
                           status_note=request.form.get('status_note', '').strip()[:100] or None,
                           remark=remark, staff_name=session['user_name'])
        db.session.add(update)
        db.session.commit()
        flash('Task updated')
        return redirect(url_for('job_detail', job_id=job_id))
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    service_types = ServiceType.query.order_by(ServiceType.name).all()
    # All jobs for same customer (for multi-task timeline)
    sibling_jobs = Job.query.filter_by(customer_id=job.customer_id).order_by(Job.created_at.asc()).all()
    partners = Partner.query.filter_by(active=True).order_by(Partner.name).all()
    quick_replies = QuickReply.query.filter(
        (QuickReply.staff_id == session.get('user_id')) | (QuickReply.is_global == True)
    ).order_by(QuickReply.label).all()
    subtask_list_names = [t.title for t in SubTaskTemplate.query.order_by(SubTaskTemplate.sort_order, SubTaskTemplate.id).all()]
    # Stages already passed — hidden from the dropdown for non-admin (rolling stages)
    locked_stages = []
    if role != 'admin':
        floor = job_stage_floor(job)
        locked_stages = [s for s, r in JOB_STAGE_RANK.items() if r < floor]
    return render_template('job_detail.html', job=job, now=now,
                           statuses=JOB_STATUSES, users=users, locked_stages=locked_stages,
                           service_types=service_types, timedelta=timedelta,
                           sibling_jobs=sibling_jobs, partners=partners,
                           wa_templates=wa_send_context(job=job), quick_replies=quick_replies,
                           subtask_list_names=subtask_list_names)

@app.route('/subtasks/<int:sub_id>/toggle', methods=['POST'])
@login_required
def subtask_toggle(sub_id):
    """Mark a sub-task done / not-done from the task detail page."""
    st = SubTask.query.get_or_404(sub_id)
    if st.status == 'Done':
        st.status = 'Pending'
        st.completed_at = None
    else:
        st.status = 'Done'
        st.completed_at = now_dubai()
    db.session.commit()
    return redirect(url_for('job_detail', job_id=st.job_id) + '#steps')

@app.route('/jobs/<int:job_id>/status-update', methods=['POST'])
@login_required
def job_status_update(job_id):
    """Send the customer a WhatsApp status update (approved template, logged).
    Dormant until the 'status_update' template is activated in WhatsApp Templates."""
    from whatsapp_webhook import send_template, log_message, normalize_phone
    job = Job.query.get_or_404(job_id)
    back = request.referrer or url_for('job_detail', job_id=job_id)
    tpl_name = get_setting('wa_status_template', 'status_update') or 'status_update'
    tpl = wa_template_active(tpl_name)
    if not tpl:
        flash(f'Status-update template isn\'t active yet — activate "{tpl_name}" in WhatsApp → Templates first.', 'error')
        return redirect(back)
    cust = job.customer
    to = normalize_phone((cust.whatsapp or cust.mobile or cust.phone or cust.phone2) if cust else '')
    if not to:
        flash('No WhatsApp number on record for this customer.', 'error')
        return redirect(back)
    status_text = (request.form.get('status_text') or '').strip()
    if not status_text:
        flash('Pick a status to send.', 'error')
        return redirect(back)
    first = ((cust.contact_person or cust.name or 'there').split() or ['there'])[0]
    service = job.job_type or 'application'
    params = [first, service, status_text]
    wam = send_template(to, tpl.meta_name, params=params, lang=tpl.lang or 'en')
    body = tpl.body_preview or ''
    for n, v in enumerate(params, start=1):
        body = body.replace('{{%d}}' % n, v)
    log_message(to, 'out', body, msg_type='template', wam_id=wam,
                handled_by=session.get('user_name', 'staff'),
                status='sent' if wam else 'failed', customer_id=cust.id if cust else None)
    flash('Status update sent on WhatsApp.' if wam else 'WhatsApp send failed — check the number/template.',
          'success' if wam else 'error')
    return redirect(back)

@app.route('/jobs/<int:job_id>/subtasks/add', methods=['POST'])
@login_required
def add_subtask(job_id):
    """Add one sub-task/step to a task. Person + due default to the main task's."""
    job = Job.query.get_or_404(job_id)
    title = (request.form.get('title') or '').strip()
    if title:
        due_str = request.form.get('due_date')
        try:
            due = datetime.strptime(due_str, '%Y-%m-%d') if due_str else (job.due_date or now_dubai() + timedelta(days=1))
        except ValueError:
            due = job.due_date or now_dubai() + timedelta(days=1)
        assigned = request.form.get('assigned_to')
        try:
            amt = float(request.form.get('amount') or 0)
        except ValueError:
            amt = 0
        st = SubTask(job_id=job.id, title=title[:200],
                     assigned_to=int(assigned) if assigned else (job.assigned_to or session['user_id']),
                     due_date=due, priority='Medium', amount=amt)
        db.session.add(st)
        # Fold add-on revenue into the task's invoice so it flows through finance
        if amt:
            job.amount_invoiced = (job.amount_invoiced or 0) + amt
        db.session.commit()
        flash('Step added.')
    return redirect(url_for('job_detail', job_id=job_id) + '#steps')

@app.route('/subtasks/<int:sub_id>/delete', methods=['POST'])
@login_required
def subtask_delete(sub_id):
    st = SubTask.query.get_or_404(sub_id)
    jid = st.job_id
    # Reverse the add-on revenue from the task's invoice when the step is removed
    if st.amount and st.job:
        st.job.amount_invoiced = max((st.job.amount_invoiced or 0) - st.amount, 0)
    db.session.delete(st)
    db.session.commit()
    return redirect(url_for('job_detail', job_id=jid) + '#steps')

@app.route('/jobs/<int:job_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_job(job_id):
    job = Job.query.get_or_404(job_id)
    if session.get('role') not in ['admin', 'operations']:
        flash('Access denied — only Operations can edit tasks')
        return redirect(url_for('job_detail', job_id=job_id))
    if job.status == 'Closed' and session['role'] != 'admin':
        flash('Closed tasks cannot be edited.')
        return redirect(url_for('job_detail', job_id=job_id))
    # Staff can only edit tasks assigned to them
    if session['role'] in ['sales', 'staff'] and job.assigned_to != session['user_id']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    customers = Customer.query.order_by(Customer.name).all()
    job_types = ServiceType.query.order_by(ServiceType.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    if request.method == 'POST':
        job.job_type = request.form['job_type']
        job.customer_id = int(request.form['customer_id'])
        assigned = request.form.get('assigned_to')
        job.assigned_to = int(assigned) if assigned else None
        due = request.form.get('due_date')
        job.due_date = datetime.strptime(due, '%Y-%m-%d') if due else None
        job.priority = request.form.get('priority', 'Medium')
        job.internal_notes = request.form.get('internal_notes')
        job.service_note = request.form.get('service_note', '').strip() or None
        if request.form.get('num_persons'):
            job.num_persons = int(request.form.get('num_persons'))
        try:
            ai = request.form.get('amount_invoiced')
            ar = request.form.get('amount_received')
            if ai: job.amount_invoiced = float(ai)
            if ar: job.amount_received = float(ar)
        except:
            pass
        db.session.commit()
        flash('Task updated')
        return redirect(url_for('job_detail', job_id=job_id))
    return render_template('edit_job.html', job=job, customers=customers,
                           job_types=job_types, users=users, statuses=JOB_STATUSES)

@app.route('/jobs/<int:job_id>/delete', methods=['POST'])
@login_required
def delete_job(job_id):
    if session['role'] not in ['admin', 'operations']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    job = Job.query.get_or_404(job_id)
    SubTask.query.filter_by(job_id=job_id).delete()
    JobUpdate.query.filter_by(job_id=job_id).delete()
    PartialRevenue.query.filter_by(job_id=job_id).delete()
    db.session.delete(job)
    db.session.commit()
    flash('Task deleted')
    return redirect(url_for('jobs'))

@app.route('/jobs/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_jobs():
    if session['role'] not in ['admin', 'operations']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    ids = request.form.getlist('job_ids')
    if not ids:
        flash('No tasks selected')
        return redirect(url_for('jobs'))
    count = 0
    for job_id in ids:
        job = Job.query.get(int(job_id))
        if job:
            SubTask.query.filter_by(job_id=job.id).delete()
            JobUpdate.query.filter_by(job_id=job.id).delete()
            PartialRevenue.query.filter_by(job_id=job.id).delete()
            db.session.delete(job)
            count += 1
    db.session.commit()
    flash(f'{count} task(s) deleted')
    return redirect(url_for('jobs'))

# ── Finance ───────────────────────────────────────────────────────────────────

@app.route('/jobs/<int:job_id>/approve', methods=['POST'])
@login_required
@finance_required
def approve_job(job_id):
    job = Job.query.get_or_404(job_id)
    amount_invoiced = request.form.get('amount_invoiced', '').strip()
    amount_received = request.form.get('amount_received', '').strip()
    try:
        if amount_invoiced:
            job.amount_invoiced = float(amount_invoiced)
        if amount_received:
            job.amount_received = float(amount_received)
    except:
        pass
    # If task was Done, Finance is closing it; if Pending Finance Approval, Finance is approving it
    if job.status == 'Done':
        job.status = 'Closed'
        job.completed_at = now_dubai()
    else:
        job.status = 'Assigned'
    job.finance_approved_by = session['user_id']
    job.finance_approved_at = now_dubai()
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        job.finance_notes = notes  # save to job record
    action = 'Closed' if job.status == 'Closed' else 'Approved'
    remark = f'{action} by Finance. Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f}'
    if notes:
        remark += f'. Notes: {notes}'
    update = JobUpdate(job_id=job.id, status=job.status, remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    msg = 'Task closed successfully.' if job.status == 'Closed' else 'Task approved and assigned to staff.'
    flash(msg)
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/jobs/<int:job_id>/payment', methods=['POST'])
@login_required
@finance_required
def update_payment(job_id):
    job = Job.query.get_or_404(job_id)
    try:
        job.amount_invoiced = float(request.form.get('amount_invoiced') or job.amount_invoiced or 0)
        job.amount_received = float(request.form.get('amount_received') or job.amount_received or 0)
    except:
        pass
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        # Replace finance notes (don't append)
        job.finance_notes = notes
    remark = f'Payment updated. Invoiced: AED {job.amount_invoiced:,.0f} / Received: AED {job.amount_received:,.0f}'
    if notes:
        remark += f'. Notes: {notes}'
    update = JobUpdate(job_id=job.id, status=job.status, remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    flash('Payment updated.')
    return redirect(request.referrer or url_for('jobs'))


@app.route('/jobs/<int:job_id>/close', methods=['POST'])
@login_required
@finance_required
def close_job(job_id):
    job = Job.query.get_or_404(job_id)
    
    # Update invoice and received amounts
    try:
        ai = request.form.get('amount_invoiced')
        ar = request.form.get('amount_received')
        if ai: job.amount_invoiced = float(ai)
        if ar: job.amount_received = float(ar)
    except:
        pass
    
    # Handle partner commission choice (mandatory)
    partner_choice = request.form.get('partner_commission_expected')
    
    if partner_choice == 'no':
        # REGULAR TASK - Revenue counted immediately
        try:
            rev = request.form.get('revenue')
            if rev:
                job.revenue = float(rev)
                # Revenue is cash-basis, dated by when the work/payment actually
                # happened — defaults to today but finance can backdate it
                # (e.g. entering a late June closure on July 1st).
                rev_date_str = request.form.get('revenue_date', '').strip()
                if rev_date_str:
                    job.revenue_date = datetime.strptime(rev_date_str, '%Y-%m-%d').date()
                else:
                    job.revenue_date = now_dubai().date()
            else:
                flash('Revenue is required for regular tasks.', 'error')
                return redirect(url_for('job_detail', job_id=job_id))
        except:
            flash('Invalid revenue amount.', 'error')
            return redirect(url_for('job_detail', job_id=job_id))
        
        job.partner_commission_expected = False
        job.partner_name = None
        job.partner_amount = None
        job.partner_due_date = None
        job.partner_status = None
        job.status = 'Closed'
        
        remark = f'Task CLOSED by Finance. Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Revenue: AED {job.revenue:,.0f} (counted for {job.revenue_date.strftime("%B %Y")})'
        
    elif partner_choice == 'yes':
        # PARTNER COMMISSION TASK - Revenue = 0 until partner pays
        partner_name = request.form.get('partner_name')
        new_partner_name = request.form.get('new_partner_name', '').strip()
        
        # Handle new partner creation
        if partner_name == '__ADD_NEW__':
            if not new_partner_name:
                flash('Please enter new partner name.', 'error')
                return redirect(url_for('job_detail', job_id=job_id))
            # Create new partner
            try:
                new_partner = Partner(name=new_partner_name)
                db.session.add(new_partner)
                db.session.flush()  # Get the ID without committing
                partner_name = new_partner_name
            except:
                flash('Partner name already exists or invalid.', 'error')
                return redirect(url_for('job_detail', job_id=job_id))
        
        if not partner_name or partner_name == '__ADD_NEW__':
            flash('Please select a partner.', 'error')
            return redirect(url_for('job_detail', job_id=job_id))
        
        try:
            partner_amount = float(request.form.get('partner_amount'))
            partner_due_date = request.form.get('partner_due_date')
            if not partner_due_date:
                raise ValueError("Due date required")
            partner_due_date = datetime.strptime(partner_due_date, '%Y-%m-%d').date()
        except:
            flash('Partner commission amount and due date are required.', 'error')
            return redirect(url_for('job_detail', job_id=job_id))
        
        job.partner_commission_expected = True
        job.partner_name = partner_name
        job.partner_amount = partner_amount
        job.partner_due_date = partner_due_date
        job.partner_status = 'Pending'
        job.revenue = 0  # Revenue NOT counted yet
        # NOTE: intentionally NOT 'Closed' — the task stays active/visible
        # (not treated as closed anywhere) until Finance marks the partner
        # commission as received, which is when it truly becomes 'Closed'.
        job.status = 'Pending Partner Commission'

        remark = f'Finance settled with customer. Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Awaiting partner commission from {partner_name}: AED {partner_amount:,.0f}. Task remains active until commission is received.'

    elif partner_choice == 'complimentary':
        # COMPLIMENTARY TASK - free service, no revenue, no partner commission
        job.partner_commission_expected = False
        job.partner_name = None
        job.partner_amount = None
        job.partner_due_date = None
        job.partner_status = None
        job.revenue = 0
        job.revenue_date = now_dubai().date()
        job.status = 'Closed'

        remark = f'Task CLOSED by Finance as COMPLIMENTARY (free service). Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Revenue: AED 0'

    else:
        flash('Please select whether partner commission is expected.', 'error')
        return redirect(url_for('job_detail', job_id=job_id))
    
    # Finance notes
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        job.finance_notes = notes
        remark += f' Notes: {notes}'
    
    # Create update record
    update = JobUpdate(job_id=job.id, status=job.status, remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    
    flash('Task closed successfully.')
    return redirect(url_for('dashboard'))

@app.route('/jobs/<int:job_id>/edit_finance', methods=['POST'])
@login_required
@finance_required
def edit_finance(job_id):
    job = Job.query.get_or_404(job_id)
    if job.status != 'Closed':
        flash('Can only edit finance details for closed tasks.')
        return redirect(url_for('job_detail', job_id=job_id))
    
    old_invoiced = job.amount_invoiced or 0
    old_received = job.amount_received or 0
    old_revenue = job.revenue or 0
    old_revenue_date = job.revenue_date
    old_created_at = job.created_at
    
    try:
        ai = request.form.get('amount_invoiced')
        ar = request.form.get('amount_received')
        rev = request.form.get('revenue')
        if ai: job.amount_invoiced = float(ai)
        if ar: job.amount_received = float(ar)
        if rev:
            job.revenue = float(rev)
            if not job.revenue_date:
                job.revenue_date = now_dubai().date()
        # Revenue date is directly editable (e.g. correcting which month's
        # revenue this counts toward), independent of the revenue amount.
        rev_date_str = request.form.get('revenue_date', '').strip()
        if rev_date_str:
            job.revenue_date = datetime.strptime(rev_date_str, '%Y-%m-%d').date()
    except:
        flash('Invalid finance values.')
        return redirect(url_for('job_detail', job_id=job_id))
    
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        # Replace finance notes (don't append)
        job.finance_notes = notes
    
    remark = f'Finance details EDITED by {session["user_name"]}. Previous — Invoiced: AED {old_invoiced:,.0f} / Received: AED {old_received:,.0f} / Revenue: AED {old_revenue:,.0f}. Updated — Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Revenue: AED {job.revenue or 0:,.0f}'
    if old_revenue_date != job.revenue_date:
        remark += f' / Revenue Date: {job.revenue_date.strftime("%d-%b-%Y") if job.revenue_date else "None"}'
    if old_created_at.date() != job.created_at.date():
        remark += f' / Created: {job.created_at.strftime("%d-%b-%Y")}'
    
    if notes:
        remark += f'. Notes: {notes}'
    
    update = JobUpdate(job_id=job.id, status='Closed', remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    flash('Finance details updated successfully.')
    return redirect(url_for('job_detail', job_id=job_id))

@app.route('/jobs/<int:job_id>/partial_revenue/add', methods=['POST'])
@login_required
def add_partial_revenue(job_id):
    if session['role'] not in ['finance', 'admin']:
        flash('Only Finance can record partial revenue')
        return redirect(url_for('job_detail', job_id=job_id))
    
    job = Job.query.get_or_404(job_id)
    amount = request.form.get('amount', type=float)
    revenue_date_str = request.form.get('revenue_date')
    notes = request.form.get('notes', '').strip()
    
    if not amount or amount <= 0:
        flash('Please enter a valid amount')
        return redirect(url_for('job_detail', job_id=job_id))
    
    if not revenue_date_str:
        flash('Please select a revenue date')
        return redirect(url_for('job_detail', job_id=job_id))
    
    try:
        revenue_date = datetime.strptime(revenue_date_str, '%Y-%m-%d').date()
    except:
        flash('Invalid date format')
        return redirect(url_for('job_detail', job_id=job_id))
    
    # Check total partial revenue doesn't exceed received amount
    existing_partial = sum(pr.amount for pr in job.partial_revenues)
    if existing_partial + amount > (job.amount_received or 0):
        flash(f'Total partial revenue ({existing_partial + amount:,.0f}) cannot exceed received amount ({job.amount_received:,.0f})')
        return redirect(url_for('job_detail', job_id=job_id))
    
    partial = PartialRevenue(
        job_id=job_id,
        amount=amount,
        revenue_date=revenue_date,
        notes=notes,
        recorded_by=session['user_id']
    )
    db.session.add(partial)
    db.session.commit()
    
    flash(f'Partial revenue of AED {amount:,.0f} recorded successfully')
    return redirect(url_for('job_detail', job_id=job_id))

# ── Daily Activity Log ────────────────────────────────────────────────────────

# ACTIVITIES loaded from DB — see get_activities()
ACTIVITY_DEFAULTS = [
    ('calls_existing',       'Calls to Existing/Potential Clients', 30),
    ('calls_cold',           'Cold Calling to Customer List',       30),
    ('dm_instagram',         'Instagram Direct Messages',           30),
    ('dm_facebook',          'Facebook Messages',                   30),
    ('dm_linkedin',          'LinkedIn Messages',                   30),
    ('posts_social',         'Social Media Posts (IG/FB/LinkedIn)', 2),
    ('videos_instagram',     'Instagram Video (Cross-post)',        1),
    ('linkedin_writing',     'LinkedIn Writing/Articles',           1),
    ('whatsapp_prospecting', 'WhatsApp Prospecting',               30),
    ('community_active',     'Active in Communities',               2),
    ('google_reviews',       'Google Review Collection',            6),
    ('real_estate_relations','Real Estate Agent Relationships',     2),
    ('content_marketing',    'Content for Marketing',               2),
    ('referral_building',    'Referral Building',                   2),
    ('networking_activities','Networking/Community Activities',     1),
    ('networking_events',    'Attended Networking Event',           1),
]

def get_activities():
    try:
        types = ActivityType.query.filter_by(active=True).order_by(ActivityType.sort_order, ActivityType.id).all()
        return [(t.field_key, t.label, t.weekly_target) for t in types]
    except:
        return ACTIVITY_DEFAULTS

@app.route('/activity')
@login_required
def activity_log():
    if session['role'] not in ['sales', 'admin']:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    now = now_dubai()
    week_start = (now - timedelta(days=now.weekday())).date()
    from_date = request.args.get('from', week_start.strftime('%Y-%m-%d'))
    to_date = request.args.get('to', now.date().strftime('%Y-%m-%d'))
    view = request.args.get('view', 'week')

    from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
    to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()

    try:
        if session['role'] == 'admin':
            sales_users = User.query.filter(User.role == 'sales', User.active == True).all()
            logs = ActivityLog.query.filter(
                ActivityLog.log_date >= from_dt,
                ActivityLog.log_date <= to_dt
            ).all()
        else:
            sales_users = [User.query.get(session['user_id'])]
            logs = ActivityLog.query.filter_by(user_id=session['user_id']).filter(
                ActivityLog.log_date >= from_dt,
                ActivityLog.log_date <= to_dt
            ).all()
    except Exception as e:
        # Table may not exist yet — use empty data
        logs = []
        sales_users = [] if session['role'] == 'admin' else [User.query.get(session['user_id'])]
        try:
            if session['role'] == 'admin':
                sales_users = User.query.filter(User.role == 'sales', User.active == True).all()
        except:
            pass

    # Build summary per user
    user_summaries = {}
    for u in sales_users:
        user_logs = [l for l in logs if l.user_id == u.id]
        summary = {}
        for field, label, target in get_activities():
            total = sum(getattr(l, field, 0) or 0 for l in user_logs)
            days = (to_dt - from_dt).days + 1
            weeks = max(1, days / 6)  # 6-day UAE working week (Sat–Thu)
            period_target = round(target * weeks)
            pct = round((total / period_target * 100) if period_target > 0 else 0)
            summary[field] = {'total': total, 'target': period_target, 'pct': pct}
        user_summaries[u.id] = {'user': u, 'summary': summary, 'logs': user_logs}

    # Today's log for current user (for the entry form)
    today_log = None
    if session['role'] == 'sales':
        today_log = ActivityLog.query.filter_by(
            user_id=session['user_id'],
            log_date=now.date()
        ).first()

    try:
        activity_types = ActivityType.query.filter_by(active=True).order_by(ActivityType.sort_order, ActivityType.id).all()
    except Exception:
        activity_types = []
    return render_template('activity_log.html',
                           activities=get_activities(),
                           activity_types=activity_types,
                           user_summaries=user_summaries,
                           sales_users=sales_users,
                           today_log=today_log,
                           from_date=from_date, to_date=to_date,
                           view=view, now=now)

@app.route('/activity/log', methods=['POST'])
@login_required
def save_activity():
    if session['role'] not in ['sales', 'admin']:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    log_date_str = request.form.get('log_date', now_dubai().date().strftime('%Y-%m-%d'))
    log_date = datetime.strptime(log_date_str, '%Y-%m-%d').date()
    # Admin can log for any user
    user_id = int(request.form.get('user_id_override') or session['user_id'])
    if request.form.get('user_id_override') and session['role'] != 'admin':
        flash('Access denied'); return redirect(url_for('activity_log'))

    # Upsert — update if exists for this date
    log = ActivityLog.query.filter_by(user_id=user_id, log_date=log_date).first()
    if not log:
        log = ActivityLog(user_id=user_id, log_date=log_date)
        db.session.add(log)

    for field, label, target in get_activities():
        val = request.form.get(field, '0').strip()
        try:
            setattr(log, field, int(val) if val else 0)
        except:
            setattr(log, field, 0)
    log.off_day = request.form.get('off_day', '') or None
    log.notes = request.form.get('notes', '')
    log.updated_at = now_dubai()
    db.session.commit()
    flash(f'Activity log saved for {log_date.strftime("%d %b %Y")}')
    return redirect(url_for('activity_log'))



@app.route('/activity/<int:log_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_activity_log(log_id):
    log = ActivityLog.query.get_or_404(log_id)
    if session['role'] != 'admin' and log.user_id != session['user_id']:
        flash('Access denied')
        return redirect(url_for('activity_log'))
    if request.method == 'POST':
        for field, label, target in get_activities():
            val = request.form.get(field, '0').strip()
            try: setattr(log, field, int(val) if val else 0)
            except: setattr(log, field, 0)
        log.off_day = request.form.get('off_day', '') or None
        log.notes = request.form.get('notes', '')
        log.updated_at = now_dubai()
        db.session.commit()
        flash(f'Activity log updated for {log.log_date.strftime("%d %b %Y")}')
        return redirect(url_for('activity_log'))
    return redirect(url_for('activity_log'))

@app.route('/activity/<int:log_id>/delete', methods=['POST'])
@login_required
def delete_activity_log(log_id):
    log = ActivityLog.query.get_or_404(log_id)
    # Admin can delete any, sales can delete their own
    if session['role'] != 'admin' and log.user_id != session['user_id']:
        flash('Access denied')
        return redirect(url_for('activity_log'))
    db.session.delete(log)
    db.session.commit()
    flash('Activity log entry deleted')
    return redirect(url_for('activity_log'))


# ── Admin — Activity Types ────────────────────────────────────────────────────

@app.route('/admin/activity-type/add', methods=['POST'])
@login_required
@admin_required
def admin_add_activity_type():
    label = request.form.get('label', '').strip()
    target = request.form.get('daily_target', '1').strip()
    if not label:
        flash('Activity name is required')
        return redirect(url_for('activity_log'))
    # Generate a safe field_key from label
    import re as re_mod
    field_key = re_mod.sub(r'[^a-z0-9]', '_', label.lower())[:40]
    field_key = re_mod.sub(r'_+', '_', field_key).strip('_')
    # Ensure unique
    base_key = field_key
    counter = 1
    while ActivityType.query.filter_by(field_key=field_key).first():
        field_key = f'{base_key}_{counter}'
        counter += 1
    try:
        target_val = float(target)
    except:
        target_val = 1.0
    max_order = db.session.query(db.func.max(ActivityType.sort_order)).scalar() or 0
    at = ActivityType(field_key=field_key, label=label, weekly_target=target_val, sort_order=max_order+1)
    db.session.add(at)
    db.session.commit()
    flash(f'Activity "{label}" added')
    return redirect(url_for('activity_log'))

@app.route('/admin/activity-type/<int:type_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_activity_type(type_id):
    at = ActivityType.query.get_or_404(type_id)
    at.label = request.form.get('label', at.label).strip()
    try:
        at.weekly_target = float(request.form.get('weekly_target', at.weekly_target))
    except:
        pass
    db.session.commit()
    flash(f'Activity updated')
    return redirect(url_for('activity_log'))

@app.route('/admin/activity-type/<int:type_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_activity_type(type_id):
    at = ActivityType.query.get_or_404(type_id)
    at.active = False  # Soft delete — preserve historical data
    db.session.commit()
    flash(f'Activity "{at.label}" removed')
    return redirect(url_for('activity_log'))


# ── Admin Edit Routes ─────────────────────────────────────────────────────────

@app.route('/admin/service/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_service(item_id):
    item = Service.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Service updated')
    return redirect(url_for('admin_panel') + '#services')

@app.route('/admin/source/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_source(item_id):
    item = Source.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Source updated')
    return redirect(url_for('admin_panel') + '#sources')

@app.route('/admin/campaign/add', methods=['POST'])
@login_required
@admin_required
def admin_add_campaign():
    name = request.form.get('name', '').strip()
    if name:
        existing = Campaign.query.filter_by(name=name).first()
        if existing:
            flash('Campaign already exists')
        else:
            db.session.add(Campaign(name=name))
            db.session.commit()
            flash(f'Campaign "{name}" added')
    return redirect(url_for('admin_panel') + '#campaigns')

@app.route('/admin/campaign/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_campaign(item_id):
    item = Campaign.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Campaign updated')
    return redirect(url_for('admin_panel') + '#campaigns')

@app.route('/admin/campaign/<int:item_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_campaign(item_id):
    item = Campaign.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash(f'Campaign "{item.name}" removed')
    return redirect(url_for('admin_panel') + '#campaigns')

@app.route('/admin/jobtype/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_jobtype(item_id):
    item = ServiceType.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
    try:
        item.default_days = int(request.form.get('default_days', 1))
    except:
        pass
    db.session.commit()
    flash('Service type updated')
    return redirect(url_for('admin_panel') + '#service-types')

@app.route('/admin/doctype/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_doctype(item_id):
    item = DocType.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Document type updated')
    return redirect(url_for('admin_panel'))

# ── Admin — Job Types ─────────────────────────────────────────────────────────

@app.route('/admin/jobtype/add', methods=['POST'])
@login_required
@admin_required
def admin_add_jobtype():
    name = request.form.get('name', '').strip()
    if name:
        if not ServiceType.query.filter_by(name=name).first():
            try:
                days = int(request.form.get('default_days', 1))
            except:
                days = 1
            new_jt = ServiceType(name=name)
            new_jt.default_days = days
            db.session.add(new_jt)
            db.session.commit()
            flash(f'Service type "{name}" added')
        else:
            flash('Job type already exists')
    return redirect(url_for('admin_panel') + '#service-types')

@app.route('/admin/jobtype/<int:jobtype_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_jobtype(jobtype_id):
    jt = ServiceType.query.get_or_404(jobtype_id)
    db.session.delete(jt)
    db.session.commit()
    flash(f'Job type "{jt.name}" removed')
    return redirect(url_for('admin_panel'))

# ── Documents ─────────────────────────────────────────────────────────────────

# ─────────────────────────── Companies ───────────────────────────
@app.route('/companies')
@login_required
def companies():
    now = now_dubai()
    search = (request.args.get('search') or '').strip().lower()
    company_list = Company.query.order_by(Company.name).all()
    if search:
        company_list = [c for c in company_list if
                        search in (c.name or '').lower() or
                        search in (c.contact_person or '').lower() or
                        search in (c.trade_license_no or '').lower()]
    def soonest(c):
        days = [(d.expiry_date.date() - now.date()).days for d in c.documents if d.expiry_date]
        return min(days) if days else None
    rows = [{'c': c, 'doc_count': len(c.documents), 'soonest': soonest(c)} for c in company_list]
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('companies.html', rows=rows, now=now, search=search, customers=customers)

@app.route('/companies/add', methods=['POST'])
@login_required
def add_company():
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('Company name is required')
        return redirect(url_for('companies'))
    c = Company(
        name=name,
        customer_id=(int(request.form['customer_id']) if request.form.get('customer_id') else None),
        contact_person=request.form.get('contact_person') or None,
        phone=request.form.get('phone') or None,
        email=request.form.get('email') or None,
        trade_license_no=request.form.get('trade_license_no') or None,
        authority=request.form.get('authority') or None,
        address=request.form.get('address') or None,
        notes=request.form.get('notes') or None,
        created_by=session.get('user_id'),
    )
    db.session.add(c)
    db.session.commit()
    flash(f'Company "{name}" added')
    return redirect(url_for('company_detail', company_id=c.id))

@app.route('/companies/<int:company_id>')
@login_required
def company_detail(company_id):
    now = now_dubai()
    company = Company.query.get_or_404(company_id)
    docs = sorted(company.documents, key=lambda d: (d.expiry_date or datetime.max))
    doc_types = DocType.query.order_by(DocType.name).all()
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('company_detail.html', company=company, docs=docs,
                           doc_types=doc_types, customers=customers, now=now)

@app.route('/companies/<int:company_id>/edit', methods=['POST'])
@login_required
def edit_company(company_id):
    c = Company.query.get_or_404(company_id)
    c.name = (request.form.get('name') or c.name).strip()
    c.customer_id = int(request.form['customer_id']) if request.form.get('customer_id') else None
    c.contact_person = request.form.get('contact_person') or None
    c.phone = request.form.get('phone') or None
    c.email = request.form.get('email') or None
    c.trade_license_no = request.form.get('trade_license_no') or None
    c.authority = request.form.get('authority') or None
    c.address = request.form.get('address') or None
    c.notes = request.form.get('notes') or None
    c.alerts_enabled = bool(request.form.get('alerts_enabled'))
    c.alert_email = request.form.get('alert_email') or None
    c.alert_whatsapp = request.form.get('alert_whatsapp') or None
    db.session.commit()
    flash('Company updated')
    return redirect(url_for('company_detail', company_id=c.id))

@app.route('/companies/<int:company_id>/toggle-alerts', methods=['POST'])
@login_required
def toggle_company_alerts(company_id):
    c = Company.query.get_or_404(company_id)
    c.alerts_enabled = not c.alerts_enabled
    db.session.commit()
    flash(('🔔 Alerts ON' if c.alerts_enabled else '🔕 Alerts OFF') + f' for {c.name}')
    return redirect(url_for('company_detail', company_id=c.id))

@app.route('/companies/<int:company_id>/documents/add', methods=['POST'])
@login_required
def add_company_document(company_id):
    company = Company.query.get_or_404(company_id)
    expiry = request.form.get('expiry_date')
    expiry_dt = datetime.strptime(expiry, '%Y-%m-%d') if expiry else None
    file_name = file_url = public_id = None
    if 'document_file' in request.files:
        f = request.files['document_file']
        if f and f.filename:
            file_name = f.filename
            file_url, public_id = upload_to_cloudinary(f)
            if not file_url:
                flash('⚠️ File could not be uploaded — document saved without attachment.', 'warning')
    doc = Document(
        doc_type=request.form['doc_type'],
        belongs_to='Company',
        owner_name=company.name,
        company_id=company.id,
        customer_id=company.customer_id,
        expiry_date=expiry_dt,
        notes=request.form.get('notes'),
        file_name=file_name,
        file_url=file_url,
        cloudinary_public_id=public_id,
        uploaded_by=session['user_id'],
        added_by=session['user_name'],
    )
    db.session.add(doc)
    db.session.commit()
    flash('Document added')
    return redirect(url_for('company_detail', company_id=company.id))

@app.route('/companies/<int:company_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_company(company_id):
    c = Company.query.get_or_404(company_id)
    Document.query.filter_by(company_id=c.id).update({'company_id': None}, synchronize_session=False)
    name = c.name
    db.session.delete(c)
    db.session.commit()
    flash(f'Company "{name}" deleted')
    return redirect(url_for('companies'))

# ─────────────────────────── Employees ───────────────────────────
@app.route('/customers/<int:customer_id>/employees/add', methods=['POST'])
@login_required
def add_employee(customer_id):
    Customer.query.get_or_404(customer_id)
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('Employee name is required')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    e = Employee(
        customer_id=customer_id, name=name,
        designation=request.form.get('designation') or None,
        nationality=request.form.get('nationality') or None,
        mobile=request.form.get('mobile') or None,
        email=request.form.get('email') or None,
        status=request.form.get('status') or 'Active',
    )
    dob = request.form.get('date_of_birth'); jd = request.form.get('join_date')
    e.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date() if dob else None
    e.join_date = datetime.strptime(jd, '%Y-%m-%d').date() if jd else None
    db.session.add(e)
    db.session.commit()
    flash(f'Employee "{name}" added')
    return redirect(url_for('employee_detail', employee_id=e.id))

@app.route('/employees/<int:employee_id>')
@login_required
def employee_detail(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    now = now_dubai()
    docs = sorted(list(emp.documents), key=lambda d: (d.expiry_date or datetime.max))
    doc_types = DocType.query.order_by(DocType.name).all()
    return render_template('employee_detail.html', emp=emp, docs=docs, doc_types=doc_types,
                           now=now, today=now.date())

@app.route('/employees/<int:employee_id>/edit', methods=['POST'])
@login_required
def edit_employee(employee_id):
    e = Employee.query.get_or_404(employee_id)
    e.name = (request.form.get('name') or e.name).strip()
    e.designation = request.form.get('designation') or None
    e.nationality = request.form.get('nationality') or None
    e.mobile = request.form.get('mobile') or None
    e.email = request.form.get('email') or None
    e.status = request.form.get('status') or 'Active'
    dob = request.form.get('date_of_birth'); jd = request.form.get('join_date')
    e.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date() if dob else None
    e.join_date = datetime.strptime(jd, '%Y-%m-%d').date() if jd else None
    db.session.commit()
    flash('Employee updated')
    return redirect(url_for('employee_detail', employee_id=e.id))

@app.route('/employees/<int:employee_id>/delete', methods=['POST'])
@login_required
def delete_employee(employee_id):
    e = Employee.query.get_or_404(employee_id)
    cid = e.customer_id
    Document.query.filter_by(employee_id=e.id).delete(synchronize_session=False)
    db.session.delete(e)
    db.session.commit()
    flash('Employee removed')
    return redirect(url_for('customer_detail', customer_id=cid))

@app.route('/employees/<int:employee_id>/documents/add', methods=['POST'])
@login_required
def add_employee_document(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    expiry = request.form.get('expiry_date')
    expiry_dt = datetime.strptime(expiry, '%Y-%m-%d') if expiry else None
    file_name = file_url = public_id = None
    if 'document_file' in request.files:
        f = request.files['document_file']
        if f and f.filename:
            file_name = f.filename
            file_url, public_id = upload_to_cloudinary(f)
            if not file_url:
                flash('⚠️ File could not be uploaded — document saved without attachment.', 'warning')
    doc = Document(
        doc_type=request.form['doc_type'], belongs_to='Employee', owner_name=emp.name,
        employee_id=emp.id, customer_id=emp.customer_id, expiry_date=expiry_dt,
        notes=request.form.get('notes'), file_name=file_name, file_url=file_url,
        cloudinary_public_id=public_id, uploaded_by=session['user_id'], added_by=session['user_name'],
    )
    db.session.add(doc)
    db.session.commit()
    flash('Document added')
    return redirect(url_for('employee_detail', employee_id=emp.id))

# ─────────────────────────── Owners / UBO ───────────────────────────
def _owner_from_form(o):
    o.name = (request.form.get('name') or o.name or '').strip()
    o.role = request.form.get('role') or None
    o.nationality = request.form.get('nationality') or None
    o.passport_no = request.form.get('passport_no') or None
    o.eid_no = request.form.get('eid_no') or None
    try:
        o.share_pct = float(request.form.get('share_pct')) if request.form.get('share_pct') else None
    except ValueError:
        o.share_pct = None
    o.mobile = (request.form.get('mobile') or '').strip() or None
    pe = request.form.get('passport_expiry'); ee = request.form.get('eid_expiry')
    o.passport_expiry = datetime.strptime(pe, '%Y-%m-%d').date() if pe else None
    o.eid_expiry = datetime.strptime(ee, '%Y-%m-%d').date() if ee else None
    dob = request.form.get('date_of_birth')
    o.date_of_birth = datetime.strptime(dob, '%Y-%m-%d').date() if dob else None
    return o

@app.route('/customers/<int:customer_id>/owners/add', methods=['POST'])
@login_required
def add_owner(customer_id):
    Customer.query.get_or_404(customer_id)
    if not (request.form.get('name') or '').strip():
        flash('Owner name is required')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    o = _owner_from_form(Owner(customer_id=customer_id))
    db.session.add(o)
    db.session.commit()
    flash(f'Added {o.name}')
    return redirect(url_for('customer_detail', customer_id=customer_id))

@app.route('/owners/<int:owner_id>/edit', methods=['POST'])
@login_required
def edit_owner(owner_id):
    o = Owner.query.get_or_404(owner_id)
    _owner_from_form(o)
    db.session.commit()
    flash('Owner updated')
    return redirect(url_for('customer_detail', customer_id=o.customer_id))

@app.route('/owners/<int:owner_id>/delete', methods=['POST'])
@login_required
def delete_owner(owner_id):
    o = Owner.query.get_or_404(owner_id)
    cid = o.customer_id
    db.session.delete(o)
    db.session.commit()
    flash('Owner removed')
    return redirect(url_for('customer_detail', customer_id=cid))

# ─────────────────────────── Email alert engine ───────────────────────────
def send_email(to_list, subject, html_body, attachments=None):
    """attachments: optional list of (filename, bytes, mime_type) tuples."""
    host = os.environ.get('SMTP_HOST'); user = os.environ.get('SMTP_USER'); pwd = os.environ.get('SMTP_PASS')
    port = int(os.environ.get('SMTP_PORT', '465')); sender = os.environ.get('SMTP_FROM', user or '')
    recipients = [r for r in to_list if r]
    if not recipients:
        return False, 'No recipient email on file'

    # Preferred transport: Resend HTTP API (port 443) — works on Railway where
    # outbound SMTP ports are blocked. Falls back to SMTP if no API key set.
    resend_key = os.environ.get('RESEND_API_KEY')
    if resend_key:
        import requests as _rq
        from_addr = os.environ.get('RESEND_FROM') or sender or 'info@tahfeel.ae'
        if '<' not in from_addr:
            from_addr = f'Tahfeel Business Solutions <{from_addr}>'
        payload = {'from': from_addr, 'to': recipients,
                   'subject': subject, 'html': html_body}
        if attachments:
            import base64
            payload['attachments'] = [
                {'filename': fname, 'content': base64.b64encode(data).decode()}
                for fname, data, _mime in attachments
            ]
        try:
            resp = _rq.post('https://api.resend.com/emails',
                            headers={'Authorization': f'Bearer {resend_key}',
                                     'Content-Type': 'application/json'},
                            json=payload,
                            timeout=30)
            if resp.status_code in (200, 201):
                return True, 'sent'
            return False, f'Resend {resp.status_code}: {resp.text[:200]}'
        except Exception as e:
            return False, f'Resend error: {e}'

    if not (host and user and pwd):
        return False, 'Email not configured (set RESEND_API_KEY, or SMTP_HOST/USER/PASS)'
    msg = MIMEMultipart('mixed' if attachments else 'alternative')
    msg['Subject'] = subject; msg['From'] = sender; msg['To'] = ', '.join(recipients)
    msg.attach(MIMEText(html_body, 'html'))
    if attachments:
        from email.mime.application import MIMEApplication
        for fname, data, mime in attachments:
            subtype = (mime or 'application/octet-stream').split('/')[-1]
            part = MIMEApplication(data, _subtype=subtype)
            part.add_header('Content-Disposition', 'attachment', filename=fname)
            msg.attach(part)

    import socket as _socket
    _orig_gai = _socket.getaddrinfo
    def _gai_ipv4(*a, **k):
        # Force IPv4 regardless of how getaddrinfo is called (positional or
        # keyword family) — avoids "multiple values for 'family'" / signature errors.
        a = list(a)
        if len(a) >= 3:
            a[2] = _socket.AF_INET
        else:
            k['family'] = _socket.AF_INET
        return _orig_gai(*a, **k)

    def _send_via(use_ssl, p):
        # Some hosts (e.g. Railway) have broken IPv6 egress which surfaces as
        # "[Errno 101] Network is unreachable" on SMTP connect → force IPv4.
        _socket.getaddrinfo = _gai_ipv4
        try:
            if use_ssl:
                with smtplib.SMTP_SSL(host, p, timeout=12) as s:
                    s.login(user, pwd); s.sendmail(sender, recipients, msg.as_string())
            else:
                with smtplib.SMTP(host, p, timeout=12) as s:
                    s.ehlo(); s.starttls(); s.ehlo()
                    s.login(user, pwd); s.sendmail(sender, recipients, msg.as_string())
        finally:
            _socket.getaddrinfo = _orig_gai

    # Try the configured transport first, then fall back to the other port/mode.
    attempts = [(True, 465), (False, 587)] if port == 465 else [(False, port), (True, 465)]
    last_err = None
    for use_ssl, p in attempts:
        try:
            _send_via(use_ssl, p)
            return True, 'sent'
        except Exception as e:
            last_err = e
            print(f'[email] {"SSL" if use_ssl else "STARTTLS"}:{p} failed: {e}')
    return False, str(last_err)

def _expiring_items(customer_id, within_days=30):
    today = now_dubai().date()
    items = []
    for d in Document.query.filter_by(customer_id=customer_id).all():
        if d.expiry_date:
            days = (d.expiry_date.date() - today).days
            if days <= within_days:
                items.append((d, days))
    items.sort(key=lambda x: x[1])
    return items

def _doc_table_html(customer, items, intro):
    rows = ''
    for d, days in items:
        status = f'Expired {-days}d ago' if days < 0 else f'{days} days left'
        color = '#B91C1C' if days <= 30 else ('#B45309' if days <= 60 else '#059669')
        rows += (f'<tr><td style="padding:6px 10px;border:1px solid #e5e7eb;">{d.doc_type}</td>'
                 f'<td style="padding:6px 10px;border:1px solid #e5e7eb;">{d.owner_name or customer.name}</td>'
                 f'<td style="padding:6px 10px;border:1px solid #e5e7eb;">{d.expiry_date.strftime("%d %b %Y")}</td>'
                 f'<td style="padding:6px 10px;border:1px solid #e5e7eb;color:{color};font-weight:bold;">{status}</td></tr>')
    return (f'<div style="font-family:Arial,sans-serif;color:#1a2333;">'
            f'<h2 style="color:#1A3B8B;">Tahfeel Business Health Check</h2>'
            f'<p><strong>{customer.name}</strong> — {intro}</p>'
            f'<table style="border-collapse:collapse;font-size:14px;"><tr style="background:#1A3B8B;color:#fff;">'
            f'<th style="padding:6px 10px;">Document</th><th style="padding:6px 10px;">Owner</th>'
            f'<th style="padding:6px 10px;">Expiry</th><th style="padding:6px 10px;">Status</th></tr>{rows}</table>'
            f'<p style="color:#888;font-size:12px;margin-top:16px;">Sent automatically by Tahfeel CRM.</p></div>')

@app.route('/cron/expiry-alerts')
def cron_expiry_alerts():
    from flask import jsonify
    if not os.environ.get('CRON_KEY') or request.args.get('key', '') != os.environ.get('CRON_KEY'):
        return 'Forbidden', 403
    if not automation_on('auto_expiry_email'):
        return jsonify({'skipped': 'weekly document-expiry email is turned OFF in the admin panel'})
    admin_email = os.environ.get('ALERT_ADMIN_EMAIL')
    alerted, results = 0, []
    for c in Customer.query.filter_by(alerts_enabled=True).all():
        items = _expiring_items(c.id, 30)
        if not items:
            continue
        ok, msg = send_email([c.alert_email or c.email, admin_email],
                             f'Document Expiry Alert — {c.name}',
                             _doc_table_html(c, items, f'has {len(items)} document(s) expiring within 30 days or already expired:'))
        if ok:
            alerted += 1
        results.append(f'{c.name}: {len(items)} doc(s) -> {msg}')
    _mark_run('expiry_email', f'{alerted} customer(s) emailed')
    return jsonify({'companies_alerted': alerted, 'details': results})

@app.route('/customers/<int:customer_id>/email-health', methods=['POST'])
@login_required
def email_health(customer_id):
    c = Customer.query.get_or_404(customer_id)
    try:
        items = _expiring_items(c.id, 36500)  # all documents that have an expiry date
        if not items:
            flash('No documents with expiry dates to report.')
            return redirect(url_for('customer_health', customer_id=customer_id))
        ok, msg = send_email([c.alert_email or c.email, os.environ.get('ALERT_ADMIN_EMAIL')],
                             f'Compliance Report — {c.name}',
                             _doc_table_html(c, items, f'full document status report ({len(items)} documents):'))
        flash('📧 Compliance report emailed.' if ok else f'Email failed: {msg}')
    except Exception as e:
        import traceback; traceback.print_exc()
        flash(f'Email failed: {e}')
    return redirect(url_for('customer_health', customer_id=customer_id))

@app.route('/health-check')
@login_required
def health_check():
    now = now_dubai()
    today = now.date()
    docs = Document.query.filter(Document.expiry_date != None).all()

    def days_left(d):
        return (d.expiry_date.date() - today).days

    # Overall summary counts
    n_expired = len([d for d in docs if days_left(d) < 0])
    n_red = len([d for d in docs if 0 <= days_left(d) <= 30])
    n_amber = len([d for d in docs if 30 < days_left(d) <= 60])
    n_green = len([d for d in docs if days_left(d) > 60])

    # Staff names behind each bucket — so an expiring staff document (e.g. a
    # visa) is visible right in the summary, not just buried in the full list.
    def staff_names(bucket_docs):
        return sorted({d.owner_name for d in bucket_docs if d.belongs_to == 'Staff' and d.owner_name})
    staff_expired = staff_names([d for d in docs if days_left(d) < 0])
    staff_red = staff_names([d for d in docs if 0 <= days_left(d) <= 30])
    staff_amber = staff_names([d for d in docs if 30 < days_left(d) <= 60])

    # Group documents by owner (customer/company, or staff name when no customer)
    customers_by_id = {c.id: c for c in Customer.query.all()}
    groups = {}
    for d in docs:
        key = d.customer_id if d.customer_id else ('staff:' + (d.owner_name or 'Unknown'))
        groups.setdefault(key, []).append(d)

    rows = []
    for key, dl in groups.items():
        dl_sorted = sorted(dl, key=lambda d: d.expiry_date)
        worst = min(days_left(d) for d in dl)
        # Compliance score: valid (>90d)=1.0, expiring (0–90d)=0.5, expired=0 → avg %
        n_valid = len([d for d in dl if days_left(d) > 90])
        n_soon = len([d for d in dl if 0 <= days_left(d) <= 90])
        score = round(100 * (n_valid + 0.5 * n_soon) / len(dl)) if dl else None
        band = ('Excellent' if score >= 90 else 'Good' if score >= 70
                else 'Average' if score >= 50 else 'Poor') if score is not None else 'No data'
        if isinstance(key, int):
            c = customers_by_id.get(key)
            owner_name = c.name if c else 'Unknown'
            owner_type = (c.customer_type if c and c.customer_type else 'Individual')
            cid = key
            ac_code = c.ac_code if c else None
        else:
            owner_name = key.split('staff:', 1)[1]
            owner_type = 'Staff'
            cid = None
            ac_code = None
        rows.append({'owner_name': owner_name, 'owner_type': owner_type, 'customer_id': cid,
                     'ac_code': ac_code, 'docs': dl_sorted, 'count': len(dl), 'worst': worst,
                     'score': score, 'band': band})
    rows.sort(key=lambda r: r['worst'])  # most urgent first

    status_filter = request.args.get('status', '')
    type_filter = request.args.get('type', '')
    search = request.args.get('q', '').strip()
    if search:
        sl = search.lower()
        rows = [r for r in rows if sl in (r['owner_name'] or '').lower()
                or sl in (r['ac_code'] or '').lower()]
    if type_filter:
        rows = [r for r in rows if r['owner_type'] == type_filter]
    if status_filter == 'expired':
        rows = [r for r in rows if r['worst'] < 0]
    elif status_filter == 'red':
        rows = [r for r in rows if 0 <= r['worst'] <= 30]
    elif status_filter == 'amber':
        rows = [r for r in rows if 30 < r['worst'] <= 60]
    elif status_filter == 'green':
        rows = [r for r in rows if r['worst'] > 60]

    return render_template('health_check.html', rows=rows, now=now, today=today,
                           n_expired=n_expired, n_red=n_red, n_amber=n_amber, n_green=n_green,
                           staff_expired=staff_expired, staff_red=staff_red, staff_amber=staff_amber,
                           total=len(docs), status_filter=status_filter, type_filter=type_filter,
                           search=search)

# ─────────────── Compliance report (printable, A4 landscape) ───────────────
DOC_CATEGORIES = [
    ('Trade License', '📜'), ('Emirates ID', '🆔'), ('Passport', '📘'),
    ('Visa', '✈️'), ('Labor Card', '💳'), ('Establishment Card', '🏛️'),
    ('Medical Certificate', '🏥'), ('Insurance', '🛡️'), ('Contract', '📝'),
    ('NOC', '📄'), ('Ejari', '🏠'), ('Tenancy Contract', '🏘️'), ('Other', '📦'),
]

def _doc_categories(docs, dl):
    """Group documents into the seeded categories with a status summary each."""
    known = [c[0] for c in DOC_CATEGORIES[:-1]]
    cats = []
    for name, icon in DOC_CATEGORIES:
        if name == 'Other':
            ds = [d for d in docs if (d.doc_type or 'Other') not in known]
        else:
            ds = [d for d in docs if (d.doc_type or '') == name]
        if not ds:
            continue
        n_exp = len([d for d in ds if dl(d) < 0])
        n_soon = len([d for d in ds if 0 <= dl(d) <= 60])
        if n_exp:
            status, color = f'{n_exp} expired', '#EF4444'
        elif n_soon:
            status, color = f'{n_soon} expiring', '#F59E0B'
        else:
            status, color = 'Active', '#16A34A'
        cats.append({'name': name, 'icon': icon, 'count': len(ds), 'status': status, 'color': color})
    return cats

def _donut_segments(band_counts, order, colors, circumference):
    """Precompute SVG stroke-dasharray/rotation for a segmented donut."""
    total = sum(band_counts.get(b, 0) for b in order) or 1
    segs, acc = [], 0.0
    for b in order:
        frac = band_counts.get(b, 0) / total
        if frac <= 0:
            continue
        segs.append({'color': colors[b], 'dash': round(frac * circumference, 1),
                     'rot': round(acc * 360 - 90, 1)})
        acc += frac
    return segs

def _customer_report_data(customer_id):
    """Everything the customer-facing report needs — THEIR documents only."""
    customer = Customer.query.get_or_404(customer_id)
    today = now_dubai().date()
    docs = [d for d in Document.query.filter_by(customer_id=customer_id).all() if d.expiry_date]
    docs.sort(key=lambda d: d.expiry_date)

    def dl(d):
        return (d.expiry_date.date() - today).days

    n_valid = len([d for d in docs if dl(d) > 90])
    n_expiring = len([d for d in docs if 0 <= dl(d) <= 90])
    n_expired = len([d for d in docs if dl(d) < 0])
    total = len(docs)
    score = round(100 * (n_valid + 0.5 * n_expiring) / total) if total else None
    # Customer-facing wording only — never 'Poor'/'Average' (owner decision)
    if score is None:
        mood, mood_color = 'No documents on file', '#94A3B8'
    elif score >= 90:
        mood, mood_color = 'Excellent standing', '#16A34A'
    elif score >= 70:
        mood, mood_color = 'Good standing', '#16A34A'
    elif n_expired:
        mood, mood_color = 'Action needed', '#B91C1C'
    else:
        mood, mood_color = 'Attention recommended', '#B45309'
    uplift = round(100 * (n_valid + n_expired + 0.5 * n_expiring) / total) if total and n_expired else None

    action_items = [(d, dl(d)) for d in docs if dl(d) <= 60]
    timeline = []
    for d, days in [(d, dl(d)) for d in docs][:5]:
        if days < 0:
            label, color, width = 'Overdue', '#EF4444', 16
        elif days <= 60:
            label, color, width = d.expiry_date.strftime('%b %Y'), '#F59E0B', 16 + min(days, 60)
        else:
            label, color, width = d.expiry_date.strftime('%b %Y'), '#16A34A', min(84, 30 + days // 6)
        timeline.append({'label': label, 'doc': d.doc_type, 'color': color, 'width': width})

    # ── Dashboard (page 1) data ──
    import calendar as _cal
    n_exp30 = len([d for d in docs if 0 <= dl(d) <= 30])
    n_exp90 = len([d for d in docs if 0 <= dl(d) <= 90])
    # Estimated non-compliance risk — simple derived index from document status mix
    risk_pct = min(95, n_expired * 15 + n_exp30 * 8 + max(0, n_exp90 - n_exp30) * 3)
    risk_label = 'VERY LOW' if risk_pct < 10 else 'LOW' if risk_pct < 25 else 'MEDIUM' if risk_pct < 50 else 'HIGH'
    risk_color = '#16A34A' if risk_pct < 25 else '#F59E0B' if risk_pct < 50 else '#EF4444'

    # One health card per document type (nearest expiry within the type)
    doc_cards = []
    seen_types = []
    for d in docs:
        t = d.doc_type or 'Other'
        if t in seen_types:
            continue
        seen_types.append(t)
        days = dl(d)
        pct = 0 if days < 0 else min(100, round(100 * days / 365))
        color = '#EF4444' if days < 0 else ('#F59E0B' if days <= 90 else '#16A34A')
        status = 'Expired' if days < 0 else ('Expiring soon' if days <= 90 else 'Healthy')
        doc_cards.append({'type': t, 'days': days, 'pct': pct, 'color': color, 'status': status,
                          'expiry': d.expiry_date.strftime('%d %b %Y')})
    doc_cards = doc_cards[:9]

    # Rule-based insights
    insights = []
    if score is not None:
        insights.append(('#16A34A' if score >= 70 else '#F59E0B',
                         f'Compliance score {score}% — {mood.lower()}'))
    insights.append(('#16A34A', 'No expired documents') if not n_expired else
                    ('#EF4444', f'{n_expired} expired document{"s" if n_expired != 1 else ""} need immediate renewal'))
    if n_exp90:
        insights.append(('#F59E0B', f'{n_exp90} document{"s" if n_exp90 != 1 else ""} require attention within 90 days'))
    if n_exp30 or n_expired:
        insights.append(('#3B82F6', 'Recommended renewal planning in the next 30 days'))
    insights.append((risk_color, f'Estimated compliance risk: {risk_label.title()}'))

    # Timeline buckets
    buckets = [
        ('Next 30 days', n_exp30, '#EF4444'),
        ('31–60 days', len([d for d in docs if 30 < dl(d) <= 60]), '#F59E0B'),
        ('61–90 days', len([d for d in docs if 60 < dl(d) <= 90]), '#EAB308'),
        ('90+ days', len([d for d in docs if dl(d) > 90]), '#16A34A'),
    ]

    # Attention required (top 5, soonest first)
    attention = [(d, dl(d)) for d in docs if dl(d) <= 90][:5]

    # Employee compliance by doc type
    employees = Employee.query.filter_by(customer_id=customer_id).all()
    emp_docs = [d for d in docs if d.employee_id]
    emp_rows = []
    for t in sorted({(d.doc_type or 'Other') for d in emp_docs}):
        ds = [d for d in emp_docs if (d.doc_type or 'Other') == t]
        ok = len([d for d in ds if dl(d) > 90])
        due = len(ds) - ok
        emp_rows.append({'type': t, 'ok': ok, 'due': due})
    owners_count = Owner.query.filter_by(customer_id=customer_id).count()

    # 3-month compliance calendar
    cal_months = []
    for i in range(3):
        mm = (today.month - 1 + i) % 12 + 1
        yy = today.year + ((today.month - 1 + i) // 12)
        items = [d for d in docs if d.expiry_date.year == yy and d.expiry_date.month == mm]
        cal_months.append({'label': f'{_cal.month_name[mm]} {yy}',
                           'items': [f'{d.expiry_date.strftime("%d %b")} — {d.doc_type}' for d in items[:3]],
                           'more': max(0, len(items) - 3)})

    # Page 2 — group documents by holder. A document belongs to a PERSON only when
    # it is linked to an employee record; everything else (incl. company docs that
    # happen to carry a contact person's name in owner_name) stays under the company.
    company_docs = [(d, dl(d)) for d in docs if not d.employee_id]
    holder_map = {}
    for d in docs:
        if d.employee_id:
            nm = (d.employee.name if d.employee else None) or d.owner_name or 'Employee'
            holder_map.setdefault(nm, []).append((d, dl(d)))
    doc_groups = []
    if company_docs:
        doc_groups.append({'holder': customer.name, 'kind': 'company',
                           'docs': sorted(company_docs, key=lambda x: x[1])})
    for name in sorted(holder_map):
        doc_groups.append({'holder': name, 'kind': 'person',
                           'docs': sorted(holder_map[name], key=lambda x: x[1])})

    report_no = f'TBS/CR/{today.year}/{customer_id:04d}'

    # ── AI-style compliance advisor (rule-based, deterministic) ──
    advisor = []
    if score is not None:
        advisor.append(f'Your company is in {mood.lower()} with a compliance score of {score}%.')
    if n_expired:
        advisor.append(f'{n_expired} expired document{"s" if n_expired != 1 else ""} '
                       f'{"significantly increase" if n_expired > 1 else "increases"} your regulatory exposure and should be renewed without delay.')
    elif n_exp30:
        advisor.append(f'{n_exp30} document{"s" if n_exp30 != 1 else ""} will require renewal within the next 30 days.')
    if uplift and uplift > (score or 0):
        advisor.append(f'Completing the identified renewals could improve your score to {uplift}% '
                       f'and reduce your estimated compliance risk from {risk_pct}%.')
    if not (customer.vat_status or customer.corp_tax_status):
        pass
    elif (customer.vat_due_date and 0 <= (customer.vat_due_date - today).days <= 60) or \
         (customer.corp_tax_due_date and 0 <= (customer.corp_tax_due_date - today).days <= 60):
        advisor.append('A tax filing is due soon — please ensure returns are submitted on time.')
    else:
        advisor.append('No immediate tax issues were detected.')
    # Priority = the category holding the most urgent document
    urgent_docs = sorted([d for d in docs if dl(d) <= 90], key=dl)
    if urgent_docs:
        advisor.append(f'Priority should be given to {(urgent_docs[0].doc_type or "document").lower()} compliance.')
    else:
        advisor.append('Maintain your current renewal discipline to keep this strong standing.')

    # ── Upcoming renewals forecast — next 3 months, grouped by type ──
    fc = {}
    for d in docs:
        if dl(d) <= 90:
            fc[d.doc_type or 'Other'] = fc.get(d.doc_type or 'Other', 0) + 1
    forecast = [{'type': t, 'count': n} for t, n in sorted(fc.items(), key=lambda x: -x[1])]
    forecast_total = sum(fc.values())

    return {
        'customer': customer, 'today': today, 'docs': [(d, dl(d)) for d in docs],
        'total': total, 'n_valid': n_valid, 'n_expiring': n_expiring, 'n_expired': n_expired,
        'score': score, 'mood': mood, 'mood_color': mood_color, 'uplift': uplift,
        'categories': _doc_categories(docs, dl), 'action_items': action_items,
        'timeline': timeline, 'month_label': now_dubai().strftime('%B %Y'),
        'n_exp30': n_exp30, 'n_exp90': n_exp90, 'risk_pct': risk_pct,
        'risk_label': risk_label, 'risk_color': risk_color, 'doc_cards': doc_cards,
        'insights': insights, 'buckets': buckets, 'attention': attention,
        'employees_count': len(employees), 'emp_rows': emp_rows, 'owners_count': owners_count,
        'cal_months': cal_months, 'doc_groups': doc_groups, 'report_no': report_no,
        'advisor': advisor, 'forecast': forecast, 'forecast_total': forecast_total,
        'account_manager': customer.rep.name if customer.rep else None,
        'generated': now_dubai(),
    }

def build_report_email_html(data):
    """Branded HTML email body for the monthly customer compliance report.
    Table-based + inline styles for mail-client compatibility."""
    c = data['customer']
    score = data['score']
    mood_color = data['mood_color']
    action_rows = ''
    for d, days in data['action_items']:
        holder = f' — {d.owner_name}' if d.owner_name and d.owner_name != c.name else ''
        if days < 0:
            badge = f'<span style="background:#B91C1C;color:#ffffff;font-size:11px;font-weight:bold;padding:2px 10px;border-radius:10px;">Expired {d.expiry_date.strftime("%d %b %Y")}</span>'
        else:
            badge = f'<span style="background:#FFF7ED;color:#B45309;font-size:11px;font-weight:bold;padding:2px 10px;border-radius:10px;">Expires {d.expiry_date.strftime("%d %b %Y")} · {days}d</span>'
        action_rows += (f'<tr><td style="padding:6px 14px;font-size:13px;color:#334155;'
                        f'border-top:1px solid #F1F5F9;">{d.doc_type}{holder}</td>'
                        f'<td style="padding:6px 14px;text-align:right;border-top:1px solid #F1F5F9;">{badge}</td></tr>')
    action_block = ''
    if action_rows:
        action_block = (
            '<table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #FECACA;border-radius:10px;margin-top:14px;">'
            '<tr><td colspan="2" style="background:#FEF2F2;color:#B91C1C;font-size:13px;font-weight:bold;'
            'padding:8px 14px;border-radius:10px 10px 0 0;">Action required</td></tr>'
            f'{action_rows}</table>')
    else:
        action_block = ('<div style="background:#ECFDF5;border-radius:10px;padding:12px 16px;margin-top:14px;'
                        'font-size:13px;color:#15803D;font-weight:bold;">All of your documents are in good standing.</div>')
    score_html = ''
    if score is not None:
        score_html = (
            f'<table width="100%" cellpadding="0" cellspacing="0" style="background:#F8FAFC;border:1px solid #E8ECF2;border-radius:10px;"><tr>'
            f'<td style="padding:16px 20px;width:110px;text-align:center;">'
            f'<div style="font-size:34px;font-weight:bold;color:{mood_color};">{score}%</div>'
            f'<div style="font-size:11px;color:#64748B;">documents valid</div></td>'
            f'<td style="padding:16px 10px;">'
            f'<div style="font-size:14px;font-weight:bold;color:#0B1B35;">Your compliance score</div>'
            f'<div style="font-size:12px;color:{mood_color};font-weight:bold;margin:2px 0 8px;">{data["mood"]}</div>'
            f'<div style="font-size:12px;color:#64748B;">Valid <b style="color:#16A34A;">{data["n_valid"]}</b>'
            f' &nbsp; Expiring <b style="color:#F59E0B;">{data["n_expiring"]}</b>'
            f' &nbsp; Expired <b style="color:#EF4444;">{data["n_expired"]}</b></div></td></tr></table>')
    n_action = len(data['action_items'])
    intro = (f'Here is your compliance health summary for {data["month_label"]}. '
             + (f'{n_action} of your documents require attention to keep your business fully compliant.'
                if n_action else 'All of your documents are in good standing — no action is needed right now.'))
    return f'''<!DOCTYPE html><html><body style="margin:0;padding:0;background:#F1F5F9;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center" style="padding:20px 10px;">
<table width="560" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:10px;overflow:hidden;">
  <tr><td style="background:#0B1B35;padding:18px;text-align:center;">
    <img src="https://tahfeelcrm.online/static/img/logo_white.png" alt="Tahfeel" width="150" style="max-width:150px;height:auto;display:block;margin:0 auto 8px;">
    <div style="color:#ffffff;font-size:16px;font-weight:bold;">Tahfeel Business Setup Services LLC</div>
    <div style="color:#85B7EB;font-size:12px;margin-top:2px;">Monthly compliance health report</div>
  </td></tr>
  <tr><td style="padding:20px 24px 8px;">
    <p style="font-size:14px;color:#334155;margin:0 0 14px;line-height:1.6;">Dear <b>{c.name}</b>,<br>{intro}</p>
    {score_html}
    {action_block}
    <table width="100%" cellpadding="0" cellspacing="0"><tr><td align="center" style="padding:18px 0 8px;">
      <div style="font-size:13px;color:#0B1B35;font-weight:bold;">Need help renewing? Tahfeel handles it end to end.</div>
      <div style="font-size:12px;color:#64748B;margin-top:4px;">Reply to this email or WhatsApp us at +971 4 585 5033 · info@tahfeel.ae</div>
    </td></tr></table>
    <div style="background:#F8FAFC;border:1px solid #E8ECF2;border-radius:8px;padding:10px 14px;margin:8px 0 16px;font-size:12px;color:#475569;">
      &#128206; Your full report is attached as a PDF — summary plus document detail.
    </div>
  </td></tr>
  <tr><td style="background:#F8FAFC;border-top:1px solid #E8ECF2;padding:12px 24px;font-size:10.5px;color:#94A3B8;text-align:center;line-height:1.6;">
    Tahfeel Business Solutions · Dubai, UAE<br>
    You receive this monthly report as a Tahfeel client. This report covers only your own documents.
  </td></tr>
</table></td></tr></table></body></html>'''

def _send_customer_report(customer_id):
    """Generate the customer's PDF + branded email and send it. Returns (ok, msg)."""
    data = _customer_report_data(customer_id)
    c = data['customer']
    to_addr = c.alert_email or c.email
    if not to_addr:
        return False, 'no email on file'
    if not data['total']:
        return False, 'no documents with expiry dates'
    html = render_template('customer_report_pdf.html', **data)
    from weasyprint import HTML as _WHTML
    pdf = _WHTML(string=html, base_url=request.url_root if request else None).write_pdf()
    import re as _re
    safe = _re.sub(r'[^A-Za-z0-9]+', '_', c.name or 'client').strip('_')[:40]
    fname = f"Compliance_Report_{safe}_{now_dubai().strftime('%b%Y')}.pdf"
    n_action = len(data['action_items'])
    subject = f'Your compliance health report — {data["month_label"]}'
    if n_action:
        subject += f' · {n_action} document{"s" if n_action != 1 else ""} need attention'
    body = build_report_email_html(data)
    return send_email([to_addr], subject, body, attachments=[(fname, pdf, 'application/pdf')])

@app.route('/customers/<int:customer_id>/send-report', methods=['POST'])
@login_required
def send_customer_report(customer_id):
    try:
        ok, msg = _send_customer_report(customer_id)
    except Exception as e:
        ok, msg = False, str(e)
    if ok:
        flash('📨 Compliance report emailed successfully.')
    else:
        flash(f'Could not send report: {msg}', 'error')
    return redirect(url_for('customer_health', customer_id=customer_id))

@app.route('/cron/monthly-reports')
def cron_monthly_reports():
    """Monthly customer compliance reports — hit by external cron on the 1st.
    GET /cron/monthly-reports?key=CRON_KEY"""
    from flask import jsonify
    if not os.environ.get('CRON_KEY') or request.args.get('key', '') != os.environ.get('CRON_KEY'):
        return 'Forbidden', 403
    if not automation_on('auto_monthly_report'):
        return jsonify({'skipped': 'monthly compliance report is turned OFF in the admin panel'})
    sent, skipped, details = 0, 0, []
    for c in Customer.query.filter_by(alerts_enabled=True).all():
        try:
            ok, msg = _send_customer_report(c.id)
        except Exception as e:
            ok, msg = False, str(e)
        if ok:
            sent += 1
        else:
            skipped += 1
        details.append(f'{c.name}: {msg}')
    _mark_run('monthly_report', f'{sent} report(s) sent')
    return jsonify({'sent': sent, 'skipped': skipped, 'details': details})

@app.route('/cron/birthday-wishes')
def cron_birthday_wishes():
    """Daily: WhatsApp the approved birthday template to any CUSTOMER whose birthday
    is today, and to any company OWNER/authorized person (sent to their own mobile;
    skipped if none saved). Once per person per year (dedupe).
    GET /cron/birthday-wishes?key=CRON_KEY"""
    from flask import jsonify
    if not os.environ.get('CRON_KEY') or request.args.get('key', '') != os.environ.get('CRON_KEY'):
        return 'Forbidden', 403
    if not automation_on('auto_birthday'):
        return jsonify({'skipped': 'birthday automation is turned OFF in the admin panel'})
    from whatsapp_webhook import send_template, log_message, normalize_phone
    tpl = wa_template_active('tahfeel_birthday')
    if not tpl:
        return jsonify({'error': 'tahfeel_birthday template is not active in WhatsApp Templates'})
    today = now_dubai().date()
    sent, skipped, details = 0, 0, []
    for c in Customer.query.filter(Customer.date_of_birth.isnot(None)).all():
        dob = c.date_of_birth
        if not (dob.month == today.month and dob.day == today.day):
            continue
        key = f'birthday:{c.id}:{today.year}'
        if AutoMessageLog.query.filter_by(dedupe_key=key).first():
            continue  # already wished this year
        to = _cust_wa_number(c)
        if not to:
            skipped += 1; details.append(f'{c.name}: no WhatsApp number'); continue
        first = ((c.contact_person or c.name or 'there').split() or ['there'])[0]
        wam = send_template(to, tpl.meta_name, params=[first], lang=tpl.lang or 'en')
        body = (tpl.body_preview or f'Happy birthday, {first}!').replace('{{1}}', first)
        log_message(to, 'out', body, msg_type='template', wam_id=wam,
                    handled_by='auto-birthday', status='sent' if wam else 'failed',
                    customer_id=c.id)
        if wam:
            db.session.add(AutoMessageLog(kind='birthday', dedupe_key=key, detail=f'{c.name} ({to})'))
            db.session.commit()
            sent += 1
        else:
            skipped += 1
        details.append(f'{c.name}: {"sent" if wam else "FAILED"}')
    # Company owners / authorized persons — wish goes to the OWNER's own mobile only
    for o in Owner.query.filter(Owner.date_of_birth.isnot(None)).all():
        dob = o.date_of_birth
        if not (dob.month == today.month and dob.day == today.day):
            continue
        key = f'birthday:owner:{o.id}:{today.year}'
        if AutoMessageLog.query.filter_by(dedupe_key=key).first():
            continue  # already wished this year
        to = normalize_phone(o.mobile or '')
        if not to:
            skipped += 1; details.append(f'{o.name} (owner): no mobile saved'); continue
        first = ((o.name or 'there').split() or ['there'])[0]
        wam = send_template(to, tpl.meta_name, params=[first], lang=tpl.lang or 'en')
        body = (tpl.body_preview or f'Happy birthday, {first}!').replace('{{1}}', first)
        log_message(to, 'out', body, msg_type='template', wam_id=wam,
                    handled_by='auto-birthday', status='sent' if wam else 'failed',
                    customer_id=o.customer_id)
        if wam:
            db.session.add(AutoMessageLog(kind='birthday', dedupe_key=key, detail=f'{o.name} — owner of {o.company.name if o.company else "?"} ({to})'))
            db.session.commit()
            sent += 1
        else:
            skipped += 1
        details.append(f'{o.name} (owner): {"sent" if wam else "FAILED"}')
    _mark_run('birthday', f'{sent} wish(es) sent')
    return jsonify({'sent': sent, 'skipped': skipped, 'details': details})

@app.route('/cron/expiry-wa')
def cron_expiry_wa():
    """Daily: WhatsApp an expiry reminder to the customer when one of their documents
    is exactly 7 or 3 days from expiry. Once per document per milestone (dedupe), and
    only for customers who have alerts enabled. GET /cron/expiry-wa?key=CRON_KEY"""
    from flask import jsonify
    if not os.environ.get('CRON_KEY') or request.args.get('key', '') != os.environ.get('CRON_KEY'):
        return 'Forbidden', 403
    if not automation_on('auto_expiry_wa'):
        return jsonify({'skipped': 'expiry-WhatsApp automation is turned OFF in the admin panel'})
    from whatsapp_webhook import send_template, log_message
    tpl = wa_template_active('compliance_alert_v1')
    if not tpl:
        return jsonify({'error': 'compliance_alert_v1 template is not active in WhatsApp Templates'})
    today = now_dubai().date()
    MILESTONES = (7, 3)
    sent, skipped, details = 0, 0, []
    docs = Document.query.filter(Document.expiry_date.isnot(None),
                                 Document.customer_id.isnot(None)).all()
    for d in docs:
        days = (d.expiry_date.date() - today).days
        if days not in MILESTONES:
            continue
        cust = d.customer
        if not cust or not cust.alerts_enabled:
            continue  # respect the per-customer alert opt-in
        key = f'expirywa:{d.id}:{d.expiry_date.date().isoformat()}:{days}'
        if AutoMessageLog.query.filter_by(dedupe_key=key).first():
            continue
        to = _cust_wa_number(cust)
        if not to:
            skipped += 1; details.append(f'{cust.name}/{d.doc_type}: no WhatsApp number'); continue
        first = ((cust.contact_person or cust.name or 'there').split() or ['there'])[0]
        item = d.doc_type or 'document'
        due = d.expiry_date.strftime('%d %b %Y')
        params = [first, item, due]
        wam = send_template(to, tpl.meta_name, params=params, lang=tpl.lang or 'en')
        body = tpl.body_preview or ''
        for n, v in enumerate(params, start=1):
            body = body.replace('{{%d}}' % n, v)
        log_message(to, 'out', body, msg_type='template', wam_id=wam,
                    handled_by='auto-expiry', status='sent' if wam else 'failed',
                    customer_id=cust.id)
        if wam:
            db.session.add(AutoMessageLog(kind='expiry_wa', dedupe_key=key,
                                          detail=f'{cust.name} · {item} · {days}d left'))
            db.session.commit()
            sent += 1
        else:
            skipped += 1
        details.append(f'{cust.name}/{item} ({days}d): {"sent" if wam else "FAILED"}')
    _mark_run('expiry_wa', f'{sent} reminder(s) sent')
    return jsonify({'sent': sent, 'skipped': skipped, 'details': details})

@app.route('/customers/<int:customer_id>/report.pdf')
@login_required
def customer_report_pdf(customer_id):
    """Download/preview the customer's personal compliance report as a real PDF."""
    data = _customer_report_data(customer_id)
    html = render_template('customer_report_pdf.html', **data)
    try:
        from weasyprint import HTML as _WHTML
        pdf = _WHTML(string=html, base_url=request.url_root).write_pdf()
    except Exception as e:
        print(f'[report] WeasyPrint failed: {e}')
        flash(f'PDF engine error: {e}', 'error')
        return redirect(url_for('customer_health', customer_id=customer_id))
    import re as _re
    safe = _re.sub(r'[^A-Za-z0-9]+', '_', data['customer'].name or 'client').strip('_')[:40]
    fname = f"Compliance_Report_{safe}_{now_dubai().strftime('%b%Y')}.pdf"
    from flask import Response
    return Response(pdf, mimetype='application/pdf',
                    headers={'Content-Disposition': f'inline; filename="{fname}"'})

@app.route('/health-check/report')
@login_required
def compliance_report():
    """Internal A4-landscape print report: executive summary + full detail."""
    import calendar
    now = now_dubai()
    today = now.date()
    docs = Document.query.filter(Document.expiry_date != None).all()

    def dl(d):
        return (d.expiry_date.date() - today).days

    total = len(docs)
    n_expired = len([d for d in docs if dl(d) < 0])
    n_red = len([d for d in docs if 0 <= dl(d) <= 30])
    n_amber = len([d for d in docs if 30 < dl(d) <= 60])
    n_green = len([d for d in docs if dl(d) > 60])
    n_valid90 = len([d for d in docs if dl(d) > 90])
    n_soon90 = len([d for d in docs if 0 <= dl(d) <= 90])
    overall = round(100 * (n_valid90 + 0.5 * n_soon90) / total) if total else None
    overall_band = ('Excellent' if overall >= 90 else 'Good' if overall >= 70
                    else 'Average' if overall >= 50 else 'Needs attention') if overall is not None else 'No data'
    ring_color = ('#16A34A' if overall >= 70 else '#F59E0B' if overall >= 50 else '#EF4444') if overall is not None else '#94A3B8'

    def bucket_split(ds):
        return {'total': len(ds), 'valid': len([d for d in ds if dl(d) > 60]),
                'expiring': len([d for d in ds if 0 <= dl(d) <= 60]),
                'expired': len([d for d in ds if dl(d) < 0])}
    comp_stats = bucket_split([d for d in docs if (d.belongs_to or '') == 'Company'])
    ppl_stats = bucket_split([d for d in docs if (d.belongs_to or '') != 'Company'])

    renew30 = len([d for d in docs if 0 <= dl(d) <= 30])
    renew60 = len([d for d in docs if 0 <= dl(d) <= 60])
    need_action = n_expired + n_red

    customers_by_id = {c.id: c for c in Customer.query.all()}
    def holder_label(d):
        c = customers_by_id.get(d.customer_id) if d.customer_id else None
        if c:
            if d.owner_name and d.owner_name.strip().lower() != (c.name or '').strip().lower():
                return f'{d.owner_name} — {c.name}'
            return c.name
        return (d.owner_name or 'Unknown') + (' (staff)' if d.belongs_to == 'Staff' else '')

    urgent = sorted([d for d in docs if dl(d) <= 60], key=dl)
    alerts = [(d, dl(d), holder_label(d)) for d in urgent[:4]]
    renewals = [(d, dl(d), holder_label(d)) for d in urgent[:8]]
    staff_alerts = [(d, dl(d)) for d in sorted(
        [d for d in docs if d.belongs_to == 'Staff' and dl(d) <= 60], key=dl)][:6]

    categories = _doc_categories(docs, dl)

    timeline = []
    month_counts = []
    for i in range(3):
        mm = (today.month - 1 + i) % 12 + 1
        yy = today.year + ((today.month - 1 + i) // 12)
        cnt = len([d for d in docs if d.expiry_date.year == yy and d.expiry_date.month == mm and dl(d) >= 0])
        month_counts.append((f'{calendar.month_abbr[mm]} {yy}', cnt))
    mx = max([c[1] for c in month_counts] + [1])
    for (lbl, cnt), color in zip(month_counts, ['#EF4444', '#F59E0B', '#16A34A']):
        timeline.append({'label': lbl, 'count': cnt, 'width': round(100 * cnt / mx), 'color': color})

    groups = {}
    for d in docs:
        key = d.customer_id if d.customer_id else ('staff:' + (d.owner_name or 'Unknown'))
        groups.setdefault(key, []).append(d)
    rows = []
    for key, dlist in groups.items():
        dsorted = sorted(dlist, key=lambda d: d.expiry_date)
        worst = min(dl(d) for d in dlist)
        n_v = len([d for d in dlist if dl(d) > 90])
        n_s = len([d for d in dlist if 0 <= dl(d) <= 90])
        score = round(100 * (n_v + 0.5 * n_s) / len(dlist))
        band = ('Excellent' if score >= 90 else 'Good' if score >= 70
                else 'Average' if score >= 50 else 'Poor')
        if isinstance(key, int):
            c = customers_by_id.get(key)
            oname = c.name if c else 'Unknown'
            otype = (c.customer_type if c and c.customer_type else 'Individual')
            ac = c.ac_code if c else None
        else:
            oname = key.split('staff:', 1)[1]
            otype = 'Staff'
            ac = None
        rows.append({'name': oname, 'type': otype, 'ac': ac, 'count': len(dlist),
                     'worst': worst, 'score': score, 'band': band,
                     'n_expired': len([d for d in dlist if dl(d) < 0]),
                     'docs': [(d, dl(d)) for d in dsorted]})
    rows.sort(key=lambda r: r['worst'])

    band_counts = {'Excellent': 0, 'Good': 0, 'Average': 0, 'Poor': 0}
    for r in rows:
        band_counts[r['band']] += 1
    donut_segs = _donut_segments(
        band_counts, ('Excellent', 'Good', 'Average', 'Poor'),
        {'Excellent': '#16A34A', 'Good': '#4ADE80', 'Average': '#F59E0B', 'Poor': '#EF4444'},
        238.8)

    return render_template('compliance_report.html', now=now, today=today,
                           total=total, n_expired=n_expired, n_red=n_red, n_amber=n_amber,
                           n_green=n_green, overall=overall, overall_band=overall_band,
                           ring_color=ring_color, comp_stats=comp_stats, ppl_stats=ppl_stats,
                           renew30=renew30, renew60=renew60, need_action=need_action,
                           alerts=alerts, renewals=renewals, staff_alerts=staff_alerts,
                           categories=categories, timeline=timeline, rows=rows,
                           band_counts=band_counts, donut_segs=donut_segs,
                           n_clients=len(rows))

@app.route('/documents')
@login_required
def documents():
    """Document Alerts — a filtered list of documents expiring within 90 days
    (including already-expired). Adding documents happens at the customer/
    company level now, so this page is alert-only, not a general browser."""
    now = now_dubai()
    # Remember the last-used filters so opening a document and returning keeps the view.
    DFK = ['search', 'expiry', 'belongs_to', 'doc_type']
    if request.args.get('reset') == '1':
        session.pop('docs_filters', None)
        return redirect(url_for('documents'))
    if any(request.args.get(k) for k in DFK):
        session['docs_filters'] = {k: request.args.get(k, '') for k in DFK}
        args = request.args
    elif 'docs_filters' in session:
        args = session['docs_filters']
    else:
        args = request.args
    search = (args.get('search', '') or '').strip().lower()
    expiry_filter = args.get('expiry', '')
    belongs_filter = args.get('belongs_to', '')
    doc_type_filter = args.get('doc_type', '')

    try:
        all_docs = Document.query.filter(Document.expiry_date.isnot(None)).order_by(Document.expiry_date).all()
    except Exception:
        # Run missing column migrations inline
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text('ALTER TABLE document ADD COLUMN IF NOT EXISTS file_name VARCHAR(255)'))
                conn.commit()
        except Exception:
            pass
        flash('System updated. Please refresh.')
        return redirect(url_for('dashboard'))

    def days_left(d):
        return (d.expiry_date - now).days

    # Alert scope: only documents expiring within 90 days (expired included)
    alert_docs = [d for d in all_docs if days_left(d) <= 90]

    expired_count = len([d for d in alert_docs if days_left(d) < 0])
    count_30 = len([d for d in alert_docs if 0 <= days_left(d) <= 30])
    count_60 = len([d for d in alert_docs if 30 < days_left(d) <= 60])
    count_90 = len([d for d in alert_docs if 60 < days_left(d) <= 90])

    doc_list = alert_docs
    if search:
        doc_list = [d for d in doc_list if
                    search in (d.owner_name or '').lower() or
                    search in (d.doc_type or '').lower() or
                    (d.customer and search in d.customer.name.lower()) or
                    (d.customer and d.customer.company and search in d.customer.company.lower())]
    if expiry_filter == 'expired':
        doc_list = [d for d in doc_list if days_left(d) < 0]
    elif expiry_filter == '30':
        doc_list = [d for d in doc_list if 0 <= days_left(d) <= 30]
    elif expiry_filter == '60':
        doc_list = [d for d in doc_list if 30 < days_left(d) <= 60]
    elif expiry_filter == '90':
        doc_list = [d for d in doc_list if 60 < days_left(d) <= 90]
    if belongs_filter:
        doc_list = [d for d in doc_list if d.belongs_to == belongs_filter]
    if doc_type_filter:
        doc_list = [d for d in doc_list if d.doc_type == doc_type_filter]

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = 50
    total = len(doc_list)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated = doc_list[(page-1)*per_page: page*per_page]

    doc_types = DocType.query.order_by(DocType.name).all()
    return render_template('documents.html',
                           documents=paginated,
                           expired_count=expired_count, count_30=count_30,
                           count_60=count_60, count_90=count_90,
                           search=search, expiry_filter=expiry_filter,
                           belongs_filter=belongs_filter, doc_type_filter=doc_type_filter,
                           doc_types=doc_types,
                           total=total, page=page, total_pages=total_pages,
                           expiry_tpl=wa_template_active('compliance_alert_v1'),
                           now=now)

@app.route('/documents/<int:doc_id>/whatsapp-remind', methods=['POST'])
@login_required
def document_whatsapp_remind(doc_id):
    """Send the approved expiry-reminder template (compliance_alert_v1) from the
    business number for one document — API path, logged to the CRM thread."""
    from whatsapp_webhook import send_template, log_message, normalize_phone
    doc = Document.query.get_or_404(doc_id)
    back = request.referrer or url_for('documents')
    tpl = wa_template_active('compliance_alert_v1')
    if not tpl:
        flash('Expiry-reminder template isn\'t active yet — activate compliance_alert_v1 in WhatsApp Templates.', 'error')
        return redirect(back)
    cust = doc.customer
    to = normalize_phone((cust.whatsapp or cust.mobile or cust.phone or cust.phone2) if cust else '')
    if not to:
        flash('No WhatsApp number on record for this customer.', 'error')
        return redirect(back)
    first = ((cust.contact_person or cust.name or 'there').split() or ['there'])[0]
    item = doc.doc_type or 'document'
    due = doc.expiry_date.strftime('%d %b %Y') if doc.expiry_date else ''
    params = [first, item, due]
    wam = send_template(to, tpl.meta_name, params=params, lang=tpl.lang or 'en')
    body = tpl.body_preview
    for n, v in enumerate(params, start=1):
        body = body.replace('{{%d}}' % n, v)
    log_message(to, 'out', body, msg_type='template', wam_id=wam,
                handled_by=session.get('user_name', 'staff'),
                status='sent' if wam else 'failed',
                customer_id=cust.id if cust else None)
    flash('Renewal reminder sent on WhatsApp.' if wam else 'WhatsApp send failed — check the number and template.',
          'success' if wam else 'error')
    return redirect(back)


@app.route('/documents/export')
@login_required
def export_documents():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Documents'
    headers = ['Customer','Company','Doc Type','Belongs To','Owner Name','Expiry Date','Days Until Expiry','Notes','Added By','Created']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    docs = Document.query.order_by(Document.expiry_date).all()
    now = now_dubai()
    for d in docs:
        days = (d.expiry_date - now).days if d.expiry_date else ''
        ws.append([
            d.customer.name if d.customer else '',
            d.customer.company if d.customer and d.customer.company else '',
            d.doc_type or '',
            d.belongs_to or '',
            d.owner_name or '',
            d.expiry_date.strftime('%d/%m/%Y') if d.expiry_date else '',
            days,
            d.notes or '',
            d.added_by or '',
            d.created_at.strftime('%d/%m/%Y') if d.created_at else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 14)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_documents.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/documents/add', methods=['GET', 'POST'])
@login_required
def add_document():
    customers = Customer.query.order_by(Customer.name).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        expiry = request.form.get('expiry_date')
        expiry_dt = datetime.strptime(expiry, '%Y-%m-%d') if expiry else None
        customer_id = request.form.get('customer_id') or None
        if request.form.get('new_customer_name'):
            new_cust = Customer(
                name=request.form['new_customer_name'],
                company=request.form.get('new_customer_company'),
                phone=request.form.get('new_customer_phone'),
                email=request.form.get('new_customer_email'),
                source=request.form.get('new_customer_source'),
            )
            db.session.add(new_cust)
            db.session.flush()
            customer_id = new_cust.id
        # Handle file upload (dummy — store filename only for now)
        file_name = None
        file_url = None
        public_id = None
        if 'document_file' in request.files:
            f = request.files['document_file']
            if f and f.filename:
                file_name = f.filename
                file_url, public_id = upload_to_cloudinary(f)
                if not file_url:
                    flash('⚠️ File could not be uploaded — document saved without attachment. Please check Cloudinary settings.', 'warning')
        doc = Document(
            doc_type=request.form['doc_type'],
            belongs_to=request.form['belongs_to'],
            owner_name=request.form['owner_name'],
            customer_id=int(customer_id) if customer_id else None,
            expiry_date=expiry_dt,
            notes=request.form.get('notes'),
            file_name=file_name,
            file_url=file_url,
            cloudinary_public_id=public_id,
            uploaded_by=session['user_id'],
            added_by=session['user_name'],
        )
        db.session.add(doc)
        db.session.commit()
        # Option A: redirect back to add form with customer pre-selected + success message
        customer_id_param = f'?customer_id={customer_id}&added=1' if customer_id else '?added=1'
        flash('Document saved successfully!')
        if request.form.get('add_another'):
            return redirect(url_for('add_document') + customer_id_param)
        elif customer_id:
            return redirect(url_for('customer_detail', customer_id=int(customer_id)))
        return redirect(url_for('documents'))
    return render_template('add_document.html', customers=customers,
                           doc_types=doc_types, sources=sources)


@app.route('/documents/<int:doc_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    doc_types = DocType.query.order_by(DocType.name).all()
    customers = Customer.query.order_by(Customer.name).all()
    if request.method == 'POST':
        doc.doc_type = request.form.get('doc_type', doc.doc_type)
        doc.belongs_to = request.form.get('belongs_to', doc.belongs_to)
        doc.owner_name = request.form.get('owner_name', doc.owner_name)
        try:
            cid = request.form.get('customer_id')
            doc.customer_id = int(cid) if cid else None
        except: pass
        try:
            ed = request.form.get('expiry_date')
            doc.expiry_date = datetime.strptime(ed, '%Y-%m-%d') if ed else None
        except: pass
        doc.notes = request.form.get('notes', doc.notes)
        # Handle new file upload
        if 'document_file' in request.files:
            f = request.files['document_file']
            if f and f.filename:
                doc.file_name = f.filename
                url, public_id = upload_to_cloudinary(f)
                if url:
                    doc.file_url = url
                    doc.cloudinary_public_id = public_id
        db.session.commit()
        flash('Document updated')
        if request.form.get('add_another'):
            cid = doc.customer_id
            return redirect(url_for('add_document') + (f'?customer_id={cid}' if cid else ''))
        nxt = request.form.get('next') or request.args.get('next')
        return _safe_redirect(nxt, 'documents')
    return render_template('edit_document.html', doc=doc, doc_types=doc_types, customers=customers,
                           next_url=request.args.get('next', ''))

@app.route('/documents/<int:doc_id>/delete', methods=['POST'])
@login_required
def delete_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    db.session.delete(doc)
    db.session.commit()
    flash('Document removed')
    next_url = request.args.get('next')
    return _safe_redirect(next_url, 'documents')

# ── Admin — Document Types ────────────────────────────────────────────────────

@app.route('/admin/doctype/add', methods=['POST'])
@login_required
@admin_required
def admin_add_doctype():
    name = request.form.get('name', '').strip()
    if name:
        if not DocType.query.filter_by(name=name).first():
            db.session.add(DocType(name=name))
            db.session.commit()
            flash(f'Document type "{name}" added')
        else:
            flash('Document type already exists')
    return redirect(url_for('admin_panel'))

@app.route('/admin/doctype/<int:doctype_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_doctype(doctype_id):
    dt = DocType.query.get_or_404(doctype_id)
    db.session.delete(dt)
    db.session.commit()
    flash(f'Document type "{dt.name}" removed')
    return redirect(url_for('admin_panel'))

# ─────────────────────────────────────────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()
        # Create monthly_target table if not exists
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS monthly_target (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER REFERENCES "user"(id),
                        month INTEGER NOT NULL,
                        year INTEGER NOT NULL,
                        lead_target INTEGER DEFAULT 0,
                        conversion_target INTEGER DEFAULT 0,
                        amount_target FLOAT DEFAULT 0
                    )
                """))
                conn.execute(db.text('ALTER TABLE monthly_target ADD COLUMN IF NOT EXISTS amount_target FLOAT DEFAULT 0'))
                conn.execute(db.text('ALTER TABLE customer ADD COLUMN IF NOT EXISTS date_of_birth DATE'))
                conn.execute(db.text('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS phone VARCHAR(20)'))
                conn.execute(db.text('ALTER TABLE customer ADD COLUMN IF NOT EXISTS phone2 VARCHAR(20)'))
                conn.execute(db.text('ALTER TABLE monthly_target ADD COLUMN IF NOT EXISTS lead_target INTEGER DEFAULT 0'))
                conn.execute(db.text('ALTER TABLE monthly_target ADD COLUMN IF NOT EXISTS conversion_target INTEGER DEFAULT 0'))
                conn.commit()
        except Exception as e:
            print(f'monthly_target table: {e}')
        # Create desk_note table if not exists
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS desk_note (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER REFERENCES "user"(id),
                        text TEXT NOT NULL,
                        reminder_date DATE,
                        mention_user_id INTEGER REFERENCES "user"(id),
                        is_done BOOLEAN DEFAULT FALSE,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                """))
                conn.commit()
        except Exception as e:
            print(f'desk_note table: {e}')
        
        # Add revenue column to job table
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS revenue FLOAT DEFAULT 0'))
                conn.commit()
                print('✓ Revenue column migration completed')
        except Exception as e:
            print(f'Revenue column migration error: {e}')
        
        # Add partner commission columns to job table
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_commission_expected BOOLEAN DEFAULT FALSE'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_name VARCHAR(100)'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_amount FLOAT'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_due_date DATE'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_status VARCHAR(20) DEFAULT \'Pending\''))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_received_date DATE'))
                conn.commit()
                print('✓ Partner commission columns migration completed')
        except Exception as e:
            print(f'Partner commission migration error: {e}')
        
        # Create partner table
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS partner (
                        id SERIAL PRIMARY KEY,
                        name VARCHAR(100) NOT NULL UNIQUE,
                        active BOOLEAN DEFAULT TRUE,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                """))
                conn.commit()
                print('✓ Partner table created')
        except Exception as e:
            print(f'Partner table creation error: {e}')
        
        migrations = [
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS phone2 VARCHAR(20)',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS amount_invoiced FLOAT DEFAULT 0',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS amount_received FLOAT DEFAULT 0',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS finance_approved_by INTEGER',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS finance_approved_at TIMESTAMP',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS num_persons INTEGER DEFAULT 1',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS finance_notes TEXT',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS service_type VARCHAR(100)',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS due_date TIMESTAMP',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS priority VARCHAR(20) DEFAULT \'Medium\'',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS amount FLOAT DEFAULT 0',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS final_remarks TEXT',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS future_work_notes TEXT',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS completed_at TIMESTAMP',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS service_note VARCHAR(200)',
            'ALTER TABLE job_update ADD COLUMN IF NOT EXISTS status_note VARCHAR(100)',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS campaign VARCHAR(100)',
            """CREATE TABLE IF NOT EXISTS campaign (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL UNIQUE
            )""",

            'ALTER TABLE job ADD COLUMN IF NOT EXISTS status VARCHAR(50) DEFAULT \'Assigned\'',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS file_name VARCHAR(255)',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS file_url TEXT',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS cloudinary_public_id VARCHAR(255)',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS revenue_date DATE',
            'ALTER TABLE activity_log ADD COLUMN IF NOT EXISTS off_day VARCHAR(20)',
            'ALTER TABLE job_type ADD COLUMN IF NOT EXISTS default_days INTEGER DEFAULT 1',
            '''CREATE TABLE IF NOT EXISTS activity_type (
                id SERIAL PRIMARY KEY,
                field_key VARCHAR(50) UNIQUE NOT NULL,
                label VARCHAR(150) NOT NULL,
                weekly_target FLOAT DEFAULT 5,
                sort_order INTEGER DEFAULT 0,
                active BOOLEAN DEFAULT TRUE
            )''',
            'ALTER TABLE activity_type ADD COLUMN IF NOT EXISTS weekly_target FLOAT DEFAULT 5',
            "UPDATE \"user\" SET role = 'sales' WHERE role = 'staff'",
            'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS on_leave BOOLEAN DEFAULT FALSE',
            'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS report_from DATE',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS meta_lead_id VARCHAR(50)',
            # Company entity + document linkage (company table itself created by create_all)
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS company_id INTEGER',
            'ALTER TABLE company ADD COLUMN IF NOT EXISTS alerts_enabled BOOLEAN DEFAULT FALSE',
            'ALTER TABLE company ADD COLUMN IF NOT EXISTS alert_email VARCHAR(120)',
            'ALTER TABLE company ADD COLUMN IF NOT EXISTS alert_whatsapp VARCHAR(30)',
            # Customer = unified entity (Individual or Company): contact person + expiry alerts
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS contact_person VARCHAR(100)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS alerts_enabled BOOLEAN DEFAULT FALSE',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS alert_email VARCHAR(120)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS alert_whatsapp VARCHAR(30)',
            # Company profile fields (UAE) on customer
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS ac_code VARCHAR(50)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS trade_name VARCHAR(150)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS legal_form VARCHAR(60)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS jurisdiction VARCHAR(30)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS licensing_authority VARCHAR(120)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS freezone_name VARCHAR(120)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS emirate VARCHAR(40)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS country_incorp VARCHAR(60)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS business_activity VARCHAR(200)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS ac_status VARCHAR(30)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS po_box VARCHAR(30)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS mobile VARCHAR(30)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS whatsapp VARCHAR(30)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS website VARCHAR(120)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS ac_opening_date DATE',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS uae_pass_number VARCHAR(50)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS uae_pass_name VARCHAR(100)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS vat_status VARCHAR(20)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS vat_due_date DATE',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS corp_tax_status VARCHAR(20)',
            'ALTER TABLE customer ADD COLUMN IF NOT EXISTS corp_tax_due_date DATE',
            # Lead redesign: quality flag + channel + timing + per-update activity
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS genuine VARCHAR(20)',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS junk_reason VARCHAR(100)',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS sub_source VARCHAR(50)',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS first_contacted_at TIMESTAMP',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS attempts INTEGER DEFAULT 0',
            'ALTER TABLE lead_update ADD COLUMN IF NOT EXISTS activity_type VARCHAR(50)',
            # One-time map of old activity-statuses -> clean pipeline stages
            "UPDATE lead SET status='Contacted' WHERE status IN ('Called — No Answer','Customer Not Responding','Call Connected','Sent WhatsApp','Sent Mail','Called — Callback Requested')",
            "UPDATE lead SET status='Qualified' WHERE status='Potential Lead'",
            "UPDATE lead SET status='Proposal' WHERE status IN ('Meeting Scheduled','Quotation Sent')",
            # Task stage rename: 'Final Stage' -> 'Partially Completed'
            "UPDATE job SET status='Partially Completed' WHERE status='Final Stage'",
            # Normalize legacy free-text Meta source -> clean 'Meta' + channel
            "UPDATE lead SET sub_source='Facebook' WHERE sub_source IS NULL AND source LIKE 'Meta%Facebook%'",
            "UPDATE lead SET sub_source='Instagram' WHERE sub_source IS NULL AND source LIKE 'Meta%Instagram%'",
            "UPDATE lead SET source='Meta' WHERE source LIKE 'Meta Ads%' OR source LIKE 'Meta —%'",
            # Document -> employee link (employee/owner tables auto-created by create_all)
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS employee_id INTEGER',
            # Tahfeel Doc: category (Tahfeel / Staff / Management) + person link
            'ALTER TABLE company_document ADD COLUMN IF NOT EXISTS category VARCHAR(20)',
            'ALTER TABLE company_document ADD COLUMN IF NOT EXISTS staff_id INTEGER',
            # WhatsApp inbox: unread tracking
            'ALTER TABLE whats_app_message ADD COLUMN IF NOT EXISTS is_read BOOLEAN DEFAULT FALSE',
            # WhatsApp: per-conversation bot pause (human takeover)
            'ALTER TABLE whats_app_thread ADD COLUMN IF NOT EXISTS bot_paused BOOLEAN DEFAULT FALSE',
            'ALTER TABLE whats_app_thread ADD COLUMN IF NOT EXISTS bot_paused_by VARCHAR(100)',
            'ALTER TABLE whats_app_thread ADD COLUMN IF NOT EXISTS resolved BOOLEAN DEFAULT FALSE',
            'ALTER TABLE whats_app_thread ADD COLUMN IF NOT EXISTS resolved_at TIMESTAMP',
            'ALTER TABLE whats_app_thread ADD COLUMN IF NOT EXISTS resolved_by VARCHAR(100)',
            # WhatsApp: media message support
            'ALTER TABLE whats_app_message ADD COLUMN IF NOT EXISTS media_url VARCHAR(500)',
            'ALTER TABLE whats_app_message ADD COLUMN IF NOT EXISTS mime_type VARCHAR(50)',
            'ALTER TABLE whats_app_message ADD COLUMN IF NOT EXISTS error VARCHAR(300)',
            # Super Admin flag: can edit other admins + self (fixed to admin@tahfeel.ae)
            'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS is_super BOOLEAN DEFAULT FALSE',
            "UPDATE \"user\" SET is_super = TRUE WHERE email = 'admin@tahfeel.ae'",
            # Owner: date of birth (auto birthday wish) + own mobile to send it to
            'ALTER TABLE owner ADD COLUMN IF NOT EXISTS date_of_birth DATE',
            'ALTER TABLE owner ADD COLUMN IF NOT EXISTS mobile VARCHAR(30)',

            # ── Performance indexes (H6) — speed up the filtered queries in reports,
            #    analytics, marketing report, cron jobs, WhatsApp matching, and the
            #    per-record lookups. IF NOT EXISTS = safe to re-run on every boot.
            'CREATE INDEX IF NOT EXISTS idx_lead_assigned_to ON lead (assigned_to)',
            'CREATE INDEX IF NOT EXISTS idx_lead_status ON lead (status)',
            'CREATE INDEX IF NOT EXISTS idx_lead_created_at ON lead (created_at)',
            'CREATE INDEX IF NOT EXISTS idx_lead_source ON lead (source)',
            'CREATE INDEX IF NOT EXISTS idx_customer_assigned_to ON customer (assigned_to)',
            'CREATE INDEX IF NOT EXISTS idx_customer_phone ON customer (phone)',
            'CREATE INDEX IF NOT EXISTS idx_job_status ON job (status)',
            'CREATE INDEX IF NOT EXISTS idx_job_revenue_date ON job (revenue_date)',
            'CREATE INDEX IF NOT EXISTS idx_job_assigned_to ON job (assigned_to)',
            'CREATE INDEX IF NOT EXISTS idx_job_customer_id ON job (customer_id)',
            'CREATE INDEX IF NOT EXISTS idx_job_created_at ON job (created_at)',
            'CREATE INDEX IF NOT EXISTS idx_document_expiry_date ON document (expiry_date)',
            'CREATE INDEX IF NOT EXISTS idx_document_customer_id ON document (customer_id)',
            'CREATE INDEX IF NOT EXISTS idx_document_employee_id ON document (employee_id)',
            'CREATE INDEX IF NOT EXISTS idx_document_company_id ON document (company_id)',
            'CREATE INDEX IF NOT EXISTS idx_lead_update_lead_id ON lead_update (lead_id)',
            'CREATE INDEX IF NOT EXISTS idx_job_update_job_id ON job_update (job_id)',
            'CREATE INDEX IF NOT EXISTS idx_partial_revenue_job_id ON partial_revenue (job_id)',
            'CREATE INDEX IF NOT EXISTS idx_partial_revenue_revenue_date ON partial_revenue (revenue_date)',
            'CREATE INDEX IF NOT EXISTS idx_wam_created_at ON whats_app_message (created_at)',
            'CREATE INDEX IF NOT EXISTS idx_login_attempt_ip_created ON login_attempt (ip, created_at)',
        ]
        for sql in migrations:
            try:
                with db.engine.connect() as conn:
                    conn.execute(db.text(sql))
                    conn.commit()
                    print(f'Migration OK: {sql[:60]}')
            except Exception as e:
                print(f'Migration skip: {sql[:60]} — {e}')
        try:
            admin = User.query.filter_by(email='admin@tahfeel.ae').first()
            if not admin:
                new_admin = User(
                    name='Admin-Tahfeel', email='admin@tahfeel.ae',
                    password=generate_password_hash('tahfeel2026'), role='admin',
                    is_super=True
                )
                db.session.add(new_admin)
                db.session.commit()
                print('Admin user created')
            elif admin.name == 'Admin':
                # Update existing admin name if it's still 'Admin'
                admin.name = 'Admin-Tahfeel'
                db.session.commit()
                print('Admin user name updated to Admin-Tahfeel')
            
            if Service.query.count() == 0:
                for s in ['Trade License', 'Family Visa', 'PRO Services', 'Healthcare License', 'Umrah Package', 'Other']:
                    db.session.add(Service(name=s))
                db.session.commit()
                print('Default services created')
            if Source.query.count() == 0:
                for s in ['Walk-in', 'WhatsApp', 'Referral', 'Social Media', 'Website', 'Other']:
                    db.session.add(Source(name=s))
                db.session.commit()
                print('Default sources created')
            # Always ensure 'Meta' source exists (leads from Meta Lead Ads)
            if not Source.query.filter_by(name='Meta').first():
                db.session.add(Source(name='Meta'))
                db.session.commit()
                print("Ensured 'Meta' source exists")
            # Source for leads converted from a WhatsApp bot conversation
            if not Source.query.filter_by(name='WhatsApp - AI Bot').first():
                db.session.add(Source(name='WhatsApp - AI Bot'))
                db.session.commit()
                print("Ensured 'WhatsApp - AI Bot' source exists")
            # Starter WhatsApp templates — seeded INACTIVE; admin activates each one
            # after creating + approving the same name/wording in Meta Business Manager
            if MessageTemplate.query.count() == 0:
                starters = [
                    ('Task completed — congratulations', 'tahfeel_job_completed', 'Utility',
                     'first_name,job_type',
                     'Dear {{1}}, great news! Your {{2}} has been completed successfully. Thank you for choosing Tahfeel Business Services. We remain at your service for any future requirements.'),
                    ('Payment received — thank you', 'tahfeel_payment_thankyou', 'Utility',
                     'first_name',
                     'Dear {{1}}, we confirm receipt of your payment. Thank you for your trust in Tahfeel Business Services.'),
                    ('Renewal reminder', 'tahfeel_renewal_reminder', 'Utility',
                     'first_name,custom,custom',
                     'Dear {{1}}, a friendly reminder from Tahfeel: your {{2}} is due for renewal on {{3}}. Reply to this message and our team will gladly assist you with the renewal.'),
                    ('Birthday greeting', 'tahfeel_birthday', 'Marketing',
                     'first_name',
                     'Dear {{1}}, the entire team at Tahfeel Business Services wishes you a very happy birthday! May the year ahead bring you success and prosperity.'),
                    ('General update / announcement', 'tahfeel_general_update', 'Marketing',
                     'first_name,custom',
                     'Dear {{1}}, an important update from Tahfeel Business Services: {{2}}. Reply to this message if you would like our assistance.'),
                ]
                for label, mname, cat, vfields, body in starters:
                    db.session.add(MessageTemplate(
                        label=label, meta_name=mname, category=cat,
                        var_fields=vfields, body_preview=body, active=False))
                db.session.commit()
                print('Starter WhatsApp templates seeded (inactive)')
            # Client-journey template batch — seeded INACTIVE, once only. Guarded by a
            # sentinel (document_request_v1): if none of the batch exists yet, add them
            # all; individual deletions afterwards are respected (won't be re-added).
            journey = [
                ('Document Request', 'document_request_v1', 'Utility', 'first_name,job_type,custom',
                 'Dear {{1}},\n\nTo continue processing your {{2}}, we require the following document(s):\n\n{{3}}\n\nPlease reply to this message with the requested document(s). If you need any assistance, kindly contact your account manager or reply to this message.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Documents Received', 'documents_received_v1', 'Utility', 'first_name,job_type',
                 'Dear {{1}},\n\nThank you. We have successfully received the documents required for your {{2}}.\n\nOur team is reviewing them and will keep you informed of the next update.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Additional Documents Required', 'additional_documents_required_v1', 'Utility', 'first_name,custom,job_type',
                 'Dear {{1}},\n\nTo continue processing your {{3}}, we require the following additional document(s):\n\n{{2}}\n\nPlease reply to this message with the requested document(s) at your earliest convenience to avoid any delay.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Application / Status Update', 'application_progress_v1', 'Utility', 'first_name,job_type,custom',
                 'Dear {{1}},\n\nWe would like to update you on your {{2}}.\n\nCurrent Status:\n{{3}}\n\nIf any action is required from your side, we will contact you separately.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Action Required', 'action_required_v1', 'Utility', 'first_name,job_type,custom',
                 'Dear {{1}},\n\nAction is required to continue processing your {{2}}.\n\nPlease complete the following:\n\n{{3}}\n\nKindly reply to this message once completed or contact your account manager if you need any assistance.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Compliance Alert', 'compliance_alert_v1', 'Utility', 'first_name,custom,custom',
                 'Dear {{1}},\n\nThis is a reminder that your {{2}} is due on {{3}}.\n\nPlease contact us if you require assistance to complete the necessary process before the due date.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Compliance Report Ready', '_compliance_report_ready_v1', 'Utility', 'first_name',
                 'Dear {{1}},\n\nYour monthly compliance report is now ready.\n\nPlease review the report sent to your inbox for upcoming renewals, pending actions, and important compliance updates.\n\nIf you have any questions, simply reply to this message.\n\nThank you,\nTahfeel Business Setup Services'),
                ('Monthly Check-In', 'service_followup_v1', 'Marketing', 'first_name',
                 'Dear {{1}},\n\nWe hope everything is going well.\n\nIf you have any questions regarding your existing services with Tahfeel, simply reply to this message and our team will be happy to assist you.\n\nThank you,\nTahfeel Business Setup Services'),
            ]
            journey_names = [m for _, m, _, _, _ in journey]
            if not MessageTemplate.query.filter(MessageTemplate.meta_name.in_(journey_names)).first():
                for label, mname, cat, vfields, body in journey:
                    db.session.add(MessageTemplate(
                        label=label, meta_name=mname, category=cat,
                        var_fields=vfields, body_preview=body, active=False))
                db.session.commit()
                print(f'Seeded {len(journey)} client-journey WhatsApp templates (inactive)')
            # Retire consolidated/dropped templates — only while still INACTIVE, so
            # anything the owner approved+activated or edited into use is never removed.
            for dead in ('application_submitted_v1', 'processing_update_v1', 'service_feedback_v1'):
                t = MessageTemplate.query.filter_by(meta_name=dead, active=False).first()
                if t:
                    db.session.delete(t)
                    print(f'Retired unused template: {dead}')
            db.session.commit()
            # One-time corrections to match how templates ended up in Meta:
            #  - compliance report name gained a leading underscore in Meta
            #  - owner reworded the compliance report body ("sent to your inbox")
            #  - Meta reclassified the monthly check-in as Marketing
            _rpt = MessageTemplate.query.filter_by(meta_name='compliance_report_ready_v1').first()
            if _rpt:
                _rpt.meta_name = '_compliance_report_ready_v1'
            _rpt = MessageTemplate.query.filter_by(meta_name='_compliance_report_ready_v1').first()
            if _rpt and 'sent to your inbox' not in (_rpt.body_preview or ''):
                _rpt.body_preview = ('Dear {{1}},\n\nYour monthly compliance report is now ready.\n\n'
                                     'Please review the report sent to your inbox for upcoming renewals, '
                                     'pending actions, and important compliance updates.\n\n'
                                     'If you have any questions, simply reply to this message.\n\n'
                                     'Thank you,\nTahfeel Business Setup Services')
            _fu = MessageTemplate.query.filter_by(meta_name='service_followup_v1').first()
            if _fu and _fu.category != 'Marketing':
                _fu.category = 'Marketing'
            db.session.commit()
            # Re-engagement template for winning back Lost leads (seeded inactive).
            # 2 variables: {{1}} first name, {{2}} the service/thing they enquired about.
            _reengage_body = ('Hello {{1}},\n\nWe noticed that you previously enquired with '
                'Tahfeel Business Setup Services about {{2}}.\n\nWe wanted to check if you\'re '
                'still interested. If you need any information or would like to move forward, '
                'simply reply to this message and our team will be happy to assist you.\n\n'
                'If you\'d prefer to speak directly with one of our Business Consultants, just '
                'reply "Speak" and we\'ll arrange it for you.\n\nThank you,')
            _re = MessageTemplate.query.filter_by(meta_name='tahfeel_reengage_v1').first()
            if not _re:
                db.session.add(MessageTemplate(
                    label='Re-engage lost lead', meta_name='tahfeel_reengage_v1',
                    category='Marketing', var_fields='first_name,service',
                    body_preview=_reengage_body, active=False))
                db.session.commit()
                print('Seeded re-engage template (inactive)')
            elif _re.var_fields == 'first_name':
                # upgrade the earlier 1-variable version to the 2-variable Meta template
                _re.var_fields = 'first_name,service'
                _re.body_preview = _reengage_body
                db.session.commit()
                print('Upgraded re-engage template to 2 variables')
            # Task status-update template (customer-facing progress update), inactive
            # until an admin approves it in Meta + activates it in WhatsApp Templates.
            _status_body = ('Hello {{1}},\n\nThis is a quick update regarding your task with Tahfeel.\n\n'
                            'Your *{{2}}* task is currently in the *{{3}}* stage and our team is actively '
                            'processing it. At this time, no action is required from your side. This message '
                            'is for your information only.\n\nWe will keep you informed as your application '
                            'progresses. If we require any documents or additional information, we will '
                            'contact you promptly.\n\nIf you have any questions in the meantime, simply reply '
                            'to this message or reach out to your Account Manager.\n\nThank you,\n'
                            'Tahfeel Business Setup.')
            _st = MessageTemplate.query.filter_by(meta_name='status_update').first()
            if not _st:
                db.session.add(MessageTemplate(
                    label='Task status update', meta_name='status_update',
                    category='Utility', var_fields='first_name,service,custom',
                    body_preview=_status_body, active=False))
                db.session.commit()
                print('Seeded status-update template (inactive)')
            elif _st.body_preview != _status_body:
                _st.body_preview = _status_body
                db.session.commit()
                print('Updated status-update template preview')
            if ServiceType.query.count() == 0:
                for jt in ['Trade License', 'Family Visa', 'PRO Services', 'Healthcare License', 'Umrah Package', 'Other']:
                    db.session.add(ServiceType(name=jt))
                db.session.commit()
                print('Default job types created')
            if ActivityType.query.count() == 0:
                for i, (key, label, target) in enumerate(ACTIVITY_DEFAULTS):
                    db.session.add(ActivityType(field_key=key, label=label, weekly_target=target, sort_order=i))
                db.session.commit()
                print('Default activity types seeded')
            if DocType.query.count() == 0:
                for dt in ['Trade License', 'Emirates ID', 'Passport', 'Visa', 'Medical Certificate', 'Insurance', 'Contract', 'NOC', 'Ejari', 'Other']:
                    db.session.add(DocType(name=dt))
                db.session.commit()
                print('Default doc types created')
            # Ensure newer doc types exist (idempotent on existing deploys)
            for dt in ['Establishment Card', 'Labor Card', 'ILOE Insurance', 'Other']:
                if not DocType.query.filter_by(name=dt).first():
                    db.session.add(DocType(name=dt))
            db.session.commit()
            # Backfill: link legacy Staff/Management docs (free-text owner) to person records
            try:
                legacy = CompanyDocument.query.filter(
                    CompanyDocument.category.in_(['Staff', 'Management']),
                    CompanyDocument.staff_id.is_(None)
                ).all()
                if legacy:
                    cache = {}
                    for d in legacy:
                        nm = (d.owner or '').strip()
                        if not nm or nm.lower() == 'tahfeel':
                            continue
                        key = (d.category, nm.lower())
                        person = cache.get(key) or TahfeelStaff.query.filter_by(category=d.category, name=nm).first()
                        if not person:
                            person = TahfeelStaff(name=nm, category=d.category)
                            db.session.add(person)
                            db.session.flush()
                        cache[key] = person
                        d.staff_id = person.id
                    db.session.commit()
                    print(f'Backfilled {len(legacy)} legacy staff/mgmt docs into person records')
            except Exception as e:
                db.session.rollback()
                print(f'Staff backfill skipped: {e}')
        except Exception as e:
            db.session.rollback()
            print(f'Init db error: {e}')

@app.route('/admin/targets', methods=['GET','POST'])
@login_required
@admin_required
def set_targets():
    now = now_dubai()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    if request.method == 'POST':
        month = int(request.form.get('month', now.month))
        year = int(request.form.get('year', now.year))
        # Get all active users (including admin)
        all_users = User.query.filter_by(active=True).all()
        saved = 0
        for u in all_users:
            val = request.form.get(f'amount_{u.id}', '').strip()
            amount_t = float(val) if val else 0.0
            try:
                t = MonthlyTarget.query.filter_by(user_id=u.id, month=month, year=year).first()
                if t:
                    t.amount_target = amount_t
                else:
                    t = MonthlyTarget(user_id=u.id, month=month, year=year,
                                     lead_target=0, conversion_target=0, amount_target=amount_t)
                    db.session.add(t)
                saved += 1
            except Exception as e:
                db.session.rollback()
                flash(f'Error saving target for {u.name}: {e}')
        try:
            db.session.commit()
            flash(f'Targets saved for {month}/{year}.')
        except Exception as e:
            db.session.rollback()
            flash(f'Save failed: {e}')
        return redirect(url_for('set_targets', month=month, year=year))
    users = User.query.filter_by(active=True).order_by(User.name).all()
    targets = {t.user_id: t for t in MonthlyTarget.query.filter_by(month=month, year=year).all()}
    return render_template('targets.html', users=users, targets=targets, month=month, year=year, now=now)

@app.route('/desk', methods=['GET','POST'])
@login_required
def my_desk():
    now = now_dubai()
    user_id = session['user_id']
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            text = request.form.get('text','').strip() or '📌'
            reminder_date = request.form.get('reminder_date','').strip()
            mention_user_ids = request.form.getlist('mention_user_ids')
            rd = datetime.strptime(reminder_date, '%Y-%m-%d').date() if reminder_date else None
            if mention_user_ids:
                # Create one note per mentioned user
                for mid in mention_user_ids:
                    note = DeskNote(user_id=user_id, text=text, reminder_date=rd, mention_user_id=int(mid))
                    db.session.add(note)
            else:
                note = DeskNote(user_id=user_id, text=text, reminder_date=rd, mention_user_id=None)
                db.session.add(note)
            db.session.commit()
        elif action == 'done':
            note_id = request.form.get('note_id')
            note = DeskNote.query.get(note_id)
            if note and (note.user_id == user_id or note.mention_user_id == user_id):
                note.is_done = not note.is_done
                db.session.commit()
        elif action == 'delete':
            note_id = request.form.get('note_id')
            note = DeskNote.query.get(note_id)
            if note and (note.user_id == user_id or note.mention_user_id == user_id):
                db.session.delete(note)
                db.session.commit()
        return redirect(url_for('my_desk'))

    # My notes + mentions
    my_notes = DeskNote.query.filter_by(user_id=user_id).order_by(DeskNote.is_done, DeskNote.reminder_date.asc().nullslast(), DeskNote.created_at.desc()).all()
    mentions = DeskNote.query.filter_by(mention_user_id=user_id, is_done=False).order_by(DeskNote.created_at.desc()).all()
    all_users = User.query.filter_by(active=True).filter(User.id != user_id).order_by(User.name).all()
    # Monthly targets + workload
    target = MonthlyTarget.query.filter_by(user_id=user_id, month=now.month, year=now.year).first()
    my_jobs_all = Job.query.filter_by(assigned_to=user_id).all()
    invoiced_actual = sum((j.amount_invoiced or 0) for j in my_jobs_all if j.status not in ['Pending Finance Approval'])
    closed_actual = sum((j.amount_received or 0) for j in my_jobs_all if j.status == 'Closed')
    amount_target = (target.amount_target or 0) if target else 0
    # Workload this month
    my_leads_month = Lead.query.filter_by(assigned_to=user_id).filter(
        db.extract('month', Lead.created_at) == now.month,
        db.extract('year', Lead.created_at) == now.year
    ).all()
    my_leads_count = len(my_leads_month)
    my_conversions_count = len([l for l in my_leads_month if l.status == 'Converted'])
    my_lost_count = len([l for l in my_leads_month if l.status == 'Lost'])
    my_overdue_leads = len([l for l in my_leads_month if l.due_date and l.due_date < now and l.status not in ['Converted','Lost']])
    my_active_tasks = len([j for j in my_jobs_all if j.status not in ['Done','Closed','Pending Finance Approval']])
    my_overdue_tasks = len([j for j in my_jobs_all if j.due_date and j.due_date < now and j.status not in ['Done','Closed','Pending Finance Approval']])

    # Create table if not exists
    try:
        db.session.execute(db.text('SELECT 1 FROM desk_note LIMIT 1'))
    except:
        db.session.rollback()
        with db.engine.connect() as conn:
            conn.execute(db.text("""
                CREATE TABLE IF NOT EXISTS desk_note (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER REFERENCES "user"(id),
                    text TEXT NOT NULL,
                    reminder_date DATE,
                    mention_user_id INTEGER REFERENCES "user"(id),
                    is_done BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()

    try:
        session['unread_mentions'] = len(mentions)
    except:
        pass
    return render_template('my_desk.html', my_notes=my_notes, mentions=mentions,
                           all_users=all_users, now=now,
                           invoiced_actual=invoiced_actual, closed_actual=closed_actual,
                           amount_target=amount_target,
                           my_leads_count=my_leads_count, my_conversions_count=my_conversions_count,
                           my_lost_count=my_lost_count, my_overdue_leads=my_overdue_leads,
                           my_active_tasks=my_active_tasks, my_overdue_tasks=my_overdue_tasks)

@app.route('/check-birthdays')
@login_required
def check_birthdays():
    today = now_dubai()
    try:
        result = db.session.execute(db.text("SELECT id, name, date_of_birth FROM customer WHERE date_of_birth IS NOT NULL")).fetchall()
        out = f"<b>Today (Dubai): {today.day}/{today.month}/{today.year}</b><br><br>Customers with DOB ({len(result)}):<br>"
        for r in result:
            dob = r[2]
            match = dob and dob.month == today.month and dob.day == today.day
            out += f"- {r[1]}: {dob} {'✅ BIRTHDAY TODAY' if match else ''}<br>"
        if not result:
            out += "No customers have DOB set yet."
        return out
    except Exception as e:
        return f'Error: {e}'
        
@app.route('/invoice')
@login_required
def invoice_generator():
    services = [s.name for s in Service.query.order_by(Service.name).all()]
    return render_template('invoice_generator.html', services=services)
   
@app.route('/admin/backup/export')
@login_required
@admin_required
def export_full_backup():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from flask import make_response
        import io

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="133E87", end_color="133E87", fill_type="solid")

        def style_headers(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Sheet 1: Leads
        try:
            ws1 = wb.active
            ws1.title = "Leads"
            users_map = {u.id: u.name for u in User.query.all()}
            style_headers(ws1, ['ID','Name','Company','Phone','Email','Service','Source','Campaign',
                                 'Status','Assigned To','Created Date','Due Date','Remarks'])
            for l in Lead.query.order_by(Lead.created_at.desc()).all():
                ws1.append([
                    l.id, l.name or '', l.company or '', l.phone or '', l.email or '',
                    l.service or '', l.source or '', l.campaign or '',
                    l.status or '', users_map.get(l.assigned_to, ''),
                    l.created_at.strftime('%d/%m/%Y %H:%M') if l.created_at else '',
                    l.due_date.strftime('%d/%m/%Y') if l.due_date else '',
                    l.remarks or ''
                ])
        except Exception as e:
            print(f"Error backing up Leads: {e}")
            flash(f'Warning: Leads backup incomplete - {str(e)}', 'warning')

        # Sheet 2: Customers
        try:
            ws2 = wb.create_sheet("Customers")
            style_headers(ws2, ['ID','Name','Company','Phone','Email','Source','Nationality',
                                 'Customer Type','Assigned To','Notes','Created Date'])
            for c in Customer.query.order_by(Customer.created_at.desc()).all():
                ws2.append([
                    c.id, c.name or '', c.company or '', c.phone or '', c.email or '',
                    c.source or '', c.nationality or '', c.customer_type or '',
                    users_map.get(c.assigned_to, ''), c.notes or '',
                    c.created_at.strftime('%d/%m/%Y %H:%M') if c.created_at else ''
                ])
        except Exception as e:
            print(f"Error backing up Customers: {e}")
            flash(f'Warning: Customers backup incomplete - {str(e)}', 'warning')

        # Sheet 3: Jobs
        try:
            ws3 = wb.create_sheet("Jobs")
            style_headers(ws3, ['ID','Customer','Company','Job Type','Assigned To','Created By',
                                 'Status','Invoiced (AED)','Received (AED)','Revenue (AED)',
                                 'Revenue Date','Partner Commission','Partner Received','Created Date','Due Date'])
            for j in Job.query.order_by(Job.created_at.desc()).all():
                try:
                    ws3.append([
                        j.id,
                        j.customer.name if j.customer else '',
                        j.customer.company if j.customer else '',
                        j.job_type or '',
                        users_map.get(j.assigned_to, ''),
                        users_map.get(j.created_by, ''),
                        j.status or '',
                        float(j.amount_invoiced or 0),
                        float(j.amount_received or 0),
                        float(j.revenue or 0),
                        j.revenue_date.strftime('%d/%m/%Y') if j.revenue_date else '',
                        'Yes' if j.partner_commission_expected else 'No',
                        'Yes' if (j.partner_status == 'Received') else 'No',
                        j.created_at.strftime('%d/%m/%Y %H:%M') if j.created_at else '',
                        j.due_date.strftime('%d/%m/%Y') if j.due_date else ''
                    ])
                except Exception as row_error:
                    print(f"Error backing up Job ID {j.id}: {row_error}")
                    continue
        except Exception as e:
            print(f"Error backing up Jobs: {e}")
            flash(f'Warning: Jobs backup incomplete - {str(e)}', 'warning')

        # Sheet 4: Documents
        try:
            ws4 = wb.create_sheet("Documents")
            style_headers(ws4, ['ID','Doc Type','Owner Name','Belongs To','Customer',
                                 'Expiry Date','Notes','Added By','Created Date'])
            for d in Document.query.order_by(Document.created_at.desc()).all():
                try:
                    ws4.append([
                        d.id, d.doc_type or '', d.owner_name or '', d.belongs_to or '',
                        d.customer.name if d.customer else '',
                        d.expiry_date.strftime('%d/%m/%Y') if d.expiry_date else '',
                        d.notes or '', d.added_by or '',
                        d.created_at.strftime('%d/%m/%Y %H:%M') if d.created_at else ''
                    ])
                except Exception as row_error:
                    print(f"Error backing up Document ID {d.id}: {row_error}")
                    continue
        except Exception as e:
            print(f"Error backing up Documents: {e}")
            flash(f'Warning: Documents backup incomplete - {str(e)}', 'warning')

        # Sheet 5: Staff
        try:
            ws5 = wb.create_sheet("Staff")
            style_headers(ws5, ['ID','Name','Email','Role','Active'])
            for u in User.query.order_by(User.name).all():
                ws5.append([u.id, u.name, u.email, u.role, 'Yes' if u.active else 'No'])
        except Exception as e:
            print(f"Error backing up Staff: {e}")
            flash(f'Warning: Staff backup incomplete - {str(e)}', 'warning')

        # Mark backup date in session
        session['last_backup_date'] = now_dubai().strftime('%Y-%m-%d')
        session.modified = True

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"tahfeel_backup_{now_dubai().strftime('%Y%m%d_%H%M')}.xlsx"
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    
    except Exception as e:
        print(f"CRITICAL ERROR in backup export: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Backup failed: {str(e)}. Please contact support.', 'error')
        return redirect(url_for('admin_panel'))


@app.route('/analytics')
@login_required
def analytics():
    if session.get('role') not in ['admin', 'finance']:
        flash('Access denied')
        return redirect(url_for('dashboard'))

    now = now_dubai()
    today = now.date()

    # Date range from request args — default this month
    period = request.args.get('period', 'this_month')
    if period == 'last_month':
        if now.month == 1:
            start = now.replace(year=now.year-1, month=12, day=1)
        else:
            start = now.replace(month=now.month-1, day=1)
        import calendar
        last_day = calendar.monthrange(start.year, start.month)[1]
        end = start.replace(day=last_day)
    elif period == 'this_year':
        start = now.replace(month=1, day=1)
        end = now
    else:  # this_month
        start = now.replace(day=1)
        end = now

    start_dt = start.replace(hour=0, minute=0, second=0) if hasattr(start, 'hour') else now.replace(day=1, hour=0, minute=0, second=0)
    end_dt = end.replace(hour=23, minute=59, second=59) if hasattr(end, 'hour') else now.replace(hour=23, minute=59, second=59)

    # All data for period
    all_leads = Lead.query.filter(Lead.created_at >= start_dt, Lead.created_at <= end_dt).all()
    all_jobs = Job.query.options(
        db.joinedload(Job.customer), db.subqueryload(Job.partial_revenues)
    ).filter(Job.created_at >= start_dt, Job.created_at <= end_dt).all()
    all_users = User.query.filter_by(active=True).order_by(User.name).all()
    users_map = {u.id: u.name for u in User.query.all()}

    # ── Lead stats
    total_leads = len(all_leads)
    won_s = {'Won', 'Converted', 'Closed-Won'}
    lost_s = {'Lost', 'Rejected', 'Closed-Lost'}
    converted = [l for l in all_leads if l.status in won_s]
    lost = [l for l in all_leads if l.status in lost_s]
    conversion_rate = round(len(converted) / total_leads * 100, 1) if total_leads > 0 else 0

    # ── Revenue stats (match dashboard Finance card - count ALL jobs)
    total_invoiced = sum(j.amount_invoiced or 0 for j in all_jobs)
    total_received = sum(j.amount_received or 0 for j in all_jobs)
    total_outstanding = total_invoiced - total_received

    # ── Lead pipeline by status
    from collections import defaultdict, Counter
    pipeline = Counter(l.status for l in all_leads)

    # ── Top services
    service_counts = Counter(l.service for l in all_leads if l.service)
    top_services = service_counts.most_common(6)

    # ── Top sources
    source_counts = Counter(l.source for l in all_leads if l.source)
    top_sources = source_counts.most_common(6)

    # ── Campaign performance (show all campaigns, not just the top few)
    campaign_counts = Counter(l.campaign for l in all_leads if l.campaign)
    top_campaigns = campaign_counts.most_common()

    # ── Monthly revenue trend (last 6 months)
    monthly_revenue = []
    for i in range(5, -1, -1):
        if now.month - i <= 0:
            m = now.month - i + 12
            y = now.year - 1
        else:
            m = now.month - i
            y = now.year
        month_jobs = Job.query.filter(
            db.extract('month', Job.created_at) == m,
            db.extract('year', Job.created_at) == y,
            Job.status.notin_(['Pending Finance Approval'])
        ).all()
        inv = sum(j.amount_invoiced or 0 for j in month_jobs)
        rec = sum(j.amount_received or 0 for j in month_jobs)
        import calendar
        monthly_revenue.append({
            'month': calendar.month_abbr[m],
            'invoiced': inv,
            'received': rec,
        })

    # ── Staff performance
    staff_stats = []
    # Get monthly targets
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    targets = {t.user_id: t.amount_target or 0 for t in MonthlyTarget.query.filter_by(month=month, year=year).all()}
    
    for u in all_users:
        if u.role != 'sales':  # Only Sales role
            continue
        u_leads = [l for l in all_leads if l.assigned_to == u.id]
        u_sales = [j for j in all_jobs if j.customer and j.customer.assigned_to == u.id]
        u_inv = sum(j.amount_invoiced or 0 for j in u_sales if j.status not in ['Pending Finance Approval'])
        u_conv = len([l for l in u_leads if l.status in won_s])
        conv_rate = round(u_conv / len(u_leads) * 100) if u_leads else 0
        
        # Count pending leads (not contacted yet - status is "New")
        u_pending = len([l for l in u_leads if l.status == 'New'])
        
        # Calculate revenue from closed jobs + partial revenues
        try:
            u_revenue = sum(j.revenue or 0 for j in u_sales if j.status == 'Closed')
            # Add partial revenues from non-closed jobs
            for j in u_sales:
                if j.status != 'Closed':
                    u_revenue += sum(pr.amount for pr in j.partial_revenues)
        except:
            u_revenue = 0
        
        staff_stats.append({
            'name': u.name,
            'role': u.role,
            'leads': len(u_leads),
            'converted': u_conv,
            'conv_rate': conv_rate,
            'invoiced': u_inv,
            'revenue': u_revenue,
            'target': targets.get(u.id, 0),
            'pending': u_pending,
        })
    staff_stats.sort(key=lambda x: x['revenue'], reverse=True)

    # Max values for bar scaling
    max_service = top_services[0][1] if top_services else 1
    max_source = top_sources[0][1] if top_sources else 1
    max_pipeline = max(pipeline.values()) if pipeline else 1
    max_rev = max((m['invoiced'] for m in monthly_revenue), default=1) or 1

    # ── Operations stats
    total_jobs = len(all_jobs)
    completed_jobs = [j for j in all_jobs if j.status in ['Done','Closed']]
    active_jobs_ops = [j for j in all_jobs if j.status not in ['Done','Closed','Pending Finance Approval']]
    overdue_jobs = [j for j in all_jobs if j.due_date and j.due_date.date() < today and j.status not in ['Done','Closed']]
    job_type_counts = Counter(j.job_type for j in all_jobs if j.job_type)
    top_job_types = job_type_counts.most_common(6)
    max_job_type = top_job_types[0][1] if top_job_types else 1
    job_status_counts = Counter(j.status for j in all_jobs)

    # Avg days to complete
    completion_days = []
    for j in completed_jobs:
        if j.created_at and j.completed_at:
            completion_days.append((j.completed_at - j.created_at).days)
    avg_completion = round(sum(completion_days) / len(completion_days), 1) if completion_days else 0

    # Avg completion days per job type
    from collections import defaultdict
    jtype_days = defaultdict(list)
    for j in Job.query.filter(Job.status.in_(['Done','Closed']), Job.completed_at.isnot(None)).all():
        if j.created_at and j.completed_at and j.job_type:
            jtype_days[j.job_type].append((j.completed_at - j.created_at).days)
    avg_by_job_type = sorted(
        [{'type': jt, 'avg': round(sum(days)/len(days), 1), 'count': len(days)}
         for jt, days in jtype_days.items()],
        key=lambda x: x['avg'], reverse=True
    )

    # ── Documents stats
    from datetime import timedelta as _td
    all_docs = Document.query.all()
    total_docs = len(all_docs)
    expired_docs = [d for d in all_docs if d.expiry_date and d.expiry_date.date() < today]
    expiring_30 = [d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date.date() - today).days <= 30]
    expiring_60 = [d for d in all_docs if d.expiry_date and 31 <= (d.expiry_date.date() - today).days <= 60]
    expiring_90 = [d for d in all_docs if d.expiry_date and 61 <= (d.expiry_date.date() - today).days <= 90]
    doc_type_counts = Counter(d.doc_type for d in all_docs if d.doc_type)
    top_doc_types = doc_type_counts.most_common(6)

    # Tab from request
    tab = request.args.get('tab', 'overview')
    
    # ── Lead breakdown by staff and status (for pivot table)
    # Get all statuses and staff
    all_statuses = sorted(set(l.status for l in all_leads if l.status), key=lambda s: (s != 'New', s))  # "New" first, then alphabetical
    sales_staff = [u for u in all_users if u.role == 'sales']  # Only Sales role, not admin
    
    # Create breakdown: {status: {staff_name: count}}
    lead_breakdown = {}
    for status in all_statuses:
        lead_breakdown[status] = {}
        for staff in sales_staff:
            count = len([l for l in all_leads if l.status == status and l.assigned_to == staff.id])
            lead_breakdown[status][staff.name] = count
    
    # Calculate totals
    status_totals = {status: sum(lead_breakdown[status].values()) for status in all_statuses}
    staff_totals = {staff.name: sum(lead_breakdown[status].get(staff.name, 0) for status in all_statuses) for staff in sales_staff}
    grand_total = sum(status_totals.values())

    # ── Daily leads received (bar chart for the Leads tab; capped to 45 bars) ──
    lead_day_counts = Counter(l.created_at.date() for l in all_leads if l.created_at)
    lead_daily = []
    if lead_day_counts:
        dmax_l, dmin_l = max(lead_day_counts), min(lead_day_counts)
        if (dmax_l - dmin_l).days > 45:
            dmin_l = dmax_l - timedelta(days=45)
        d = dmin_l
        while d <= dmax_l:
            lead_daily.append({'date': d, 'count': lead_day_counts.get(d, 0)})
            d += timedelta(days=1)
    lead_daily_max = max((x['count'] for x in lead_daily), default=0)
    lead_daily_avg = round(sum(x['count'] for x in lead_daily) / len(lead_daily), 1) if lead_daily else 0
    # ── Lead quality breakdown (tallies to total leads, incl. Not reviewed) ──
    q_genuine = sum(1 for l in all_leads if l.genuine == 'Genuine')
    q_junk = sum(1 for l in all_leads if l.genuine == 'Junk')
    q_unreach = sum(1 for l in all_leads if l.genuine == 'Unreachable')
    lead_quality = {'genuine': q_genuine, 'junk': q_junk, 'unreachable': q_unreach,
                    'not_reviewed': total_leads - (q_genuine + q_junk + q_unreach), 'total': total_leads}

    return render_template('analytics.html',
        now=now, period=period, tab=tab,
        total_leads=total_leads, converted=len(converted), lost=len(lost),
        conversion_rate=conversion_rate,
        total_invoiced=total_invoiced, total_received=total_received,
        total_outstanding=total_outstanding,
        pipeline=pipeline, top_services=top_services, top_sources=top_sources,
        top_campaigns=top_campaigns, monthly_revenue=monthly_revenue,
        staff_stats=staff_stats,
        max_service=max_service, max_source=max_source,
        max_pipeline=max_pipeline, max_rev=max_rev,
        users_map=users_map,
        lead_breakdown=lead_breakdown, all_statuses=all_statuses, sales_staff=sales_staff,
        status_totals=status_totals, staff_totals=staff_totals, grand_total=grand_total,
        lead_daily=lead_daily, lead_daily_max=lead_daily_max, lead_daily_avg=lead_daily_avg,
        lead_quality=lead_quality,
        total_jobs=total_jobs, completed_jobs=len(completed_jobs),
        active_jobs_ops=len(active_jobs_ops), overdue_jobs=len(overdue_jobs),
        avg_completion=avg_completion, top_job_types=top_job_types, avg_by_job_type=avg_by_job_type,
        max_job_type=max_job_type, job_status_counts=job_status_counts,
        total_docs=total_docs, expired_docs=expired_docs,
        expiring_30=expiring_30, expiring_60=expiring_60, expiring_90=expiring_90,
        top_doc_types=top_doc_types,
    )

from reports import reports_bp
app.register_blueprint(reports_bp)
from meta_webhook import meta_bp
app.register_blueprint(meta_bp)
from whatsapp_webhook import wa_bp
app.register_blueprint(wa_bp)
# Webhooks are external POSTs authenticated by HMAC signature (not a session
# cookie), so CSRF tokens don't apply — exempt them or Meta/WhatsApp POSTs 400.
csrf.exempt(meta_bp)
csrf.exempt(wa_bp)

# ── WhatsApp inbox (CRM-side UI) ──────────────────────────────────────────────
WA_WINDOW = timedelta(hours=24)  # Meta free-reply window after last inbound

@app.route('/whatsapp')
@login_required
def whatsapp_inbox():
    """List of WhatsApp conversations, newest activity first. Supports search + filter."""
    q = (request.args.get('q') or '').strip()
    flt = request.args.get('filter', 'unread')   # unread(default) | all | unmatched | mine | done
    msgs = WhatsAppMessage.query.order_by(WhatsAppMessage.created_at.desc()).all()
    thread_assignments = {t.wa_id: t for t in WhatsAppThread.query.all()}
    threads = {}
    for m in msgs:
        if m.wa_id not in threads:
            threads[m.wa_id] = {
                'wa_id': m.wa_id,
                'name': m.contact_name,
                'last_body': m.body,
                'last_at': m.created_at,
                'last_dir': m.direction,
                'lead_id': None, 'customer_id': None, 'unread': 0,
            }
        t = threads[m.wa_id]
        if m.direction == 'in' and not m.is_read:
            t['unread'] += 1
        if m.lead_id and not t['lead_id']:
            t['lead_id'] = m.lead_id
        if m.customer_id and not t['customer_id']:
            t['customer_id'] = m.customer_id
        if not t['name'] and m.contact_name:
            t['name'] = m.contact_name
    for t in threads.values():
        t['lead'] = Lead.query.get(t['lead_id']) if t['lead_id'] else None
        t['customer'] = Customer.query.get(t['customer_id']) if t['customer_id'] else None
        wt = thread_assignments.get(t['wa_id'])
        t['assignee'] = wt.assignee if wt else None
        t['resolved'] = bool(wt.resolved) if wt else False
    thread_list = sorted(threads.values(), key=lambda x: x['last_at'], reverse=True)

    uid = session.get('user_id')
    # counts for the filter tabs (before filtering). Active views exclude 'done'.
    counts = {
        'all': len(thread_list),
        'unread': sum(1 for t in thread_list if t['unread'] and not t['resolved']),
        'unmatched': sum(1 for t in thread_list if not t['lead'] and not t['customer'] and not t['resolved']),
        'leads': sum(1 for t in thread_list if t['lead'] and not t['resolved']),
        'mine': sum(1 for t in thread_list if t['assignee'] and t['assignee'].id == uid and not t['resolved']),
        'done': sum(1 for t in thread_list if t['resolved']),
    }
    # apply search
    if q:
        ql = q.lower()
        thread_list = [t for t in thread_list
                       if ql in (t['name'] or '').lower() or ql in (t['wa_id'] or '')]
    # apply filter
    if flt == 'unread':
        thread_list = [t for t in thread_list if t['unread'] and not t['resolved']]
    elif flt == 'unmatched':
        thread_list = [t for t in thread_list if not t['lead'] and not t['customer'] and not t['resolved']]
    elif flt == 'leads':
        thread_list = [t for t in thread_list if t['lead'] and not t['resolved']]
    elif flt == 'mine':
        thread_list = [t for t in thread_list if t['assignee'] and t['assignee'].id == uid and not t['resolved']]
    elif flt == 'done':
        thread_list = [t for t in thread_list if t['resolved']]

    staff = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).order_by(User.name).all()
    return render_template('whatsapp_inbox.html', threads=thread_list, now=now_dubai(),
                           q=q, flt=flt, counts=counts, staff=staff,
                           wa_templates=wa_send_context(), wa_health=_wa_health())

def _wa_health():
    """Plain-English WhatsApp sending health for the inbox banner. Looks at whether
    the API is configured, whether the auto-bot is on, and how many of the most
    recent OUTGOING messages failed — so a billing/token outage is visible in the
    CRM instead of only in server logs."""
    def _envflag(k):
        return os.environ.get(k, '').strip().lower() in ('1', 'true', 'yes', 'on')
    configured = bool(os.environ.get('WA_ACCESS_TOKEN') and os.environ.get('WA_PHONE_NUMBER_ID'))
    bot_on = _envflag('WA_BOT_ENABLED')
    try:
        # Only the last 24h — an old failure shouldn't keep re-triggering the banner
        # for days just because few messages were sent since.
        recent = WhatsAppMessage.query.filter_by(direction='out')\
                 .filter(WhatsAppMessage.created_at >= now_dubai() - timedelta(hours=24))\
                 .order_by(WhatsAppMessage.created_at.desc()).limit(10).all()
    except Exception:
        recent = []
    total = len(recent)
    failed = sum(1 for m in recent if (m.status or '') == 'failed')
    fail_id = max((m.id for m in recent if (m.status or '') == 'failed'), default=0)
    if not configured:
        level, msg = 'down', ('WhatsApp API is not configured (access token missing). '
                              'Outgoing messages cannot be sent. Check the Railway settings.')
    elif total and failed >= max(3, total // 2):
        level, msg = 'down', (f'WhatsApp sending may be DOWN — {failed} of the last {total} '
                              f'outgoing messages failed. Check your Meta WhatsApp billing and '
                              f'that the access token has not expired.')
    elif failed:
        level, msg = 'warn', (f'{failed} of the last {total} outgoing messages failed — '
                              f'usually a lead whose number isn\'t on WhatsApp. Click "See why" for the exact reason.')
    else:
        level, msg = 'ok', 'WhatsApp sending is healthy.'
    return {'level': level, 'msg': msg, 'bot_on': bot_on, 'configured': configured,
            'failed': failed, 'total': total, 'fail_id': fail_id}

@app.route('/whatsapp/<wa_id>')
@login_required
def whatsapp_thread(wa_id):
    msgs = WhatsAppMessage.query.filter_by(wa_id=wa_id)\
            .order_by(WhatsAppMessage.created_at.asc()).all()
    if not msgs:
        flash('Conversation not found')
        return redirect(url_for('whatsapp_inbox'))
    # Mark this conversation's incoming messages as read (clears the unread badge).
    # The 🔄 Refresh button loads with ?refresh=1 — in that case we DON'T mark as
    # read, so refreshing to check for new messages never silently clears unread.
    if not request.args.get('refresh'):
        WhatsAppMessage.query.filter_by(wa_id=wa_id, direction='in', is_read=False)\
                .update({'is_read': True})
        db.session.commit()

    last_in = WhatsAppMessage.query.filter_by(wa_id=wa_id, direction='in')\
            .order_by(WhatsAppMessage.created_at.desc()).first()
    window_open = bool(last_in) and (now_dubai() - last_in.created_at) < WA_WINDOW
    # Human-friendly countdown for the 24h free-reply window
    window_left = ''
    if window_open:
        rem = WA_WINDOW - (now_dubai() - last_in.created_at)
        hrs = int(rem.total_seconds() // 3600)
        mins = int((rem.total_seconds() % 3600) // 60)
        window_left = f'{hrs}h {mins}m'
    name = next((m.contact_name for m in msgs if m.contact_name), None)
    lead_id = next((m.lead_id for m in msgs if m.lead_id), None)
    customer_id = next((m.customer_id for m in msgs if m.customer_id), None)
    # Fallback: if no message was stamped with a lead/customer, match by phone so an
    # existing lead/client is still detected (and the "Convert to Lead" button hidden).
    if not lead_id and not customer_id:
        try:
            from whatsapp_webhook import find_contact
            lead_id, customer_id = find_contact(wa_id)
        except Exception:
            pass
    thread = WhatsAppThread.query.get(wa_id)
    staff = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).order_by(User.name).all()
    quick_replies = QuickReply.query.filter(
        (QuickReply.staff_id == session.get('user_id')) | (QuickReply.is_global == True)
    ).order_by(QuickReply.label).all()
    cust = Customer.query.get(customer_id) if customer_id else None
    return render_template('whatsapp_thread.html',
        wa_id=wa_id, messages=msgs, name=name, window_open=window_open,
        window_left=window_left,
        lead=Lead.query.get(lead_id) if lead_id else None,
        customer=cust,
        assignee=thread.assignee if thread else None, staff=staff,
        bot_paused=thread.bot_paused if thread else False,
        resolved=thread.resolved if thread else False,
        quick_replies=quick_replies,
        wa_templates=wa_send_context(customer=cust),
        now=now_dubai())

@app.route('/whatsapp/<wa_id>/assign', methods=['POST'])
@login_required
def whatsapp_assign(wa_id):
    """Assign (or reassign / unassign) a WhatsApp conversation to a staff member."""
    staff_id = request.form.get('staff_id')
    thread = WhatsAppThread.query.get(wa_id)
    if not thread:
        thread = WhatsAppThread(wa_id=wa_id)
        db.session.add(thread)
    thread.assigned_to = int(staff_id) if staff_id else None
    thread.assigned_at = now_dubai() if staff_id else None
    db.session.commit()
    if staff_id:
        assignee = User.query.get(int(staff_id))
        flash(f'Conversation assigned to {assignee.name if assignee else "staff"}.')
    else:
        flash('Conversation unassigned.')
    return redirect(request.referrer or url_for('whatsapp_thread', wa_id=wa_id))

@app.route('/whatsapp/<wa_id>/reply', methods=['POST'])
@login_required
def whatsapp_reply(wa_id):
    from whatsapp_webhook import send_text, send_media, log_message
    body = (request.form.get('body') or '').strip()
    media_file = request.files.get('media')
    has_media = media_file and media_file.filename
    if not body and not has_media:
        return redirect(url_for('whatsapp_thread', wa_id=wa_id))
    last_in = WhatsAppMessage.query.filter_by(wa_id=wa_id, direction='in')\
            .order_by(WhatsAppMessage.created_at.desc()).first()
    if not last_in or (now_dubai() - last_in.created_at) >= WA_WINDOW:
        flash('24-hour reply window has closed — only an approved template can be sent now.', 'warning')
        return redirect(url_for('whatsapp_thread', wa_id=wa_id))

    if has_media:
        # public=True: Meta's servers must fetch this URL to deliver the media
        url, public_id = upload_to_cloudinary(media_file, folder='tahfeel-whatsapp', public=True)
        if not url:
            flash('Media upload failed.', 'error')
            return redirect(url_for('whatsapp_thread', wa_id=wa_id))
        mime = media_file.mimetype or ''
        media_type = 'image' if mime.startswith('image') else \
                     'audio' if mime.startswith('audio') else \
                     'video' if mime.startswith('video') else 'document'
        wam = send_media(wa_id, media_type, url, caption=body or None)
        log_message(wa_id, 'out', body or f'[{media_type}]', msg_type=media_type, wam_id=wam,
                    handled_by=session.get('user_name', 'staff'), status='sent',
                    media_url=url, mime_type=mime)
    else:
        wam = send_text(wa_id, body)
        log_message(wa_id, 'out', body, wam_id=wam,
                    handled_by=session.get('user_name', 'staff'), status='sent')
    return redirect(url_for('whatsapp_thread', wa_id=wa_id))

@app.route('/api/whatsapp-unread-count')
@login_required
def api_whatsapp_unread_count():
    return jsonify({'unread': wa_unread_count()})

@app.route('/whatsapp/quick-replies', methods=['GET', 'POST'])
@login_required
def quick_replies_list():
    """List + add canned replies (own + global)."""
    if request.method == 'POST':
        label = (request.form.get('label') or '').strip()
        body = (request.form.get('body') or '').strip()
        is_global = request.form.get('is_global') == '1' and session.get('role') == 'admin'
        if label and body:
            db.session.add(QuickReply(
                label=label, body=body, is_global=is_global,
                staff_id=session.get('user_id'), created_at=now_dubai(),
            ))
            db.session.commit()
            flash('Quick reply added.')
        return redirect(url_for('quick_replies_list'))
    replies = QuickReply.query.filter(
        (QuickReply.staff_id == session.get('user_id')) | (QuickReply.is_global == True)
    ).order_by(QuickReply.label).all()
    return render_template('quick_replies.html', replies=replies)

@app.route('/whatsapp/quick-replies/<int:reply_id>/delete', methods=['POST'])
@login_required
def delete_quick_reply(reply_id):
    reply = QuickReply.query.get_or_404(reply_id)
    if reply.staff_id != session.get('user_id') and not (reply.is_global and session.get('role') == 'admin'):
        flash('Not authorized to delete this quick reply.', 'warning')
        return redirect(url_for('quick_replies_list'))
    db.session.delete(reply)
    db.session.commit()
    flash('Quick reply deleted.')
    return redirect(url_for('quick_replies_list'))

@app.route('/whatsapp/templates', methods=['GET', 'POST'])
@login_required
@admin_required
def whatsapp_templates():
    """Manage the approved Meta templates the CRM can send (admin only)."""
    if request.method == 'POST':
        label = (request.form.get('label') or '').strip()
        meta_name = (request.form.get('meta_name') or '').strip()
        body = (request.form.get('body_preview') or '').strip()
        if not (label and meta_name and body):
            flash('Label, Meta template name and body are required.', 'error')
        else:
            db.session.add(MessageTemplate(
                label=label, meta_name=meta_name,
                category=request.form.get('category', 'Utility'),
                lang=(request.form.get('lang') or 'en').strip() or 'en',
                body_preview=body,
                var_fields=(request.form.get('var_fields') or '').strip(),
                active=request.form.get('active') == '1',
                created_at=now_dubai(),
            ))
            db.session.commit()
            flash(f'Template "{label}" added.')
        return redirect(url_for('whatsapp_templates'))
    tpls = MessageTemplate.query.order_by(MessageTemplate.label).all()
    return render_template('whatsapp_templates.html', templates=tpls, var_keys=WA_VAR_LABELS)

@app.route('/whatsapp/templates/<int:tpl_id>/edit', methods=['POST'])
@login_required
@admin_required
def whatsapp_template_edit(tpl_id):
    t = MessageTemplate.query.get_or_404(tpl_id)
    t.label = (request.form.get('label') or t.label).strip()
    t.meta_name = (request.form.get('meta_name') or t.meta_name).strip()
    t.category = request.form.get('category', t.category)
    t.lang = (request.form.get('lang') or t.lang).strip() or 'en'
    t.body_preview = (request.form.get('body_preview') or t.body_preview).strip()
    t.var_fields = (request.form.get('var_fields') or '').strip()
    db.session.commit()
    flash(f'Template "{t.label}" updated.')
    return redirect(url_for('whatsapp_templates'))

@app.route('/whatsapp/templates/<int:tpl_id>/toggle', methods=['POST'])
@login_required
@admin_required
def whatsapp_template_toggle(tpl_id):
    t = MessageTemplate.query.get_or_404(tpl_id)
    t.active = not t.active
    db.session.commit()
    flash(f'Template "{t.label}" {"activated" if t.active else "deactivated"}.')
    return redirect(url_for('whatsapp_templates'))

@app.route('/whatsapp/templates/<int:tpl_id>/delete', methods=['POST'])
@login_required
@admin_required
def whatsapp_template_delete(tpl_id):
    t = MessageTemplate.query.get_or_404(tpl_id)
    # keep past broadcasts (they store their own label) but drop the FK
    Broadcast.query.filter_by(template_id=t.id).update({'template_id': None})
    db.session.delete(t)
    db.session.commit()
    flash('Template deleted.')
    return redirect(url_for('whatsapp_templates'))

@app.route('/whatsapp/send-template', methods=['POST'])
@login_required
def whatsapp_send_template():
    """Send an approved template to one contact (works outside the 24h window).
    Destination = explicit wa_id (from a thread) or the customer's best number."""
    from whatsapp_webhook import send_template, log_message, normalize_phone
    tpl = MessageTemplate.query.get_or_404(request.form.get('template_id', type=int))
    back = request.form.get('return_url') or request.referrer or url_for('whatsapp_inbox')
    customer = Customer.query.get(request.form.get('customer_id', type=int) or 0)
    to = (request.form.get('wa_id') or '').strip()
    if not to and customer:
        to = customer.whatsapp or customer.mobile or customer.phone or customer.phone2 or ''
    to = normalize_phone(to)
    if not to:
        flash('No WhatsApp number on record for this contact.', 'error')
        return redirect(back)
    params, i = [], 1
    while f'param_{i}' in request.form:
        params.append((request.form.get(f'param_{i}') or '').strip())
        i += 1
    wam = send_template(to, tpl.meta_name, params=params or None, lang=tpl.lang or 'en')
    body = tpl.body_preview
    for n, v in enumerate(params, start=1):
        body = body.replace('{{%d}}' % n, v)
    log_message(to, 'out', body, msg_type='template', wam_id=wam,
                handled_by=session.get('user_name', 'staff'),
                status='sent' if wam else 'failed',
                customer_id=customer.id if customer else None)
    if wam:
        flash(f'"{tpl.label}" sent on WhatsApp.')
    else:
        flash('WhatsApp send failed — check the number and that this template name is approved in Meta.', 'error')
    return redirect(back)

@app.route('/whatsapp/broadcast')
@login_required
@admin_required
def whatsapp_broadcast():
    """Filter customers and preview a bulk WhatsApp send (admin only)."""
    applied = bool(request.args.get('apply'))
    matched, with_wa, without_wa = [], [], []
    if applied:
        for c in broadcast_filter_customers(request.args):
            (with_wa if _cust_wa_number(c) else without_wa).append(c)
        matched = with_wa + without_wa
    # Distinct dropdown values from existing data
    def _distinct(col):
        rows = db.session.query(col).filter(col.isnot(None), col != '').distinct().all()
        return sorted({r[0] for r in rows})
    templates = MessageTemplate.query.filter_by(active=True).order_by(MessageTemplate.label).all()
    tpl_ctx = [{'id': t.id, 'label': t.label, 'category': t.category, 'body': t.body_preview,
                'vars': [k.strip() for k in (t.var_fields or '').split(',') if k.strip()]}
               for t in templates]
    today_sent = WhatsAppMessage.query.filter(
        WhatsAppMessage.direction == 'out', WhatsAppMessage.msg_type == 'template',
        WhatsAppMessage.created_at >= now_dubai().replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    daily_cap = int(os.environ.get('WA_DAILY_CAP', '250'))
    recent = Broadcast.query.order_by(Broadcast.created_at.desc()).limit(8).all()
    return render_template('whatsapp_broadcast.html',
        applied=applied, args=request.args, with_wa=with_wa, without_wa=without_wa,
        matched=matched, templates=tpl_ctx, var_labels=WA_VAR_LABELS,
        jurisdictions=_distinct(Customer.jurisdiction), emirates=_distinct(Customer.emirate),
        authorities=_distinct(Customer.licensing_authority),
        nationalities=_distinct(Customer.nationality),
        today_sent=today_sent, daily_cap=daily_cap, recent=recent, now=now_dubai())

@app.route('/whatsapp/broadcast/export')
@login_required
@admin_required
def whatsapp_broadcast_export():
    """Download the filtered customer list as Excel (same filters as the preview)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    custs = broadcast_filter_customers(request.args)
    wb = Workbook(); ws = wb.active; ws.title = 'Customers'
    headers = ['Name', 'Company', 'Assigned Rep', 'Business Activity', 'Type', 'Jurisdiction',
               'Emirate', 'Authority', 'Nationality', 'WhatsApp', 'Email', 'Next Doc Expiry']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    today = now_dubai().date()
    for c in custs:
        docs = Document.query.filter_by(customer_id=c.id).filter(Document.expiry_date.isnot(None)).all()
        next_exp = min([d.expiry_date.date() for d in docs], default=None)
        ws.append([c.name or '', c.company or '', (c.rep.name if c.rep else ''),
                   c.business_activity or '', c.customer_type or '',
                   c.jurisdiction or '', c.emirate or '', c.licensing_authority or '',
                   c.nationality or '', _cust_wa_number(c) or '', c.email or '',
                   next_exp.strftime('%d/%m/%Y') if next_exp else ''])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 14)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_broadcast_list.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/whatsapp/broadcast/search')
@login_required
@admin_required
def whatsapp_broadcast_search():
    """Typeahead for manually adding recipients — clients with a WhatsApp number."""
    q = (request.args.get('q') or '').strip()
    out = []
    if len(q) >= 2:
        like = f'%{q}%'
        rows = Customer.query.filter(db.or_(
            Customer.name.ilike(like), Customer.company.ilike(like),
            Customer.phone.ilike(like), Customer.mobile.ilike(like),
            Customer.whatsapp.ilike(like), Customer.phone2.ilike(like),
        )).order_by(Customer.name).limit(20).all()
        for c in rows:
            wa = _cust_wa_number(c)
            if not wa:
                continue
            out.append({'id': c.id, 'name': c.name or '', 'company': c.company or '',
                        'rep': (c.rep.name if c.rep else ''),
                        'activity': c.business_activity or '', 'phone': wa})
    return jsonify(out)

@app.route('/whatsapp/broadcast/import-excel', methods=['POST'])
@login_required
@admin_required
def whatsapp_broadcast_import_excel():
    """Parse an uploaded Excel of EXTERNAL (non-CRM) numbers for a broadcast.
    Columns: Name, Phone, Consent (Yes/No), Company. Only Consent=Yes rows are
    returned; blanks/no-consent/no-phone/duplicates are skipped and counted.
    Nothing is stored — these are used for a one-off send + log only."""
    from openpyxl import load_workbook
    from whatsapp_webhook import normalize_phone
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'No file uploaded.'}), 400
    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return jsonify({'error': 'Could not read that file — please upload a valid .xlsx.'}), 400
    if not rows:
        return jsonify({'recipients': [], 'stats': {'loaded': 0}})
    header = [(str(c).strip().lower() if c is not None else '') for c in rows[0]]
    def col(*names):
        for nm in names:
            if nm in header:
                return header.index(nm)
        return None
    ci_name = col('name', 'customer', 'full name')
    ci_phone = col('phone', 'number', 'mobile', 'whatsapp', 'phone number', 'contact')
    ci_consent = col('consent', 'opt-in', 'optin', 'opted in')
    ci_company = col('company', 'company name', 'business')
    has_header = any(x is not None for x in (ci_name, ci_phone, ci_consent, ci_company))
    data = rows[1:] if has_header else rows
    if ci_name is None:
        ci_name = 0
    if ci_phone is None:
        ci_phone = 1 if (rows and len(rows[0]) > 1) else 0
    def cell(r, i):
        return (str(r[i]).strip() if (i is not None and i < len(r) and r[i] is not None) else '')
    seen = set()
    out, skip_consent, skip_phone, dupes = [], 0, 0, 0
    for r in data:
        phone = normalize_phone(cell(r, ci_phone))
        if not phone or len(phone) < 8:
            skip_phone += 1
            continue
        if ci_consent is not None:
            consent = cell(r, ci_consent).lower()
            if consent not in ('yes', 'y', 'true', '1', 'opted in', 'consent', 'opt-in'):
                skip_consent += 1
                continue
        if phone in seen:
            dupes += 1
            continue
        seen.add(phone)
        out.append({'phone': phone, 'name': cell(r, ci_name), 'company': cell(r, ci_company)})
    return jsonify({'recipients': out, 'has_consent_col': ci_consent is not None,
                    'stats': {'loaded': len(out), 'skip_consent': skip_consent,
                              'skip_phone': skip_phone, 'dupes': dupes}})

@app.route('/whatsapp/broadcast/sample-excel')
@login_required
@admin_required
def whatsapp_broadcast_sample_excel():
    """A blank Excel with the exact columns the import expects, so staff fill it right."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook(); ws = wb.active; ws.title = 'Numbers'
    for i, h in enumerate(['Name', 'Phone', 'Consent', 'Company'], 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    ws.append(['Ahmed Ali', '971501234567', 'Yes', 'Al Noor Trading'])
    ws.append(['Sara Khan', '971559876543', 'Yes', ''])
    ws.append(['(no consent — will be skipped)', '971500000000', 'No', ''])
    for i, w in enumerate([28, 20, 12, 24], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_broadcast_numbers_sample.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _run_broadcast(app_obj, broadcast_id, recipients, template_id, custom_vars, sender_name):
    """Background worker: send one approved template to many recipients, throttled.
    recipients = list of dicts {'cust_id': int|None, 'to': digits, 'first': str, 'company': str}.
    External (Excel-imported) recipients have cust_id=None and carry their own first/company."""
    import time
    from whatsapp_webhook import send_template, log_message
    with app_obj.app_context():
        tpl = MessageTemplate.query.get(template_id)
        bc = Broadcast.query.get(broadcast_id)
        if not tpl or not bc:
            return
        keys = [k.strip() for k in (tpl.var_fields or '').split(',') if k.strip()]
        for r in recipients:
            customer = Customer.query.get(r['cust_id']) if r.get('cust_id') else None
            params, ci = [], 0
            for k in keys:
                if k == 'custom':
                    params.append(custom_vars[ci] if ci < len(custom_vars) else '')
                    ci += 1
                elif customer:
                    params.append(_wa_resolve_var(k, customer))
                elif k in ('first_name', 'full_name'):
                    params.append(r.get('first') or '')
                elif k == 'company':
                    params.append(r.get('company') or '')
                elif k == 'service':
                    params.append(r.get('service') or '')
                else:
                    params.append('')
            to = r['to']
            wam = send_template(to, tpl.meta_name, params=params or None, lang=tpl.lang or 'en')
            body = tpl.body_preview
            for n, v in enumerate(params, start=1):
                body = body.replace('{{%d}}' % n, v)
            log_message(to, 'out', body, msg_type='template', wam_id=wam,
                        handled_by=sender_name, status='sent' if wam else 'failed',
                        customer_id=r.get('cust_id'))
            if wam:
                bc.sent = (bc.sent or 0) + 1
            else:
                bc.failed = (bc.failed or 0) + 1
            db.session.commit()
            time.sleep(1)  # ~1 msg/sec — stays well under Meta's rate limits
        bc.status = 'done'
        db.session.commit()

@app.route('/whatsapp/broadcast/send', methods=['POST'])
@login_required
@admin_required
def whatsapp_broadcast_send():
    import threading
    from whatsapp_webhook import normalize_phone
    back = request.form.get('return_url') or url_for('whatsapp_broadcast')
    tpl = MessageTemplate.query.get_or_404(request.form.get('template_id', type=int))
    custom_vars = []
    i = 1
    while f'custom_{i}' in request.form:
        custom_vars.append((request.form.get(f'custom_{i}') or '').strip())
        i += 1
    recipients, seen = [], set()
    # CRM customers
    for cid in request.form.getlist('customer_ids'):
        c = Customer.query.get(int(cid)) if str(cid).isdigit() else None
        if not c:
            continue
        to = _cust_wa_number(c)
        if to and to not in seen:
            seen.add(to)
            recipients.append({'cust_id': c.id, 'to': to,
                               'first': ((c.contact_person or c.name or '').split() or [''])[0],
                               'company': c.trade_name or c.company or ''})
    # External recipients — Excel import or re-engage leads (parallel arrays)
    ext_phones = request.form.getlist('ext_phone')
    ext_firsts = request.form.getlist('ext_first')
    ext_comps = request.form.getlist('ext_company')
    ext_services = request.form.getlist('ext_service')
    for idx, ph in enumerate(ext_phones):
        to = normalize_phone(ph)
        if not to or to in seen:
            continue
        seen.add(to)
        recipients.append({'cust_id': None, 'to': to,
                           'first': (ext_firsts[idx] if idx < len(ext_firsts) else ''),
                           'company': (ext_comps[idx] if idx < len(ext_comps) else ''),
                           'service': (ext_services[idx] if idx < len(ext_services) else '')})
    if not recipients:
        flash('No recipients with a WhatsApp number were selected.', 'error')
        return redirect(back)
    # Daily-cap guard (Meta limits business-initiated conversations per 24h)
    daily_cap = int(os.environ.get('WA_DAILY_CAP', '250'))
    today_sent = WhatsAppMessage.query.filter(
        WhatsAppMessage.direction == 'out', WhatsAppMessage.msg_type == 'template',
        WhatsAppMessage.created_at >= now_dubai().replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    if today_sent + len(recipients) > daily_cap:
        flash(f'This send ({len(recipients)}) would exceed today\'s cap of {daily_cap} '
              f'({today_sent} already sent). Reduce the list or raise WA_DAILY_CAP.', 'error')
        return redirect(back)
    bc = Broadcast(template_id=tpl.id, template_label=tpl.label,
                   filter_summary=(request.form.get('filter_summary') or '')[:400],
                   total=len(recipients), sent=0, failed=0, status='sending',
                   created_by=session.get('user_id'), created_at=now_dubai())
    db.session.add(bc); db.session.commit()
    threading.Thread(target=_run_broadcast,
                     args=(app, bc.id, recipients, tpl.id, custom_vars,
                           session.get('user_name', 'staff')),
                     daemon=True).start()
    flash(f'Broadcast started — sending "{tpl.label}" to {len(recipients)} recipients. '
          f'Progress updates below.')
    return redirect(back)

@app.route('/whatsapp/reengage')
@login_required
@admin_required
def whatsapp_reengage():
    """Win-back tool: filter LOST leads (excluding junk) by date range and send an
    approved re-engagement template. Reuses the broadcast send path (leads are
    submitted as external recipients; replies auto-link to the lead by phone)."""
    applied = bool(request.args.get('apply'))
    leads = []
    lead_statuses = ['Lost', 'New', 'Contacted', 'Qualified', 'Proposal', 'Future']  # Converted excluded — wrong audience for win-back
    if applied:
        q = Lead.query.filter(Lead.phone.isnot(None), Lead.phone != '')
        status = (request.args.get('status') or 'Lost').strip()
        if status and status != 'All':
            q = q.filter(Lead.status == status)
        quality = (request.args.get('quality') or 'exclude_junk').strip()
        if quality == 'exclude_junk':
            q = q.filter(db.or_(Lead.genuine.is_(None), Lead.genuine != 'Junk'))
        elif quality == 'junk_only':
            q = q.filter(Lead.genuine == 'Junk')
        # 'all' → no quality filter
        df = (request.args.get('from') or '').strip()
        dt = (request.args.get('to') or '').strip()
        if df:
            try:
                q = q.filter(Lead.created_at >= datetime.strptime(df, '%Y-%m-%d'))
            except Exception:
                pass
        if dt:
            try:
                q = q.filter(Lead.created_at < datetime.strptime(dt, '%Y-%m-%d') + timedelta(days=1))
            except Exception:
                pass
        src = (request.args.get('source') or '').strip()
        if src and src != 'All':
            q = q.filter(Lead.source == src)
        leads = q.order_by(Lead.created_at.desc()).all()
    sources = sorted({r[0] for r in db.session.query(Lead.source)
                      .filter(Lead.source.isnot(None), Lead.source != '').distinct().all()})
    daily_cap = int(os.environ.get('WA_DAILY_CAP', '250'))
    today_sent = WhatsAppMessage.query.filter(
        WhatsAppMessage.direction == 'out', WhatsAppMessage.msg_type == 'template',
        WhatsAppMessage.created_at >= now_dubai().replace(hour=0, minute=0, second=0, microsecond=0)
    ).count()
    recent = Broadcast.query.order_by(Broadcast.created_at.desc()).limit(6).all()
    return render_template('reengage.html', applied=applied, args=request.args, leads=leads,
                           templates=wa_send_context(), sources=sources, lead_statuses=lead_statuses,
                           today_sent=today_sent, daily_cap=daily_cap, recent=recent, now=now_dubai())

@app.route('/whatsapp/task-updates')
@login_required
@admin_required
def whatsapp_task_updates():
    """Bulk task status broadcast: pick active tasks, send each customer the approved
    status_update template. Dormant until that template is active."""
    applied = bool(request.args.get('apply'))
    status_f = (request.args.get('status') or '').strip()
    active_q = Job.query.filter(Job.status.notin_(['Done', 'Closed', 'Closed - Pending Partner Commission']))
    jobs = []
    if applied:
        q = active_q.options(db.joinedload(Job.customer))
        if status_f and status_f != 'All':
            q = q.filter(Job.status == status_f)
        jobs = [j for j in q.order_by(Job.due_date.asc()).all()
                if j.customer and _cust_wa_number(j.customer)]
    statuses = sorted({j.status for j in active_q.all() if j.status})
    tpl = wa_template_active(get_setting('wa_status_template', 'status_update') or 'status_update')
    return render_template('task_updates.html', applied=applied, jobs=jobs, status_f=status_f,
                           statuses=statuses, tpl_active=bool(tpl), now=now_dubai())

@app.route('/whatsapp/task-updates/send', methods=['POST'])
@login_required
@admin_required
def whatsapp_task_updates_send():
    from whatsapp_webhook import send_template, log_message
    tpl = wa_template_active(get_setting('wa_status_template', 'status_update') or 'status_update')
    if not tpl:
        flash('Activate the status_update template in WhatsApp → Templates first.', 'error')
        return redirect(url_for('whatsapp_task_updates'))
    status_text = (request.form.get('status_text') or '').strip()
    job_ids = request.form.getlist('job_ids')
    if not status_text or not job_ids:
        flash('Pick a status and at least one task.', 'error')
        return redirect(url_for('whatsapp_task_updates'))
    sent = 0
    for jid in job_ids:
        job = Job.query.get(int(jid)) if str(jid).isdigit() else None
        if not job or not job.customer:
            continue
        to = _cust_wa_number(job.customer)
        if not to:
            continue
        first = ((job.customer.contact_person or job.customer.name or 'there').split() or ['there'])[0]
        params = [first, job.job_type or 'application', status_text]
        wam = send_template(to, tpl.meta_name, params=params, lang=tpl.lang or 'en')
        body = tpl.body_preview or ''
        for n, v in enumerate(params, start=1):
            body = body.replace('{{%d}}' % n, v)
        log_message(to, 'out', body, msg_type='template', wam_id=wam,
                    handled_by=session.get('user_name', 'staff'),
                    status='sent' if wam else 'failed', customer_id=job.customer.id)
        if wam:
            sent += 1
    flash(f'Status update sent to {sent} customer(s).')
    return redirect(url_for('whatsapp_task_updates'))

@app.route('/whatsapp/test-send', methods=['GET', 'POST'])
@login_required
@admin_required
def whatsapp_test_send():
    """Send the welcome template to a number on demand and show Meta's exact result
    — so we can diagnose failures live without digging in Railway logs."""
    result = None
    number = (request.form.get('number') or request.args.get('number') or '').strip()
    if request.method == 'POST' and number:
        from whatsapp_webhook import send_template, last_send_error, normalize_phone, log_message
        to = normalize_phone(number)
        tmpl = os.environ.get('WA_WELCOME_TEMPLATE', 'general')
        lang = os.environ.get('WA_WELCOME_LANG', 'en_GB')
        pname = os.environ.get('WA_WELCOME_PARAM_NAME', 'customer_name')
        param_names = [pname] if pname else None
        wam = send_template(to, tmpl, params=['Test'], lang=lang, param_names=param_names)
        # Log the test like any other outbound — CRITICAL: the wam_id row is what
        # Meta's delivery receipt attaches to, so tick-status + failure reasons work.
        log_message(to, 'out', f'[test] Welcome template ({tmpl}) — test send', msg_type='template',
                    wam_id=wam, handled_by=session.get('user_name', 'admin'),
                    status='sent' if wam else 'failed')
        if wam:
            result = {'ok': True, 'msg': f'✅ Meta accepted it (message id …{wam[-12:]}). Now open this number\'s chat in the CRM: '
                                          f'one tick = sent, two ticks = DELIVERED (proof). If it fails on delivery instead, '
                                          f'it will appear on Failed sends with Meta\'s reason.'}
        else:
            result = {'ok': False, 'msg': last_send_error() or 'Failed, but Meta returned no reason.'}
    return render_template('wa_test_send.html', result=result, number=number)

@app.route('/whatsapp/failures')
@login_required
@admin_required
def whatsapp_failures():
    """Recent failed outgoing WhatsApp sends + Meta's reason — so admins can see
    'why' directly in the CRM instead of digging through Railway logs."""
    fails = (WhatsAppMessage.query.filter_by(direction='out', status='failed')
             .order_by(WhatsAppMessage.created_at.desc()).limit(60).all())
    return render_template('whatsapp_failures.html', fails=fails)

@app.route('/whatsapp/<wa_id>/done', methods=['POST'])
@login_required
def whatsapp_mark_done(wa_id):
    """Mark a conversation resolved (Done) — it leaves the active list. Toggles back."""
    thread = WhatsAppThread.query.get(wa_id)
    if not thread:
        thread = WhatsAppThread(wa_id=wa_id)
        db.session.add(thread)
    resolving = not thread.resolved
    thread.resolved = resolving
    thread.resolved_at = now_dubai() if resolving else None
    thread.resolved_by = session.get('user_name') if resolving else None
    db.session.commit()
    flash('Marked done — moved to the Done filter.' if resolving else 'Conversation reopened.')
    return redirect(request.referrer or url_for('whatsapp_inbox'))

@app.route('/whatsapp/bulk-done', methods=['POST'])
@login_required
def whatsapp_bulk_done():
    """Mark several conversations Done at once (from the inbox checkboxes)."""
    wa_ids = request.form.getlist('wa_ids')
    count = 0
    for wid in wa_ids:
        thread = WhatsAppThread.query.get(wid)
        if not thread:
            thread = WhatsAppThread(wa_id=wid)
            db.session.add(thread)
        if not thread.resolved:
            thread.resolved = True
            thread.resolved_at = now_dubai()
            thread.resolved_by = session.get('user_name')
            count += 1
    db.session.commit()
    flash(f'{count} conversation(s) marked done.')
    return redirect(request.referrer or url_for('whatsapp_inbox'))

@app.route('/whatsapp/<wa_id>/bot-toggle', methods=['POST'])
@login_required
def whatsapp_bot_toggle(wa_id):
    """Pause/resume the auto-reply bot for one conversation (human takeover)."""
    thread = WhatsAppThread.query.get(wa_id)
    if not thread:
        thread = WhatsAppThread(wa_id=wa_id)
        db.session.add(thread)
    thread.bot_paused = not thread.bot_paused
    thread.bot_paused_by = session.get('user_name') if thread.bot_paused else None
    db.session.commit()
    flash(f"Bot {'paused' if thread.bot_paused else 'resumed'} for this conversation.")
    return redirect(url_for('whatsapp_thread', wa_id=wa_id))

@app.route('/whatsapp/<wa_id>/delete', methods=['POST'])
@login_required
def whatsapp_delete_thread(wa_id):
    """Delete an entire WhatsApp conversation (all messages + thread state)."""
    WhatsAppMessage.query.filter_by(wa_id=wa_id).delete()
    WhatsAppThread.query.filter_by(wa_id=wa_id).delete()
    db.session.commit()
    flash('Conversation deleted.')
    return redirect(url_for('whatsapp_inbox'))

@app.route('/whatsapp/<wa_id>/message/<int:msg_id>/delete', methods=['POST'])
@login_required
def whatsapp_delete_message(wa_id, msg_id):
    """Delete a single message from a conversation."""
    msg = WhatsAppMessage.query.get_or_404(msg_id)
    if msg.wa_id != wa_id:
        return 'Not found', 404
    db.session.delete(msg)
    db.session.commit()
    return ('', 204)

@app.route('/whatsapp/<wa_id>/convert', methods=['POST'])
@login_required
def whatsapp_convert(wa_id):
    """Flow B — staff turns a logged conversation into a CRM lead (round-robin assigned)."""
    from meta_webhook import get_next_sales_staff
    existing = WhatsAppMessage.query.filter_by(wa_id=wa_id, direction='in')\
            .order_by(WhatsAppMessage.created_at.asc()).first()
    name = (request.form.get('name') or
            next((m.contact_name for m in WhatsAppMessage.query.filter_by(wa_id=wa_id).all() if m.contact_name), None) or
            f'WhatsApp {wa_id}')
    # Keep the lead with the rep already handling this chat (the reviewer); else round-robin.
    thread = WhatsAppThread.query.get(wa_id)
    rep_id = thread.assigned_to if (thread and thread.assigned_to) else None
    if not rep_id:
        assigned = get_next_sales_staff(db, User, Lead)
        rep_id = assigned.id if assigned else None
    lead = Lead(
        name=name.title(), phone=wa_id, source='WhatsApp - AI Bot', sub_source='WhatsApp Bot',
        lead_type='New', status='New', representative=session.get('user_name'),
        assigned_to=rep_id,
        created_at=now_dubai(), due_date=now_dubai() + timedelta(days=1),
        remarks='Created from WhatsApp conversation.',
    )
    db.session.add(lead)
    db.session.flush()
    # link this contact's whole history to the new lead
    WhatsAppMessage.query.filter_by(wa_id=wa_id).update({'lead_id': lead.id})
    db.session.add(LeadUpdate(lead_id=lead.id, stage='New — WhatsApp',
        remark='Lead created from WhatsApp inbox.', staff_name=session.get('user_name', 'System'),
        created_at=now_dubai()))
    db.session.commit()
    flash(f'Lead created and assigned to {assigned.name if assigned else "nobody"}.')
    return redirect(url_for('lead_detail', lead_id=lead.id))

@app.route('/partners')
@login_required
def partners():
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('dashboard'))
    
    partners = Partner.query.order_by(Partner.name).all()
    return render_template('partners.html', partners=partners)

@app.route('/partners/add', methods=['POST'])
@login_required
def add_partner():
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    name = request.form.get('name', '').strip()
    if not name:
        flash('Partner name is required.', 'error')
        return redirect(url_for('partners'))
    
    existing = Partner.query.filter_by(name=name).first()
    if existing:
        flash('Partner already exists.', 'error')
        return redirect(url_for('partners'))
    
    partner = Partner(name=name)
    db.session.add(partner)
    db.session.commit()
    flash(f'Partner "{name}" added successfully.')
    return redirect(url_for('partners'))

@app.route('/partners/<int:partner_id>/edit', methods=['POST'])
@login_required
def edit_partner(partner_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    partner = Partner.query.get_or_404(partner_id)
    new_name = request.form.get('name', '').strip()
    
    if not new_name:
        flash('Partner name cannot be empty.', 'error')
        return redirect(url_for('partners'))
    
    # Check if new name already exists (excluding current partner)
    existing = Partner.query.filter(Partner.name == new_name, Partner.id != partner_id).first()
    if existing:
        flash('Partner name already exists.', 'error')
        return redirect(url_for('partners'))
    
    old_name = partner.name
    partner.name = new_name
    db.session.commit()
    flash(f'Partner renamed from "{old_name}" to "{new_name}".')
    return redirect(url_for('partners'))

@app.route('/partners/<int:partner_id>/delete', methods=['POST'])
@login_required
def delete_partner(partner_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    partner = Partner.query.get_or_404(partner_id)
    
    # Check if partner has any jobs associated
    jobs_with_partner = Job.query.filter_by(partner_name=partner.name).first()
    if jobs_with_partner:
        flash(f'Cannot delete "{partner.name}" - has associated transactions. Deactivate instead.', 'error')
        return redirect(url_for('partners'))
    
    name = partner.name
    db.session.delete(partner)
    db.session.commit()
    flash(f'Partner "{name}" deleted successfully.')
    return redirect(url_for('partners'))

@app.route('/partners/<int:partner_id>/toggle', methods=['POST'])
@login_required
def toggle_partner(partner_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    partner = Partner.query.get_or_404(partner_id)
    partner.active = not partner.active
    db.session.commit()
    status = 'activated' if partner.active else 'deactivated'
    flash(f'Partner "{partner.name}" {status}.')
    return redirect(url_for('partners'))

@app.route('/partner-commissions')
@login_required
def partner_commissions():
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('dashboard'))
    
    try:
        # Get all jobs with pending partner commissions
        all_pending = Job.query.filter_by(partner_commission_expected=True, partner_status='Pending').all()
    except Exception as e:
        print(f"Error querying pending jobs: {e}")
        all_pending = []
    
    # Filter options
    partner_filter = request.args.get('partner', '')
    status_filter = request.args.get('status', 'pending')
    
    # Apply filters
    try:
        if status_filter == 'pending':
            jobs = all_pending
        elif status_filter == 'received':
            jobs = Job.query.filter_by(partner_commission_expected=True, partner_status='Received').all()
        else:  # all
            jobs = Job.query.filter_by(partner_commission_expected=True).all()
    except Exception as e:
        print(f"Error filtering jobs: {e}")
        jobs = []
    
    if partner_filter:
        jobs = [j for j in jobs if j.partner_name == partner_filter]
    
    # Get unique partners for filter dropdown
    try:
        all_partners = Partner.query.filter_by(active=True).order_by(Partner.name).all()
    except Exception as e:
        print(f"Error querying partners: {e}")
        all_partners = []
    
    # Calculate totals
    try:
        total_pending = sum((j.partner_amount or 0) for j in all_pending)
        total_received = sum((j.partner_amount or 0) for j in Job.query.filter_by(partner_commission_expected=True, partner_status='Received').all())
    except Exception as e:
        print(f"Error calculating totals: {e}")
        total_pending = 0
        total_received = 0
    
    now = now_dubai()
    
    return render_template('partner_commissions.html', 
                          jobs=jobs, 
                          all_partners=all_partners,
                          partner_filter=partner_filter,
                          status_filter=status_filter,
                          total_pending=total_pending,
                          total_received=total_received,
                          now=now)

@app.route('/partner-commissions/<int:job_id>/mark-received', methods=['POST'])
@login_required
def mark_partner_received(job_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partner_commissions'))
    
    job = Job.query.get_or_404(job_id)
    customer = Customer.query.get(job.customer_id)
    
    if not job.partner_commission_expected or job.partner_status != 'Pending':
        flash('This task does not have a pending partner commission.', 'error')
        return redirect(url_for('partner_commissions'))
    
    try:
        # Get revenue amount from form (user enters it)
        revenue_amount = request.form.get('revenue_amount')
        if not revenue_amount:
            flash('Revenue amount is required.', 'error')
            return redirect(url_for('partner_commissions'))
        
        revenue = float(revenue_amount)
        invoice_amount = float(job.amount_invoiced or 0)
        partner_amount = float(job.partner_amount or 0)
        
        # Validation
        if revenue < 0:
            flash('Revenue cannot be negative.', 'error')
            return redirect(url_for('partner_commissions'))
        
        if revenue > invoice_amount:
            flash(f'Revenue (AED {revenue:,.0f}) cannot exceed invoice amount (AED {invoice_amount:,.0f}).', 'error')
            return redirect(url_for('partner_commissions'))
        
        if revenue > partner_amount:
            flash(f'Revenue (AED {revenue:,.0f}) cannot exceed partner reimbursement (AED {partner_amount:,.0f}).', 'error')
            return redirect(url_for('partner_commissions'))
        
        # Mark as received and book revenue
        job.partner_status = 'Received'
        job.partner_received_date = now_dubai().date()
        job.revenue = revenue
        job.revenue_date = now_dubai().date()
        job.status = 'Closed'
        
        # Add timeline update
        remark = f'Partner commission RECEIVED from {job.partner_name}: AED {partner_amount:,.0f}. Revenue booked: AED {revenue:,.0f} for {now_dubai().strftime("%B %Y")} (cash-basis). Marked by {session["user_name"]}.'
        update = JobUpdate(job_id=job.id, status='Closed', remark=remark, staff_name=session['user_name'])
        db.session.add(update)
        db.session.commit()
        
        flash(f'✓ Partner commission from {job.partner_name} marked received. Revenue AED {revenue:,.0f} booked.')
        return redirect(url_for('partner_commissions'))
        
    except ValueError:
        flash('Invalid revenue amount.', 'error')
        return redirect(url_for('partner_commissions'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        print(f"Error marking partner received: {e}")
        return redirect(url_for('partner_commissions'))

@app.route('/partner-commissions/<int:job_id>/edit-revenue', methods=['POST'])
@login_required
def edit_partner_revenue(job_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partner_commissions'))
    
    job = Job.query.get_or_404(job_id)
    
    if job.partner_status != 'Received':
        flash('Can only edit revenue for received commissions.', 'error')
        return redirect(url_for('partner_commissions'))
    
    try:
        revenue_amount = request.form.get('revenue_amount')
        if not revenue_amount:
            flash('Revenue amount is required.', 'error')
            return redirect(url_for('partner_commissions'))
        
        revenue = float(revenue_amount)
        invoice_amount = float(job.amount_invoiced or 0)
        partner_amount = float(job.partner_amount or 0)
        old_revenue = job.revenue
        
        # Validation
        if revenue < 0:
            flash('Revenue cannot be negative.', 'error')
            return redirect(url_for('partner_commissions'))
        
        if revenue > invoice_amount:
            flash(f'Revenue (AED {revenue:,.0f}) cannot exceed invoice amount (AED {invoice_amount:,.0f}).', 'error')
            return redirect(url_for('partner_commissions'))
        
        if revenue > partner_amount:
            flash(f'Revenue (AED {revenue:,.0f}) cannot exceed partner reimbursement (AED {partner_amount:,.0f}).', 'error')
            return redirect(url_for('partner_commissions'))
        
        # Update revenue
        job.revenue = revenue
        job.revenue_date = now_dubai().date()
        
        # Add timeline update
        remark = f'Partner commission revenue updated: AED {old_revenue:,.0f} → AED {revenue:,.0f}. Edited by {session["user_name"]}.'
        update = JobUpdate(job_id=job.id, status='Closed', remark=remark, staff_name=session['user_name'])
        db.session.add(update)
        db.session.commit()
        
        flash(f'✓ Revenue updated to AED {revenue:,.0f}')
        return redirect(url_for('partner_commissions'))
        
    except ValueError:
        flash('Invalid revenue amount.', 'error')
        return redirect(url_for('partner_commissions'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        print(f"Error editing partner revenue: {e}")
        return redirect(url_for('partner_commissions'))

# ── TAHFEEL DOCUMENTS MANAGEMENT
@app.route('/tahfeel-doc')
@login_required
def tahfeel_doc():
    # Access: Admin OR Saada only
    if session.get('role') != 'admin' and session.get('user_email') != 'saadatahfeel@gmail.com':
        flash('Access denied.')
        return redirect(url_for('dashboard'))

    try:
        all_docs = CompanyDocument.query.order_by(CompanyDocument.expiry_date).all()
        expiring_docs = [d for d in all_docs if d.expiry_status() in ['critical', 'warning', 'expired']]
        critical_count = len([d for d in all_docs if d.expiry_status() == 'critical'])
        warning_count = len([d for d in all_docs if d.expiry_status() == 'warning'])
        expired_count = len([d for d in all_docs if d.expiry_status() == 'expired'])

        # Tahfeel = company docs (legacy/null category counts as Tahfeel)
        tahfeel_docs = [d for d in all_docs if (d.category or 'Tahfeel') == 'Tahfeel']

        # Staff/Management: group docs under each person record; returns [(name, [docs]), ...]
        def grouped(cat):
            people = {p.id: p for p in TahfeelStaff.query.filter_by(category=cat).all()}
            by_id, no_person = {}, {}
            for d in all_docs:
                if d.category != cat:
                    continue
                if d.staff_id:
                    by_id.setdefault(d.staff_id, []).append(d)
                else:
                    no_person.setdefault(d.owner or '—', []).append(d)
            result = []
            for pid, docs in by_id.items():
                person = people.get(pid)
                result.append((person.name if person else (docs[0].owner or '—'), docs))
            for nm, docs in no_person.items():
                result.append((nm, docs))
            result.sort(key=lambda x: x[0].lower())
            return result
        staff_groups = grouped('Staff')
        mgmt_groups = grouped('Management')

        # Active people for the add-document dropdowns
        staff_people = [{'id': p.id, 'name': p.name} for p in
                        TahfeelStaff.query.filter_by(category='Staff', active=True).order_by(TahfeelStaff.name).all()]
        mgmt_people = [{'id': p.id, 'name': p.name} for p in
                       TahfeelStaff.query.filter_by(category='Management', active=True).order_by(TahfeelStaff.name).all()]

    except Exception as e:
        print(f"Error loading documents: {e}")
        all_docs = []
        expiring_docs = []
        critical_count = warning_count = expired_count = 0
        tahfeel_docs = []
        staff_groups = mgmt_groups = []
        staff_people = mgmt_people = []

    # Same admin-managed Document Type list used by the customer/company
    # Documents feature, so there's one list to maintain, not two.
    doc_type_names = [dt.name for dt in DocType.query.order_by(DocType.name).all()]

    return render_template('tahfeel_doc_simple.html',
                          all_docs=all_docs,
                          expiring_docs=expiring_docs,
                          tahfeel_docs=tahfeel_docs,
                          staff_groups=staff_groups,
                          mgmt_groups=mgmt_groups,
                          staff_people=staff_people,
                          mgmt_people=mgmt_people,
                          doc_type_names=doc_type_names,
                          critical_count=critical_count,
                          warning_count=warning_count,
                          expired_count=expired_count)

def _is_ajax():
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'

@app.route('/tahfeel-doc/add', methods=['POST'])
@login_required
def add_tahfeel_doc():
    ajax = _is_ajax()
    def fail(msg):
        if ajax:
            return jsonify(success=False, message=msg)
        flash(msg, 'error')
        return redirect(url_for('tahfeel_doc'))

    if session.get('role') != 'admin' and session.get('user_email') != 'saadatahfeel@gmail.com':
        return fail('Access denied.')

    try:
        category = request.form.get('category', '').strip()
        doc_type = request.form.get('doc_type', '').strip()
        authority = request.form.get('authority', '').strip()
        issue_date = request.form.get('issue_date')
        expiry_date = request.form.get('expiry_date')

        if category not in ('Tahfeel', 'Staff', 'Management') or not doc_type or not expiry_date:
            return fail('Please fill all required fields.')

        # Resolve owner: Tahfeel = company itself; Staff/Management = a person record
        staff_id = None
        owner = 'Tahfeel'
        if category == 'Tahfeel':
            pass  # keep authority
        else:
            authority = ''  # authority only applies to Tahfeel company docs
            person = _resolve_tahfeel_person(category)
            if person is None:
                return fail('Please select a person or add a new one.')
            staff_id = person.id
            owner = person.name

        # Parse dates
        try:
            issue_dt = datetime.strptime(issue_date, '%Y-%m-%d').date() if issue_date else None
            expiry_dt = datetime.strptime(expiry_date, '%Y-%m-%d').date()
        except ValueError:
            return fail('Invalid date format.')

        # Handle optional file upload
        doc_url = None
        cloudinary_id = None
        file = request.files.get('file')
        upload_warning = None
        if file and file.filename:
            try:
                upload_result = cloudinary.uploader.upload(
                    file, folder='tahfeel-documents', resource_type='auto',
                    access_mode='authenticated'
                )
                doc_url = upload_result['secure_url']
                cloudinary_id = upload_result['public_id']
            except Exception as e:
                print(f"Cloudinary upload error: {e}")
                upload_warning = 'File could not be uploaded — document saved without attachment.'

        doc = CompanyDocument(
            name=doc_type,            # name field kept for compatibility = the doc type
            doc_type=doc_type,
            category=category,
            staff_id=staff_id,
            issue_date=issue_dt,
            expiry_date=expiry_dt,
            authority=authority,
            owner=owner,
            document_url=doc_url,
            cloudinary_public_id=cloudinary_id,
            created_by=session['user_name']
        )
        db.session.add(doc)
        db.session.commit()

        if ajax:
            return jsonify(success=True, message=upload_warning or 'Document added successfully.',
                            warning=bool(upload_warning), person_id=staff_id)
        if upload_warning:
            flash(upload_warning, 'warning')
        flash('✓ Document added successfully.')
        return redirect(url_for('tahfeel_doc'))

    except Exception as e:
        db.session.rollback()
        print(f"Error adding document: {e}")
        return fail(f'Error: {str(e)}')


def _resolve_tahfeel_person(category):
    """Resolve the person for a Staff/Management doc from the submitted form.
    Returns a TahfeelStaff (existing or newly-created), or None if neither given."""
    sid = (request.form.get('staff_id') or '').strip()
    new_name = (request.form.get('new_person_name') or '').strip()
    if sid:
        return TahfeelStaff.query.get(int(sid))
    if new_name:
        person = TahfeelStaff.query.filter_by(name=new_name, category=category).first()
        if not person:
            person = TahfeelStaff(name=new_name, category=category)
            db.session.add(person)
            db.session.flush()
        return person
    return None

@app.route('/tahfeel-doc/<int:doc_id>/edit', methods=['POST'])
@login_required
def edit_tahfeel_doc(doc_id):
    ajax = _is_ajax()
    def fail(msg):
        if ajax:
            return jsonify(success=False, message=msg)
        flash(msg, 'error')
        return redirect(url_for('tahfeel_doc'))

    # Access: Admin OR Saada only
    if session.get('role') != 'admin' and session.get('user_email') != 'saadatahfeel@gmail.com':
        return fail('Access denied.')

    doc = CompanyDocument.query.get_or_404(doc_id)

    try:
        category = request.form.get('category', '').strip() or doc.category or 'Tahfeel'
        doc_type = request.form.get('doc_type', '').strip()
        issue_date = request.form.get('issue_date')
        expiry_date = request.form.get('expiry_date')
        authority = request.form.get('authority', '').strip()

        # Resolve owner/person
        staff_id = doc.staff_id
        owner = 'Tahfeel'
        if category == 'Tahfeel':
            staff_id = None
        else:
            authority = ''
            person = _resolve_tahfeel_person(category)
            if person is not None:
                staff_id = person.id
                owner = person.name
            elif doc.staff and doc.category == category:
                owner = doc.staff.name  # unchanged person
            else:
                return fail('Please select a person or add a new one.')

        # Validation
        if not doc_type or not expiry_date:
            return fail('All required fields must be filled.')

        # Parse dates
        try:
            issue_dt = datetime.strptime(issue_date, '%Y-%m-%d').date() if issue_date else None
            expiry_dt = datetime.strptime(expiry_date, '%Y-%m-%d').date()
        except ValueError:
            return fail('Invalid date format.')

        # Update document
        doc.name = doc_type
        doc.doc_type = doc_type
        doc.category = category
        doc.staff_id = staff_id
        doc.issue_date = issue_dt
        doc.expiry_date = expiry_dt
        doc.authority = authority
        doc.owner = owner

        upload_warning = None
        # Handle file replacement
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename:
                try:
                    # Delete old file if exists
                    if doc.cloudinary_public_id:
                        cloudinary.uploader.destroy(doc.cloudinary_public_id)
                    
                    # Upload new file
                    upload_result = cloudinary.uploader.upload(
                        file,
                        folder='tahfeel-documents',
                        resource_type='auto',
                        access_mode='authenticated'
                    )
                    doc.document_url = upload_result['secure_url']
                    doc.cloudinary_public_id = upload_result['public_id']
                except Exception as e:
                    print(f"Cloudinary upload error: {e}")
                    upload_warning = 'File update failed, but record updated.'

        db.session.commit()
        if ajax:
            return jsonify(success=True, message=upload_warning or 'Document updated successfully.',
                            warning=bool(upload_warning), person_id=staff_id)
        if upload_warning:
            flash(upload_warning, 'warning')
        flash(f'✓ Document "{doc_type}" updated successfully.')
        return redirect(url_for('tahfeel_doc'))

    except Exception as e:
        db.session.rollback()
        print(f"Error editing document: {e}")
        return fail(f'Error: {str(e)}')

@app.route('/tahfeel-doc/<int:doc_id>/upload', methods=['POST'])
@login_required
def upload_tahfeel_doc_file(doc_id):
    if session.get('role') != 'admin' and session.get('user_email') != 'saadatahfeel@gmail.com':
        flash('Access denied.', 'error')
        return redirect(url_for('tahfeel_doc'))
    
    doc = CompanyDocument.query.get_or_404(doc_id)
    
    try:
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('tahfeel_doc'))
        
        file = request.files['file']
        if not file or not file.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('tahfeel_doc'))
        
        # Delete old file if exists
        if doc.cloudinary_public_id:
            try:
                cloudinary.uploader.destroy(doc.cloudinary_public_id)
            except Exception as e:
                print(f"Cloudinary delete error: {e}")
        
        # Upload new file
        upload_result = cloudinary.uploader.upload(
            file,
            folder='tahfeel-documents',
            resource_type='auto',
            access_mode='authenticated',
            public_id=f"doc_{doc.name.replace(' ', '_')}_{datetime.now().timestamp()}"
        )
        
        doc.document_url = upload_result.get('secure_url')
        doc.cloudinary_public_id = upload_result.get('public_id')
        
        db.session.commit()
        flash(f'✓ File uploaded successfully for "{doc.name}".', 'success')
        return redirect(url_for('tahfeel_doc'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error uploading file: {str(e)}', 'error')
        print(f"Error uploading file: {e}")
        return redirect(url_for('tahfeel_doc'))

@app.route('/tahfeel-doc/add-bulk', methods=['POST'])
@login_required
def add_tahfeel_doc_bulk():
    if session['role'] != 'admin':
        flash('Only admin can add documents.', 'error')
        return redirect(url_for('tahfeel_doc'))
    
    try:
        # Count how many documents are being added (look for name_0, name_1, etc.)
        doc_indices = []
        i = 0
        while request.form.get(f'name_{i}'):
            doc_indices.append(i)
            i += 1
        
        if not doc_indices:
            flash('Please add at least one document.', 'error')
            return redirect(url_for('tahfeel_doc'))
        
        created_count = 0
        
        # Process each document
        for idx in doc_indices:
            name = request.form.get(f'name_{idx}', '').strip()
            doc_type = request.form.get(f'doc_type_{idx}', '').strip()
            owner = request.form.get(f'owner_{idx}', '').strip()
            issue_date_str = request.form.get(f'issue_date_{idx}', '')
            expiry_date_str = request.form.get(f'expiry_date_{idx}', '')
            authority = request.form.get(f'authority_{idx}', '').strip()
            
            # Validation
            if not name or not doc_type or not expiry_date_str or not owner:
                print(f"Skipping document {idx}: missing required fields")
                continue
            
            try:
                issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date() if issue_date_str else None
                expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
            except ValueError as ve:
                print(f"Date parsing error for doc {idx}: {ve}")
                continue
            
            # Handle file upload
            doc_url = None
            cloudinary_id = None
            file_key = f'file_{idx}'
            
            if file_key in request.files and request.files[file_key].filename:
                try:
                    file = request.files[file_key]
                    result = cloudinary.uploader.upload(
                        file,
                        folder='tahfeel-documents',
                        resource_type='auto',
                        access_mode='authenticated',
                        public_id=f"doc_{name.replace(' ', '_')}_{datetime.now().timestamp()}"
                    )
                    doc_url = result.get('secure_url')
                    cloudinary_id = result.get('public_id')
                except Exception as e:
                    print(f"Cloudinary upload error for doc {idx}: {e}")
            
            # Create document
            doc = CompanyDocument(
                name=name,
                doc_type=doc_type,
                owner=owner,
                issue_date=issue_date,
                expiry_date=expiry_date,
                authority=authority,
                document_url=doc_url,
                cloudinary_public_id=cloudinary_id,
                created_by=session.get('username', 'Unknown')
            )
            
            db.session.add(doc)
            created_count += 1
        
        if created_count > 0:
            db.session.commit()
            flash(f'✓ {created_count} document{"s" if created_count != 1 else ""} added successfully!', 'success')
        else:
            flash('No valid documents to add.', 'error')
        
        return redirect(url_for('tahfeel_doc'))
        
    except Exception as e:
        print(f"Error adding documents: {e}")
        flash('Error adding documents. Please try again.', 'error')
        return redirect(url_for('tahfeel_doc'))

@app.route('/tahfeel-doc/<int:doc_id>/delete', methods=['POST'])
@login_required
def delete_tahfeel_doc(doc_id):
    # Allow: Admin OR Saada (saadatahfeel@gmail.com)
    if session['role'] != 'admin' and session.get('user_email') != 'saadatahfeel@gmail.com':
        flash('Only admin can delete documents.')
        return redirect(url_for('tahfeel_doc'))
    
    doc = CompanyDocument.query.get_or_404(doc_id)
    
    try:
        # Delete from Cloudinary
        if doc.cloudinary_public_id:
            try:
                cloudinary.uploader.destroy(doc.cloudinary_public_id)
            except Exception as e:
                print(f"Cloudinary delete error: {e}")
        
        name = doc.name
        db.session.delete(doc)
        db.session.commit()
        flash(f'✓ Document "{name}" deleted.')
        return redirect(url_for('tahfeel_doc'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        print(f"Error deleting document: {e}")
        return redirect(url_for('tahfeel_doc'))

# ── Admin Panel Partner Routes (simpler pattern)
@app.route('/admin/partner/add', methods=['POST'])
@login_required
@admin_required
def admin_add_partner():
    name = request.form.get('name', '').strip()
    if name:
        existing = Partner.query.filter_by(name=name).first()
        if not existing:
            partner = Partner(name=name)
            db.session.add(partner)
            db.session.commit()
            flash(f'Partner "{name}" added.')
        else:
            flash('Partner already exists.', 'error')
    return redirect(url_for('admin_panel'))

@app.route('/admin/partner/<int:partner_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_partner(partner_id):
    partner = Partner.query.get_or_404(partner_id)
    new_name = request.form.get('name', '').strip()
    if new_name:
        existing = Partner.query.filter(Partner.name == new_name, Partner.id != partner_id).first()
        if not existing:
            partner.name = new_name
            db.session.commit()
            flash(f'Partner updated.')
        else:
            flash('Partner name already exists.', 'error')
    return redirect(url_for('admin_panel'))

@app.route('/admin/partner/<int:partner_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_partner(partner_id):
    partner = Partner.query.get_or_404(partner_id)
    jobs_with_partner = Job.query.filter_by(partner_name=partner.name).first()
    if not jobs_with_partner:
        db.session.delete(partner)
        db.session.commit()
        flash(f'Partner "{partner.name}" deleted.')
    else:
        flash('Cannot delete - has associated tasks.', 'error')
    return redirect(url_for('admin_panel'))

# ══════════════════════════════════════════════════════════════════════════════
# TEMPORARY ADMIN ROUTE - Fix April 30 Revenue Dates
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════

# ── App entry point (MUST stay at the very end so every @app.route above is
#    registered before the server starts; gunicorn imports the module, which
#    runs the else branch).
if __name__ == '__main__':
    init_db()
    # debug defaults OFF; enable locally with FLASK_DEBUG=true. Never enable in production
    # (the Werkzeug debugger allows remote code execution). Railway runs via gunicorn anyway.
    _debug = os.environ.get('FLASK_DEBUG', '').lower() == 'true'
    app.run(debug=_debug, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
else:
    init_db()
