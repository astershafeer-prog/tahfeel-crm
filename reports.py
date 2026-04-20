from flask import Blueprint, render_template, request, send_file, session, redirect, url_for, flash
from io import BytesIO
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)

HEADER_BG  = "133E87"
HEADER_FG  = "FFFFFF"
SUBHDR_BG  = "DCDCBF"
SUBHDR_FG  = "133E87"
ALT_ROW_BG = "EEF2F8"
TOTAL_BG   = "D0D8EC"
BORDER_CLR = "BBBBBB"

def _get_models():
    import app as _a
    return (_a.db, _a.Lead, _a.LeadUpdate, _a.Customer,
            _a.Job, _a.JobUpdate, _a.User, _a.Document)

def _border():
    s = Side(style='thin', color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, bg=HEADER_BG, fg=HEADER_FG):
    cell.font = Font(name='Arial', bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _border()

def _dat(cell, ridx, right=False, bold=False):
    cell.font = Font(name='Arial', size=10, bold=bold)
    cell.alignment = Alignment(horizontal='right' if right else 'left', vertical='center')
    cell.border = _border()
    if ridx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)

def _tot(cell, right=False):
    cell.font = Font(name='Arial', bold=True, size=10, color=HEADER_BG)
    cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
    cell.alignment = Alignment(horizontal='right' if right else 'center', vertical='center')
    cell.border = _border()

def _title_block(ws, title, df, dt, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="TAHFEEL BUSINESS SOLUTIONS")
    c.font = Font(name='Arial', bold=True, size=13, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=title)
    c.font = Font(name='Arial', bold=True, size=11, color=SUBHDR_FG)
    c.fill = PatternFill("solid", fgColor=SUBHDR_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1,
                value=f"Period: {df}  to  {dt}    |    Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}")
    c.font = Font(name='Arial', italic=True, size=9, color="666666")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14

def _headers(ws, cols):
    for i, h in enumerate(cols, 1):
        _hdr(ws.cell(row=4, column=i, value=h))
    ws.row_dimensions[4].height = 28

def _write_rows(ws, rows, start=5, num_cols=None):
    num_cols = num_cols or set()
    for ri, row in enumerate(rows):
        er = start + ri
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            _dat(cell, ri, right=(ci in num_cols))
            if ci in num_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[er].height = 15
    return start + len(rows)

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws):
    ws.freeze_panes = ws.cell(row=5, column=1)

def _filter(ws, ncols):
    ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}4"

def _respond(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)

def _dates(req):
    today = date.today()
    df = req.args.get('date_from', today.replace(day=1).strftime('%Y-%m-%d'))
    dt = req.args.get('date_to',   today.strftime('%Y-%m-%d'))
    try: df_d = datetime.strptime(df, '%Y-%m-%d')
    except: df_d = datetime(today.year, today.month, 1)
    try: dt_d = datetime.strptime(dt, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except: dt_d = datetime(today.year, today.month, today.day, 23, 59, 59)
    return df_d, dt_d, df, dt

def _guard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') not in ('admin', 'finance'):
        flash("Access denied.", "danger")
        return redirect(url_for('dashboard'))
    return None

# ── Reports index page
@reports_bp.route('/reports')
def reports_index():
    g = _guard()
    if g: return g
    today = date.today()
    defaults = {
        'date_from': today.replace(day=1).strftime('%Y-%m-%d'),
        'date_to':   today.strftime('%Y-%m-%d'),
    }
    return render_template('reports.html', defaults=defaults)


# ── 1. Lead Detail Report
@reports_bp.route('/reports/leads/export')
def export_lead_report():
    g = _guard()
    if g: return g
    db, Lead, LeadUpdate, Customer, Job, JobUpdate, User, Document = _get_models()
    df_d, dt_d, df, dt = _dates(request)

    users = {u.id: u.name for u in db.session.query(User).all()}
    leads = (db.session.query(Lead)
             .filter(Lead.created_at >= df_d, Lead.created_at <= dt_d)
             .order_by(Lead.created_at.desc()).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Detail"

    cols = ['#', 'Lead Name', 'Company', 'Phone', 'Service', 'Source',
            'Assigned To', 'Status', 'Created Date', 'Due Date',
            'Potential Value (AED)', 'Days Open', 'Interactions', 'Status History']
    _title_block(ws, "LEAD DETAIL REPORT", df, dt, len(cols))
    _headers(ws, cols)

    rows = []
    for i, lead in enumerate(leads, 1):
        updates = (db.session.query(LeadUpdate)
                   .filter(LeadUpdate.lead_id == lead.id)
                   .order_by(LeadUpdate.created_at.asc()).all())
        hist = []
        for u in updates:
            ts = u.created_at.strftime('%d/%m/%y %H:%M') if u.created_at else ''
            hist.append(f"[{ts}] {u.stage or ''} — {u.remark or ''} ({u.staff_name or ''})")
        days_open = (datetime.now() - lead.created_at).days if lead.created_at else ''
        rows.append([
            i, lead.name or '', lead.company or '', lead.phone or '',
            lead.service or '', lead.source or '',
            users.get(lead.assigned_to, '—'),
            lead.status or '',
            lead.created_at.strftime('%d/%m/%Y') if lead.created_at else '',
            lead.due_date.strftime('%d/%m/%Y') if lead.due_date else '',
            float(lead.potential_value or 0),
            days_open, len(updates),
            " | ".join(hist),
        ])

    nr = _write_rows(ws, rows, num_cols={11})
    ws.cell(row=nr, column=1, value='TOTAL'); _tot(ws.cell(row=nr, column=1))
    c = ws.cell(row=nr, column=11, value=f'=SUM(K5:K{nr-1})')
    _tot(c, right=True); c.number_format = '#,##0.00'

    _col_widths(ws, [4, 22, 22, 15, 18, 14, 18, 14, 13, 13, 18, 10, 12, 70])
    _freeze(ws); _filter(ws, len(cols))
    return _respond(wb, f"Lead_Detail_{df}_{dt}.xlsx")


# ── 2. Sales Report
@reports_bp.route('/reports/sales/export')
def export_sales_report():
    g = _guard()
    if g: return g
    db, Lead, LeadUpdate, Customer, Job, JobUpdate, User, Document = _get_models()
    df_d, dt_d, df, dt = _dates(request)

    users = {u.id: u.name for u in db.session.query(User).all()}
    jobs = (db.session.query(Job, Customer)
            .join(Customer, Job.customer_id == Customer.id)
            .filter(Job.created_at >= df_d, Job.created_at <= dt_d)
            .order_by(Job.created_at.desc()).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    cols = ['#', 'Customer Name', 'Company', 'Phone', 'Email',
            'Job Type', 'Source', 'Assigned To', 'Created By',
            'Status', 'Created Date', 'Due Date',
            'Invoiced (AED)', 'Received (AED)', 'Outstanding (AED)']
    _title_block(ws, "SALES REPORT", df, dt, len(cols))
    _headers(ws, cols)

    rows = []
    for i, (job, cust) in enumerate(jobs, 1):
        inv = float(job.amount_invoiced or 0)
        rec = float(job.amount_received or 0)
        rows.append([
            i, cust.name or '', cust.company or '', cust.phone or '', cust.email or '',
            job.job_type or '', cust.source or '',
            users.get(job.assigned_to, '—'),
            users.get(job.created_by, '—'),
            job.status or '',
            job.created_at.strftime('%d/%m/%Y') if job.created_at else '',
            job.due_date.strftime('%d/%m/%Y') if job.due_date else '',
            inv, rec, inv - rec,
        ])

    nr = _write_rows(ws, rows, num_cols={13, 14, 15})
    ws.cell(row=nr, column=1, value='TOTAL'); _tot(ws.cell(row=nr, column=1))
    for col in [13, 14, 15]:
        ltr = get_column_letter(col)
        c = ws.cell(row=nr, column=col, value=f'=SUM({ltr}5:{ltr}{nr-1})')
        _tot(c, right=True); c.number_format = '#,##0.00'

    _col_widths(ws, [4, 22, 22, 15, 22, 20, 14, 18, 18, 16, 13, 13, 18, 18, 18])
    _freeze(ws); _filter(ws, len(cols))
    return _respond(wb, f"Sales_Report_{df}_{dt}.xlsx")


# ── 3. Finance Report
@reports_bp.route('/reports/finance/export')
def export_finance_report():
    g = _guard()
    if g: return g
    db, Lead, LeadUpdate, Customer, Job, JobUpdate, User, Document = _get_models()
    df_d, dt_d, df, dt = _dates(request)

    users = {u.id: u.name for u in db.session.query(User).all()}
    jobs = (db.session.query(Job, Customer)
            .join(Customer, Job.customer_id == Customer.id)
            .filter(Job.created_at >= df_d, Job.created_at <= dt_d)
            .all())

    from collections import defaultdict
    summary = defaultdict(lambda: {'inv': 0.0, 'rec': 0.0, 'jobs': 0})
    for job, cust in jobs:
        month = job.created_at.strftime('%Y-%m') if job.created_at else 'Unknown'
        staff = users.get(job.assigned_to, 'Unassigned')
        summary[(month, staff)]['inv'] += float(job.amount_invoiced or 0)
        summary[(month, staff)]['rec'] += float(job.amount_received or 0)
        summary[(month, staff)]['jobs'] += 1

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Monthly Summary"
    cols1 = ['Month', 'Staff Member', 'Jobs', 'Invoiced (AED)', 'Received (AED)', 'Pending (AED)', 'Collection %']
    _title_block(ws1, "FINANCE REPORT — MONTHLY SUMMARY BY STAFF", df, dt, len(cols1))
    _headers(ws1, cols1)

    s_rows = []
    for (month, staff), d in sorted(summary.items()):
        inv, rec = d['inv'], d['rec']
        rate = f"{rec/inv*100:.1f}%" if inv > 0 else "—"
        s_rows.append([month, staff, d['jobs'], inv, rec, inv - rec, rate])

    nr1 = _write_rows(ws1, s_rows, num_cols={3, 4, 5, 6})
    ws1.cell(row=nr1, column=1, value='TOTAL'); _tot(ws1.cell(row=nr1, column=1))
    for col in [3, 4, 5, 6]:
        ltr = get_column_letter(col)
        c = ws1.cell(row=nr1, column=col, value=f'=SUM({ltr}5:{ltr}{nr1-1})')
        _tot(c, right=True)
        if col > 3: c.number_format = '#,##0.00'
    _col_widths(ws1, [12, 22, 10, 18, 18, 18, 14])
    _freeze(ws1); _filter(ws1, len(cols1))

    ws2 = wb.create_sheet("Job Detail")
    cols2 = ['#', 'Month', 'Customer', 'Company', 'Job Type', 'Assigned To',
             'Status', 'Finance Approved By', 'Created Date', 'Due Date',
             'Invoiced (AED)', 'Received (AED)', 'Pending (AED)']
    _title_block(ws2, "FINANCE REPORT — JOB DETAIL", df, dt, len(cols2))
    _headers(ws2, cols2)

    d_rows = []
    for i, (job, cust) in enumerate(jobs, 1):
        inv = float(job.amount_invoiced or 0)
        rec = float(job.amount_received or 0)
        d_rows.append([
            i,
            job.created_at.strftime('%Y-%m') if job.created_at else '',
            cust.name or '', cust.company or '',
            job.job_type or '',
            users.get(job.assigned_to, '—'),
            job.status or '',
            users.get(job.finance_approved_by, '—') if job.finance_approved_by else '—',
            job.created_at.strftime('%d/%m/%Y') if job.created_at else '',
            job.due_date.strftime('%d/%m/%Y') if job.due_date else '',
            inv, rec, inv - rec,
        ])

    nr2 = _write_rows(ws2, d_rows, num_cols={11, 12, 13})
    ws2.cell(row=nr2, column=1, value='TOTAL'); _tot(ws2.cell(row=nr2, column=1))
    for col in [11, 12, 13]:
        ltr = get_column_letter(col)
        c = ws2.cell(row=nr2, column=col, value=f'=SUM({ltr}5:{ltr}{nr2-1})')
        _tot(c, right=True); c.number_format = '#,##0.00'
    _col_widths(ws2, [4, 10, 22, 22, 20, 18, 18, 18, 13, 13, 18, 18, 18])
    _freeze(ws2); _filter(ws2, len(cols2))

    return _respond(wb, f"Finance_Report_{df}_{dt}.xlsx")


# ── 4. Task Report
@reports_bp.route('/reports/tasks/export')
def export_task_report():
    g = _guard()
    if g: return g
    db, Lead, LeadUpdate, Customer, Job, JobUpdate, User, Document = _get_models()
    df_d, dt_d, df, dt = _dates(request)

    users = {u.id: u.name for u in db.session.query(User).all()}
    jobs = (db.session.query(Job, Customer)
            .join(Customer, Job.customer_id == Customer.id)
            .filter(Job.created_at >= df_d, Job.created_at <= dt_d)
            .order_by(Job.created_at.desc()).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task Report"

    cols = ['#', 'Customer', 'Company', 'Job Type', 'Assigned To',
            'Status', 'Priority', 'Created Date', 'Due Date', 'Completed At',
            'Days to Complete', 'Interactions', 'Status History']
    _title_block(ws, "TASK REPORT", df, dt, len(cols))
    _headers(ws, cols)

    rows = []
    for i, (job, cust) in enumerate(jobs, 1):
        updates = (db.session.query(JobUpdate)
                   .filter(JobUpdate.job_id == job.id)
                   .order_by(JobUpdate.created_at.asc()).all())
        hist = []
        for u in updates:
            ts = u.created_at.strftime('%d/%m/%y %H:%M') if u.created_at else ''
            hist.append(f"[{ts}] {u.status or ''} — {u.remark or ''} ({u.staff_name or ''})")
        days_to_complete = ''
        completed_at_str = job.completed_at.strftime('%d/%m/%Y') if job.completed_at else ''
        if job.completed_at and job.created_at:
            days_to_complete = (job.completed_at - job.created_at).days
        rows.append([
            i, cust.name or '', cust.company or '',
            job.job_type or '',
            users.get(job.assigned_to, '—'),
            job.status or '', job.priority or '',
            job.created_at.strftime('%d/%m/%Y') if job.created_at else '',
            job.due_date.strftime('%d/%m/%Y') if job.due_date else '',
            completed_at_str, days_to_complete, len(updates),
            " | ".join(hist),
        ])

    _write_rows(ws, rows, num_cols={11, 12})
    _col_widths(ws, [4, 22, 22, 22, 18, 18, 10, 13, 13, 13, 14, 12, 70])
    _freeze(ws); _filter(ws, len(cols))
    return _respond(wb, f"Task_Report_{df}_{dt}.xlsx")


# ── 5. Document Expiry Report
@reports_bp.route('/reports/documents/export')
def export_document_report():
    g = _guard()
    if g: return g
    db, Lead, LeadUpdate, Customer, Job, JobUpdate, User, Document = _get_models()
    df_d, dt_d, df, dt = _dates(request)

    docs = (db.session.query(Document)
            .filter(Document.expiry_date >= df_d, Document.expiry_date <= dt_d)
            .order_by(Document.expiry_date.asc()).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Document Expiry"

    cols = ['#', 'Owner Name', 'Belongs To', 'Company', 'Phone',
            'Document Type', 'Expiry Date', 'Days Remaining', 'Status', 'Notes', 'Added By']
    _title_block(ws, "DOCUMENT EXPIRY REPORT", df, dt, len(cols))
    _headers(ws, cols)

    today = date.today()
    rows = []
    for i, doc in enumerate(docs, 1):
        days_rem = ''
        status = ''
        if doc.expiry_date:
            exp = doc.expiry_date.date() if hasattr(doc.expiry_date, 'date') else doc.expiry_date
            delta = (exp - today).days
            days_rem = delta
            if delta < 0:      status = 'EXPIRED'
            elif delta <= 30:  status = 'CRITICAL'
            elif delta <= 90:  status = 'EXPIRING SOON'
            else:              status = 'VALID'
        cust = doc.customer
        rows.append([
            i, doc.owner_name or '', doc.belongs_to or '',
            cust.company if cust else '', cust.phone if cust else '',
            doc.doc_type or '',
            doc.expiry_date.strftime('%d/%m/%Y') if doc.expiry_date else '',
            days_rem, status, doc.notes or '', doc.added_by or '',
        ])

    nr = _write_rows(ws, rows, num_cols={8})
    status_colors = {
        'EXPIRED':       ('FF3333', 'FFFFFF'),
        'CRITICAL':      ('FF8800', 'FFFFFF'),
        'EXPIRING SOON': ('FFD700', '000000'),
        'VALID':         ('22AA55', 'FFFFFF'),
    }
    for r in range(5, nr):
        sc = ws.cell(row=r, column=9)
        if sc.value in status_colors:
            bg, fg = status_colors[sc.value]
            sc.fill = PatternFill("solid", fgColor=bg)
            sc.font = Font(name='Arial', bold=True, size=10, color=fg)
            sc.alignment = Alignment(horizontal='center', vertical='center')

    _col_widths(ws, [4, 22, 14, 22, 15, 20, 13, 14, 14, 22, 16])
    _freeze(ws); _filter(ws, len(cols))
    return _respond(wb, f"Document_Expiry_{df}_{dt}.xlsx")


# ── 6. Staff Performance Report
@reports_bp.route('/reports/staff/export')
def export_staff_report():
    g = _guard()
    if g: return g
    db, Lead, LeadUpdate, Customer, Job, JobUpdate, User, Document = _get_models()
    df_d, dt_d, df, dt = _dates(request)

    users = db.session.query(User).filter(User.active == True).all()

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Staff Summary"

    cols1 = ['Staff Member', 'Role',
             'Leads Assigned', 'Leads Won', 'Leads Lost', 'Conversion %',
             'Jobs Assigned', 'Jobs Completed', 'Completion %',
             'Invoiced (AED)', 'Received (AED)', 'Outstanding (AED)']
    _title_block(ws1, "STAFF PERFORMANCE REPORT", df, dt, len(cols1))
    _headers(ws1, cols1)

    won_s  = {'Won', 'Converted', 'Closed-Won'}
    lost_s = {'Lost', 'Rejected', 'Closed-Lost'}
    done_s = {'Completed', 'Done', 'Closed', 'Delivered'}

    rows = []
    for user in users:
        leads = db.session.query(Lead).filter(
            Lead.assigned_to == user.id,
            Lead.created_at >= df_d, Lead.created_at <= dt_d).all()
        lt  = len(leads)
        lw  = sum(1 for l in leads if l.status in won_s)
        ll  = sum(1 for l in leads if l.status in lost_s)
        lcr = f"{lw/lt*100:.1f}%" if lt > 0 else "—"

        # Tasks assigned to this person (for completion tracking)
        jobs = db.session.query(Job).filter(
            Job.assigned_to == user.id,
            Job.created_at >= df_d, Job.created_at <= dt_d).all()
        jt  = len(jobs)
        jd  = sum(1 for j in jobs if j.status in done_s)
        jcr = f"{jd/jt*100:.1f}%" if jt > 0 else "—"

        # Sales value: credited to job CREATOR (primary representative), not assignee
        sales_jobs = db.session.query(Job).filter(
            Job.created_by == user.id,
            Job.created_at >= df_d, Job.created_at <= dt_d).all()
        inv = sum(float(j.amount_invoiced or 0) for j in sales_jobs)
        rec = sum(float(j.amount_received or 0) for j in sales_jobs)
        rows.append([user.name, user.role, lt, lw, ll, lcr, jt, jd, jcr, inv, rec, inv - rec])

    nr1 = _write_rows(ws1, rows, num_cols={3, 4, 5, 7, 8, 10, 11, 12})
    ws1.cell(row=nr1, column=1, value='TOTAL'); _tot(ws1.cell(row=nr1, column=1))
    for col in [3, 4, 5, 7, 8, 10, 11, 12]:
        ltr = get_column_letter(col)
        c = ws1.cell(row=nr1, column=col, value=f'=SUM({ltr}5:{ltr}{nr1-1})')
        _tot(c, right=True)
        if col >= 10: c.number_format = '#,##0.00'
    _col_widths(ws1, [22, 12, 14, 12, 12, 14, 14, 14, 14, 18, 18, 18])
    _freeze(ws1); _filter(ws1, len(cols1))

    ws2 = wb.create_sheet("Leads by Staff")
    cols2 = ['Staff', 'Lead Name', 'Company', 'Service', 'Source', 'Status',
             'Created', 'Due Date', 'Potential Value (AED)', 'Interactions']
    _title_block(ws2, "LEAD DETAIL BY STAFF", df, dt, len(cols2))
    _headers(ws2, cols2)

    users_map = {u.id: u.name for u in users}
    all_leads = (db.session.query(Lead)
                 .filter(Lead.created_at >= df_d, Lead.created_at <= dt_d)
                 .order_by(Lead.assigned_to, Lead.created_at.desc()).all())

    l_rows = []
    for l in all_leads:
        cnt = db.session.query(LeadUpdate).filter(LeadUpdate.lead_id == l.id).count()
        l_rows.append([
            users_map.get(l.assigned_to, '—'),
            l.name or '', l.company or '', l.service or '', l.source or '',
            l.status or '',
            l.created_at.strftime('%d/%m/%Y') if l.created_at else '',
            l.due_date.strftime('%d/%m/%Y') if l.due_date else '',
            float(l.potential_value or 0), cnt,
        ])
    _write_rows(ws2, l_rows, num_cols={9, 10})
    _col_widths(ws2, [20, 22, 22, 18, 14, 14, 13, 13, 18, 12])
    _freeze(ws2); _filter(ws2, len(cols2))

    return _respond(wb, f"Staff_Performance_{df}_{dt}.xlsx")
